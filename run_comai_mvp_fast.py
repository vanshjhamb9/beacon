"""Fast COMAI MVP Pipeline - Process leads without slow HTTP requests."""

from __future__ import annotations

import json
import logging
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / "packages"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("comai_mvp_fast")


# 35 Indian D2C seed brands with pre-populated data
SEED_BRANDS = [
    # Beauty & Skincare
    ("mamaearth.in", "Mamaearth", "beauty", "Mumbai", "Maharashtra", "Indian D2C beauty brand offering natural personal care products.", 500, "shopify", True, False, True, False, "Varun Dhamija", "varun@mamaearth.in", "+91 9876543210"),
    ("beardo.in", "Beardo", "grooming", "Mumbai", "Maharashtra", "Indian men's grooming brand selling beard care, hair care, and skincare products.", 200, "shopify", False, False, True, False, "Aditya Sharma", "aditya@beardo.in", "+91 9876543211"),
    ("mcaffeine.com", "mCaffeine", "skincare", "Mumbai", "Maharashtra", "India's first caffeinated personal care brand for millennials.", 300, "shopify", False, False, True, False, "Tarun Sharma", "tarun@mcaffeine.com", "+91 9876543212"),
    ("sugarcosmetics.com", "Sugar Cosmetics", "cosmetics", "Mumbai", "Maharashtra", "Premium Indian makeup brand with 10,000+ retail touchpoints.", 500, "shopify", True, True, True, False, "Vineeta Singh", "vineeta@sugarcosmetics.com", "+91 9876543213"),
    ("wowskinscience.com", "WOW Skin Science", "skincare", "Bangalore", "Karnataka", "Indian D2C personal care brand with natural and active ingredients.", 300, "shopify", False, False, True, False, "Manish Chowdhary", "manish@wowskinscience.com", "+91 9876543214"),
    ("plumgoodness.com", "Plum Goodness", "beauty", "Mumbai", "Maharashtra", "100% vegan and cruelty-free Indian beauty brand.", 250, "shopify", False, False, True, False, "Shankar Prasad", "shankar@plumgoodness.com", "+91 9876543215"),
    ("bombayshavingcompany.com", "Bombay Shaving Company", "grooming", "New Delhi", "Delhi", "Indian men's grooming D2C brand.", 200, "shopify", False, False, True, False, "Shantanu Deshpande", "shantanu@bombayshaving.com", "+91 9876543216"),
    ("dermaco.in", "Derma Co", "skincare", "New Delhi", "Delhi", "Indian derma-cosmetics brand by Mamaearth parent.", 250, "shopify", False, False, True, False, "Gazal Alagh", "gazal@dermaco.in", "+91 9876543217"),
    ("juicychemistry.com", "Juicy Chemistry", "organic beauty", "Coimbatore", "Tamil Nadu", "Indian organic and natural skincare brand.", 150, "shopify", False, False, True, False, "Pritesh Asher", "pritesh@juicychemistry.com", "+91 9876543218"),
    ("khadinatural.com", "Khadi Natural", "natural products", "Ahmedabad", "Gujarat", "Indian natural personal care brand.", 200, "shopify", True, True, True, False, "Pradeep Ghodilya", "pradeep@khadinatural.com", "+91 9876543219"),

    # Fashion
    ("thesouledstore.com", "The Souled Store", "fashion", "Mumbai", "Maharashtra", "India's largest pop-culture fashion brand.", 400, "shopify", False, False, True, False, "Vedang Patel", "vedang@thesouledstore.com", "+91 9876543220"),
    ("bewakoof.com", "Bewakoof", "fashion", "Mumbai", "Maharashtra", "India's leading D2C fashion brand for youth.", 500, "shopify", False, False, True, False, "Prabhkiran Singh", "prabhkiran@bewakoof.com", "+91 9876543221"),
    ("snitch.co.in", "Snitch", "fashion", "Bangalore", "Karnataka", "Fast-growing Indian menswear D2C brand.", 300, "shopify", False, False, True, False, "Siddharth Daga", "siddharth@snitch.co.in", "+91 9876543222"),
    ("vanheusen.com", "Van Heusen", "fashion", "Mumbai", "Maharashtra", "Premium Indian fashion brand by Aditya Birla.", 400, "shopify", False, False, True, False, "Ashish Dixit", "ashish@vanheusen.com", "+91 9876543223"),
    ("parkavenue.in", "Park Avenue", "fashion", "Mumbai", "Maharashtra", "Indian premium menswear brand.", 300, "shopify", False, False, True, False, "Rajiv Sabharwal", "rajiv@parkavenue.in", "+91 9876543224"),
    ("wooplr.com", "Wooplr", "fashion", "Bangalore", "Karnataka", "Indian fashion discovery platform.", 200, "shopify", False, False, True, False, "Soumen Sarkar", "soumen@wooplr.com", "+91 9876543225"),
    ("styched.in", "Styched", "fashion", "Bangalore", "Karnataka", "Indian online fashion brand for youth.", 150, "shopify", False, False, True, False, "Aditya Singh", "aditya@styched.in", "+91 9876543226"),

    # Electronics
    ("ambraneindia.com", "Ambrane", "electronics", "New Delhi", "Delhi", "Indian consumer electronics brand for power banks and accessories.", 150, "shopify", False, False, True, False, "Ashok Rajpal", "ashok@ambraneindia.com", "+91 9876543227"),
    ("pTron.com", "pTron", "electronics", "Hyderabad", "Telangana", "Indian budget audio and mobile accessories brand.", 200, "shopify", False, False, True, False, "Ameen Khwaja", "ameen@ptron.com", "+91 9876543228"),
    ("boat-lifestyle.com", "boAt", "electronics", "New Delhi", "Delhi", "India's #1 audio and wearables brand with 8M+ products sold.", 500, "shopify", True, True, True, False, "Aman Gupta", "aman@boat-lifestyle.com", "+91 9876543229"),
    ("noise.tech", "Noise", "electronics", "Gurugram", "Haryana", "India's leading connected lifestyle brand for smartwatches and audio.", 400, "shopify", False, False, True, False, "Gaurav Khatri", "gaurav@noise.tech", "+91 9876543230"),
    ("fireboltt.com", "Fire-Boltt", "electronics", "New Delhi", "Delhi", "India's fastest growing smartwatch and audio brand.", 350, "shopify", False, False, True, False, "Nitin Khatri", "nitin@fireboltt.com", "+91 9876543231"),

    # Jewellery
    ("caratlane.com", "CaratLane", "jewelry", "Chennai", "Tamil Nadu", "India's largest omnichannel jewelry brand.", 500, "shopify", True, True, True, False, "Mithun Sacheti", "mithun@caratlane.com", "+91 9876543232"),
    ("bluestone.com", "BlueStone", "jewelry", "Bangalore", "Karnataka", "Leading online jewelry brand in India.", 400, "shopify", False, False, True, False, "Gaurav Singh Kushwaha", "gaurav@bluestone.com", "+91 9876543233"),

    # Home Decor
    ("addresshome.com", "Address Home", "home decor", "New Delhi", "Delhi", "Indian premium home decor and lifestyle brand.", 150, "shopify", False, False, True, False, "Raj Shetya", "raj@addresshome.com", "+91 9876543234"),
    ("jaypore.com", "Jaypore", "lifestyle", "New Delhi", "Delhi", "Indian online lifestyle and home decor brand.", 200, "shopify", False, False, True, False, "Shilpa Gupta", "shilpa@jaypore.com", "+91 9876543235"),

    # Kids
    ("hopscotch.in", "Hopscotch", "kids", "Mumbai", "Maharashtra", "India's leading online store for kids' fashion.", 250, "shopify", False, False, True, False, "Rahul Anand", "rahul@hopscotch.in", "+91 9876543236"),
    ("firstcry.com", "FirstCry", "kids", "Pune", "Maharashtra", "Asia's largest online store for kids and baby products.", 1000, "shopify", True, True, True, False, "Supam Maheshwari", "supam@firstcry.com", "+91 9876543237"),

    # Lifestyle
    ("nicobar.com", "Nicobar", "lifestyle", "Mumbai", "Maharashtra", "Modern Indian lifestyle brand for travel and home.", 150, "shopify", False, False, True, False, "Simran Lal", "simran@nicobar.com", "+91 9876543238"),
    ("okhai.org", "Okhai", "handicrafts", "Ahmedabad", "Gujarat", "Handcrafted lifestyle brand empowering rural artisans.", 100, "shopify", False, False, True, False, "Nimish Shah", "nimish@okhai.org", "+91 9876543239"),

    # Pets
    ("headsupfortails.com", "Heads Up For Tails", "pets", "New Delhi", "Delhi", "Indian premium pet care brand.", 200, "shopify", False, False, True, False, "Rashi Narang", "rashi@headsupfortails.com", "+91 9876543240"),

    # Food & Beverage
    ("teabox.com", "Teabox", "tea", "Kolkata", "West Bengal", "Indian premium tea D2C brand.", 100, "shopify", False, False, True, False, "Kaus Munot", "kaus@teabox.com", "+91 9876543241"),
    ("rawpressery.com", "Raw Pressery", "beverages", "Mumbai", "Maharashtra", "India's largest cold-pressed juice brand.", 200, "shopify", False, False, True, False, "Saurabh Kahatra", "saurabh@rawpressery.com", "+91 9876543242"),
    ("box8.in", "BOX8", "food", "Mumbai", "Maharashtra", "Indian all-in-one meal delivery brand.", 150, "shopify", False, False, True, False, "Anil Gehlot", "anil@box8.in", "+91 9876543243"),
]


def run_fast_pipeline() -> dict:
    """Run fast pipeline without HTTP requests."""
    from packages.qualification_engine.icp_loader import load_icp
    from packages.qualification_engine.scorer import QualificationScorer
    from packages.qualification_engine.sales_intelligence import SalesIntelligenceGenerator
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # Load COMAI ICP
    comai_icp = load_icp("comai")
    scorer = QualificationScorer(comai_icp)
    sales_intel_gen = SalesIntelligenceGenerator(comai_icp)

    start_time = time.time()
    all_leads = []
    qualified_leads = []
    sales_ready_leads = []
    nurture_leads = []
    rejected_leads = []
    seen_domains = set()

    # Stats
    stats = {
        "discovered": 0,
        "valid": 0,
        "excluded": 0,
        "icp_matched": 0,
        "qualified": 0,
        "sales_ready": 0,
        "nurture": 0,
        "rejected": 0,
        "scores": [],
    }

    logger.info("=" * 70)
    logger.info("COMAI MVP PIPELINE (FAST) - Processing seed brands")
    logger.info("=" * 70)

    for i, (domain, name, category, city, state, desc, products, platform, has_chatbot, has_whatsapp, has_crm, has_instagram, founder, email, phone) in enumerate(SEED_BRANDS):
        stats["discovered"] += 1

        if domain in seen_domains:
            continue
        seen_domains.add(domain)

        # Create lead with pre-populated data
        from packages.ecommerce_leads.models import RawEcommerceLead, EnrichedEcommerceLead

        raw = RawEcommerceLead(
            company_name=name,
            website=f"https://{domain}",
            domain=domain,
            platform=platform,
            industry=category,
            category=category,
            country="India",
            city=city,
            state=state,
            description=desc,
            product_count=products,
            source="mvp_seed_data",
        )

        lead = EnrichedEcommerceLead(raw=raw)
        lead.shopify_detected = platform == "shopify"
        lead.chatbot_detected = has_chatbot
        lead.whatsapp_detected = has_whatsapp
        lead.crm_detected = has_crm
        lead.founder_name = founder
        lead.email = email
        lead.phone = phone

        # Qualification scoring
        qual_result = scorer.score(lead)

        # Generate sales intelligence
        sales_intel = sales_intel_gen.generate(lead)

        # Store results
        lead.raw.metadata["qualification"] = {
            "total_score": qual_result.total_score,
            "grade": qual_result.grade,
            "business_unit": qual_result.business_unit,
            "icp_fit_score": qual_result.icp_fit_score,
            "business_size_score": qual_result.business_size_score,
            "growth_signals_score": qual_result.growth_signals_score,
            "technology_fit_score": qual_result.technology_fit_score,
            "support_pain_score": qual_result.support_pain_score,
            "decision_maker_score": qual_result.decision_maker_score,
            "buying_intent_score": qual_result.buying_intent_score,
            "evidence": qual_result.evidence,
            "disqualification_reasons": qual_result.disqualification_reasons,
        }
        lead.raw.metadata["sales_intelligence"] = {
            "company_summary": sales_intel.company_summary,
            "business_challenges": sales_intel.business_challenges,
            "technology_stack": sales_intel.technology_stack,
            "decision_makers": sales_intel.decision_makers,
            "buying_signals": sales_intel.buying_signals,
            "recommended_service": sales_intel.recommended_service,
            "recommended_pitch": sales_intel.recommended_pitch,
            "likely_objections": sales_intel.likely_objections,
            "meeting_strategy": sales_intel.meeting_strategy,
            "probability_of_closing": sales_intel.probability_of_closing,
        }

        all_leads.append(lead)
        stats["scores"].append(qual_result.total_score)

        # Categorize
        if qual_result.grade == "SALES_READY":
            sales_ready_leads.append(lead)
            stats["sales_ready"] += 1
            stats["qualified"] += 1
            logger.info("[✓] %s - SALES_READY (Score: %.1f)", name, qual_result.total_score)
        elif qual_result.grade == "QUALIFIED":
            qualified_leads.append(lead)
            stats["qualified"] += 1
            logger.info("[✓] %s - QUALIFIED (Score: %.1f)", name, qual_result.total_score)
        elif qual_result.grade == "NURTURE":
            nurture_leads.append(lead)
            stats["nurture"] += 1
            logger.info("[~] %s - NURTURE (Score: %.1f)", name, qual_result.total_score)
        else:
            rejected_leads.append(lead)
            stats["rejected"] += 1
            logger.info("[✗] %s - REJECT (Score: %.1f)", name, qual_result.total_score)

        # Check if valid
        if not qual_result.disqualification_reasons:
            stats["valid"] += 1
        else:
            stats["excluded"] += 1

        # Check if ICP matched
        if qual_result.total_score > 50:
            stats["icp_matched"] += 1

    elapsed = time.time() - start_time

    # Calculate averages
    avg_score = sum(stats["scores"]) / len(stats["scores"]) if stats["scores"] else 0
    score_distribution = {
        "85-100 (SALES_READY)": len([s for s in stats["scores"] if s >= 85]),
        "70-84 (QUALIFIED)": len([s for s in stats["scores"] if 70 <= s < 85]),
        "50-69 (NURTURE)": len([s for s in stats["scores"] if 50 <= s < 70]),
        "0-49 (REJECT)": len([s for s in stats["scores"] if s < 50]),
    }

    logger.info("=" * 70)
    logger.info("MVP Pipeline completed in %.1f seconds", elapsed)
    logger.info("=" * 70)

    # Print summary
    logger.info("")
    logger.info("RESULTS SUMMARY:")
    logger.info("  Total discovered: %d", stats["discovered"])
    logger.info("  Valid (no exclusions): %d", stats["valid"])
    logger.info("  Excluded: %d", stats["excluded"])
    logger.info("  ICP matched (score > 50): %d", stats["icp_matched"])
    logger.info("  Sales Ready (score >= 85): %d", stats["sales_ready"])
    logger.info("  Qualified (score 70-84): %d", len(qualified_leads))
    logger.info("  Nurture (score 50-69): %d", stats["nurture"])
    logger.info("  Rejected (score < 50): %d", stats["rejected"])
    logger.info("  Average score: %.1f", avg_score)
    logger.info("")
    logger.info("SCORE DISTRIBUTION:")
    for range_name, count in score_distribution.items():
        logger.info("  %s: %d", range_name, count)

    # Export to Excel
    output_path = Path(__file__).parent / "exports" / "comai_qualified_leads.xlsx"
    output_path.parent.mkdir(exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["COMAI MVP PIPELINE RESULTS"])
    ws_summary.append([])
    ws_summary.append(["Generated", time.strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Total Discovered", stats["discovered"]])
    ws_summary.append(["Valid", stats["valid"]])
    ws_summary.append(["Excluded", stats["excluded"]])
    ws_summary.append(["ICP Matched", stats["icp_matched"]])
    ws_summary.append(["Sales Ready", stats["sales_ready"]])
    ws_summary.append(["Qualified", len(qualified_leads)])
    ws_summary.append(["Nurture", stats["nurture"]])
    ws_summary.append(["Rejected", stats["rejected"]])
    ws_summary.append(["Average Score", f"{avg_score:.1f}"])
    ws_summary.append([])
    ws_summary.append(["SCORE DISTRIBUTION"])
    for range_name, count in score_distribution.items():
        ws_summary.append([range_name, count])

    # Sales Ready Leads sheet
    ws_leads = wb.create_sheet("Sales Ready Leads")
    headers = [
        "Company", "Website", "Platform", "Industry", "City",
        "Email", "Phone", "Founder",
        "Total Score", "ICP Fit", "Business Size", "Growth", "Tech Fit",
        "Support Pain", "Decision Maker", "Buying Intent",
        "Recommended Service", "Probability", "Pitch",
    ]
    ws_leads.append(headers)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for cell in ws_leads[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for lead in sales_ready_leads:
        qual = lead.raw.metadata.get("qualification", {})
        intel = lead.raw.metadata.get("sales_intelligence", {})
        ws_leads.append([
            lead.raw.company_name,
            lead.raw.website,
            lead.raw.platform,
            lead.raw.industry,
            lead.raw.city,
            lead.email or "",
            lead.phone or "",
            lead.founder_name or "",
            qual.get("total_score", 0),
            qual.get("icp_fit_score", 0),
            qual.get("business_size_score", 0),
            qual.get("growth_signals_score", 0),
            qual.get("technology_fit_score", 0),
            qual.get("support_pain_score", 0),
            qual.get("decision_maker_score", 0),
            qual.get("buying_intent_score", 0),
            intel.get("recommended_service", ""),
            intel.get("probability_of_closing", 0),
            intel.get("recommended_pitch", "")[:200],
        ])

    # Qualified Leads sheet
    ws_qualified = wb.create_sheet("Qualified Leads")
    ws_qualified.append(headers)
    for cell in ws_qualified[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for lead in qualified_leads:
        qual = lead.raw.metadata.get("qualification", {})
        intel = lead.raw.metadata.get("sales_intelligence", {})
        ws_qualified.append([
            lead.raw.company_name,
            lead.raw.website,
            lead.raw.platform,
            lead.raw.industry,
            lead.raw.city,
            lead.email or "",
            lead.phone or "",
            lead.founder_name or "",
            qual.get("total_score", 0),
            qual.get("icp_fit_score", 0),
            qual.get("business_size_score", 0),
            qual.get("growth_signals_score", 0),
            qual.get("technology_fit_score", 0),
            qual.get("support_pain_score", 0),
            qual.get("decision_maker_score", 0),
            qual.get("buying_intent_score", 0),
            intel.get("recommended_service", ""),
            intel.get("probability_of_closing", 0),
            intel.get("recommended_pitch", "")[:200],
        ])

    # Auto-width
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(str(output_path))
    logger.info("Exported to: %s", output_path)

    # Save JSON report
    json_report = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "elapsed_seconds": round(elapsed, 1),
        "stats": {
            "discovered": stats["discovered"],
            "valid": stats["valid"],
            "excluded": stats["excluded"],
            "icp_matched": stats["icp_matched"],
            "sales_ready": stats["sales_ready"],
            "qualified": len(qualified_leads),
            "nurture": stats["nurture"],
            "rejected": stats["rejected"],
            "average_score": round(avg_score, 1),
            "score_distribution": score_distribution,
        },
        "sales_ready_leads": [],
        "qualified_leads": [],
    }

    for lead in sales_ready_leads:
        qual = lead.raw.metadata.get("qualification", {})
        intel = lead.raw.metadata.get("sales_intelligence", {})
        json_report["sales_ready_leads"].append({
            "company": lead.raw.company_name,
            "website": lead.raw.website,
            "platform": lead.raw.platform,
            "industry": lead.raw.industry,
            "city": lead.raw.city,
            "email": lead.email,
            "phone": lead.phone,
            "founder": lead.founder_name,
            "scores": {
                "total": qual.get("total_score", 0),
                "icp_fit": qual.get("icp_fit_score", 0),
                "business_size": qual.get("business_size_score", 0),
                "growth_signals": qual.get("growth_signals_score", 0),
                "technology_fit": qual.get("technology_fit_score", 0),
                "support_pain": qual.get("support_pain_score", 0),
                "decision_maker": qual.get("decision_maker_score", 0),
                "buying_intent": qual.get("buying_intent_score", 0),
            },
            "sales_intelligence": intel,
        })

    json_path = Path(__file__).parent / "exports" / "comai_mvp_results.json"
    with open(json_path, "w") as f:
        json.dump(json_report, f, indent=2, default=str)
    logger.info("JSON report saved to: %s", json_path)

    return {
        "stats": stats,
        "sales_ready_leads": sales_ready_leads,
        "qualified_leads": qualified_leads,
        "all_leads": all_leads,
        "elapsed": elapsed,
    }


if __name__ == "__main__":
    result = run_fast_pipeline()
    print(f"\nDone! Processed {result['stats']['discovered']} brands.")
    print(f"Sales Ready: {result['stats']['sales_ready']}")
    print(f"Qualified: {result['stats']['qualified']}")
    print(f"Average Score: {result['stats']['average_score']:.1f}")
