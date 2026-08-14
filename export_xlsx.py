"""Export sales queue and outreach drafts to XLSX format."""

from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PROJECT_ROOT = Path(__file__).parent
EXPORTS_DIR = PROJECT_ROOT / "exports"


def load_json(filename: str) -> list[dict[str, Any]]:
    """Load JSON file."""
    filepath = EXPORTS_DIR / filename
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def export_sales_queue_xlsx() -> None:
    """Export sales queue to XLSX."""
    if not HAS_OPENPYXL:
        print("openpyxl not available. Skipping XLSX export.")
        return

    companies = load_json("final_sales_queue.json")

    wb = Workbook()

    # Sheet 1: Sales Queue
    ws = wb.active
    ws.title = "Sales Queue"

    # Headers
    headers = [
        "Company", "Requirement", "Intent", "Intent Score",
        "Outsourcing Fit", "Decision Maker", "DM Role",
        "DM Confidence", "Email", "Email Status",
        "LinkedIn", "Service Match", "Why Now",
        "Recommended Channel", "Pitch Angle", "Queue"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row, company in enumerate(companies, 2):
        data = [
            company.get("company"),
            company.get("requirement"),
            company.get("intent"),
            company.get("intent_score"),
            company.get("outsourcing_fit"),
            company.get("decision_maker"),
            company.get("decision_maker_role"),
            company.get("decision_maker_confidence"),
            company.get("email"),
            company.get("email_status"),
            company.get("linkedin"),
            company.get("service_match"),
            company.get("why_now"),
            company.get("recommended_channel"),
            company.get("pitch_angle"),
            company.get("queue"),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Color code by queue
        queue = company.get("queue")
        if queue == "OUTREACH_READY":
            fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        elif queue == "NEEDS_RESEARCH":
            fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    # Save
    output_file = EXPORTS_DIR / "final_sales_queue.xlsx"
    wb.save(output_file)
    print(f"Saved sales queue to {output_file}")


def export_outreach_drafts_xlsx() -> None:
    """Export outreach drafts to XLSX."""
    if not HAS_OPENPYXL:
        print("openpyxl not available. Skipping XLSX export.")
        return

    drafts = load_json("final_outreach_drafts.json")

    wb = Workbook()

    # Sheet 1: Email Drafts
    ws_email = wb.active
    ws_email.title = "Email Drafts"

    email_headers = ["Company", "To", "To Status", "Subject", "Body", "Generated At"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(email_headers, 1):
        cell = ws_email.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    email_drafts = [d for d in drafts if d["channel"] == "email"]
    for row, draft in enumerate(email_drafts, 2):
        data = [
            draft.get("company"),
            draft.get("to"),
            draft.get("to_status"),
            draft.get("subject"),
            draft.get("body"),
            draft.get("generated_at"),
        ]

        for col, value in enumerate(data, 1):
            cell = ws_email.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 2: LinkedIn Drafts
    ws_linkedin = wb.create_sheet("LinkedIn Drafts")

    linkedin_headers = ["Company", "To", "To LinkedIn", "Message", "Connection Note", "Generated At"]

    for col, header in enumerate(linkedin_headers, 1):
        cell = ws_linkedin.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    linkedin_drafts = [d for d in drafts if d["channel"] == "linkedin"]
    for row, draft in enumerate(linkedin_drafts, 2):
        data = [
            draft.get("company"),
            draft.get("to"),
            draft.get("to_linkedin"),
            draft.get("message"),
            draft.get("connection_note"),
            draft.get("generated_at"),
        ]

        for col, value in enumerate(data, 1):
            cell = ws_linkedin.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths
    for ws in [ws_email, ws_linkedin]:
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 25

    # Save
    output_file = EXPORTS_DIR / "final_outreach_drafts.xlsx"
    wb.save(output_file)
    print(f"Saved outreach drafts to {output_file}")


def main():
    """Export all files."""
    print("Exporting sales queue to XLSX...")
    export_sales_queue_xlsx()

    print("\nExporting outreach drafts to XLSX...")
    export_outreach_drafts_xlsx()

    print("\nExport complete!")


if __name__ == "__main__":
    main()
