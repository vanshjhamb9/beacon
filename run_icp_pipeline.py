"""ICP-Based Pipeline - Discover only businesses that match COMAI's Ideal Customer Profile."""

import asyncio
import logging
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / "packages"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("icp_pipeline")


async def run_icp_pipeline():
    from packages.ecommerce_leads.models import RawEcommerceLead, EnrichedEcommerceLead
    from packages.ecommerce_leads.enrichment.technology_enrichment import TechnologyEnricher
    from packages.ecommerce_leads.enrichment.contact_enrichment import ContactEnricher
    from packages.ecommerce_leads.collectors.social_collector import SocialCollector
    from packages.ecommerce_leads.scoring.ecommerce_score import EcommerceScorer
    from packages.ecommerce_leads.scoring.quality_gate import QualityGate
    from packages.ecommerce_leads.scoring.sales_intelligence import SalesIntelligenceGenerator
    from packages.ecommerce_leads.scoring.icp_scorer import ICPScorer, ICPScoringResult
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    tech_enricher = TechnologyEnricher(timeout=20.0)
    contact_enricher = ContactEnricher(timeout=20.0)
    social_collector = SocialCollector(timeout=20.0)
    scorer = EcommerceScorer()
    quality_gate = QualityGate()
    sales_intel = SalesIntelligenceGenerator()
    icp_scorer = ICPScorer()

    # Curated list of known Indian D2C brands that are strong ICP candidates
    # Focus on: Fashion, Beauty, Skincare, Cosmetics, Jewellery, Home Decor,
    # Pet Products, Health, Food, Footwear, Electronics, Baby Products, Lifestyle
    icp_candidates = [
        # Beauty & Skincare (High priority)
        ("mamaearth.in", "Mamaearth", "beauty", "Mumbai", "Maharashtra", "Indian D2C beauty brand offering natural personal care products. Direct to consumer model with strong online presence.", 500),
        ("beardo.in", "Beardo", "grooming", "Mumbai", "Maharashtra", "Indian men's grooming brand selling beard care, hair care, and skincare products. D2C focused.", 200),
        ("mcaffeine.com", "mCaffeine", "skincare", "Mumbai", "Maharashtra", "India's first caffeinated personal care brand for millennials. D2C beauty brand.", 300),
        ("sugarcosmetics.com", "Sugar Cosmetics", "cosmetics", "Mumbai", "Maharashtra", "Premium Indian makeup brand with 10,000+ retail touchpoints. D2C cosmetics.", 500),
        ("wowskinscience.com", "WOW Skin Science", "skincare", "Bangalore", "Karnataka", "Indian D2C personal care brand. Natural and active ingredients.", 300),
        ("plumgoodness.com", "Plum Goodness", "beauty", "Mumbai", "Maharashtra", "100% vegan and cruelty-free Indian beauty brand. D2C beauty.", 250),
        ("bombayshavingcompany.com", "Bombay Shaving Company", "grooming", "New Delhi", "Delhi", "Indian men's grooming D2C brand. Direct to consumer剃须产品.", 200),
        ("dermaco.in", "Derma Co", "skincare", "New Delhi", "Delhi", "Indian derma-cosmetics brand by Mamaearth parent. D2C skincare.", 250),
        ("juicychemistry.com", "Juicy Chemistry", "organic beauty", "Coimbatore", "Tamil Nadu", "Indian organic and natural skincare brand. D2C organic beauty.", 150),
        ("khadinatural.com", "Khadi Natural", "natural products", "Ahmedabad", "Gujarat", "Indian natural personal care brand. D2C natural products.", 200),
        ("rusticart.in", "Rustic Art", "personal care", "Pune", "Maharashtra", "Indian organic personal care brand. D2C organic products.", 150),
        ("colorbarcosmetics.com", "Colorbar Cosmetics", "cosmetics", "New Delhi", "Delhi", "Premium Indian cosmetics brand. D2C cosmetics.", 300),
        ("theskinstory.in", "The Skin Story", "skincare", "Mumbai", "Maharashtra", "Indian skincare brand for modern consumers. D2C skincare.", 150),
        ("deconstruct.in", "Deconstruct", "skincare", "Bangalore", "Karnataka", "Indian skincare brand with active ingredients. D2C skincare.", 200),

        # Fashion (High priority)
        ("thesouledstore.com", "The Souled Store", "fashion", "Mumbai", "Maharashtra", "India's largest pop-culture fashion brand. D2C fashion.", 400),
        ("bewakoof.com", "Bewakoof", "fashion", "Mumbai", "Maharashtra", "India's leading D2C fashion brand for youth. Direct to consumer fashion.", 500),
        ("snitch.co.in", "Snitch", "fashion", "Bangalore", "Karnataka", "Fast-growing Indian menswear D2C brand. D2C fashion.", 300),
        ("vanheusen.com", "Van Heusen", "fashion", "Mumbai", "Maharashtra", "Premium Indian fashion brand by Aditya Birla. D2C fashion.", 400),
        ("parkavenue.in", "Park Avenue", "fashion", "Mumbai", "Maharashtra", "Indian premium menswear brand. D2C fashion.", 300),
        ("wooplr.com", "Wooplr", "fashion", "Bangalore", "Karnataka", "Indian fashion discovery platform. D2C fashion.", 200),
        ("styched.in", "Styched", "fashion", "Bangalore", "Karnataka", "Indian online fashion brand for youth. D2C fashion.", 150),

        # Electronics Accessories (Medium-high priority)
        ("ambraneindia.com", "Ambrane", "electronics", "New Delhi", "Delhi", "Indian consumer electronics brand for power banks. D2C electronics.", 150),
        ("pTron.com", "pTron", "electronics", "Hyderabad", "Telangana", "Indian budget audio and mobile accessories brand. D2C electronics.", 200),

        # Jewellery (High priority)
        ("caratlane.com", "CaratLane", "jewelry", "Chennai", "Tamil Nadu", "India's largest omnichannel jewelry brand. D2C jewelry.", 500),
        ("bluestone.com", "BlueStone", "jewelry", "Bangalore", "Karnataka", "Leading online jewelry brand in India. D2C jewelry.", 400),

        # Home Decor (Medium priority)
        ("addresshome.com", "Address Home", "home decor", "New Delhi", "Delhi", "Indian premium home decor and lifestyle brand. D2C home decor.", 150),
        ("jaypore.com", "Jaypore", "lifestyle", "New Delhi", "Delhi", "Indian online lifestyle and home decor brand. D2C lifestyle.", 200),

        # Kids & Baby (Medium priority)
        ("hopscotch.in", "Hopscotch", "kids", "Mumbai", "Maharashtra", "India's leading online store for kids' fashion. D2C kids.", 250),

        # Lifestyle (Medium priority)
        ("nicobar.com", "Nicobar", "lifestyle", "Mumbai", "Maharashtra", "Modern Indian lifestyle brand for travel and home. D2C lifestyle.", 150),
        ("okhai.org", "Okhai", "handicrafts", "Ahmedabad", "Gujarat", "Handcrafted lifestyle brand empowering rural artisans. D2C handicrafts.", 100),

        # Pet Products (High priority - underserved market)
        ("headsupfortails.com", "Heads Up For Tails", "pets", "New Delhi", "Delhi", "Indian premium pet care brand. D2C pet products.", 200),

        # Food & Beverage (Medium priority)
        ("teabox.com", "Teabox", "tea", "Kolkata", "West Bengal", "Indian premium tea D2C brand. D2C food.", 100),
        ("rawpressery.com", "Raw Pressery", "beverages", "Mumbai", "Maharashtra", "India's largest cold-pressed juice brand. D2C beverages.", 200),
        ("box8.in", "BOX8", "food", "Mumbai", "Maharashtra", "Indian all-in-one meal delivery brand. D2C food.", 150),

        # Footwear (Medium priority)
        ("boat-lifestyle.com", "boAt", "electronics", "New Delhi", "Delhi", "India's #1 audio and wearables brand with 8M+ products sold. D2C electronics.", 500),
        ("noise.tech", "Noise", "electronics", "Gurugram", "Haryana", "India's leading connected lifestyle brand for smartwatches and audio. D2C electronics.", 400),
        ("fireboltt.com", "Fire-Boltt", "electronics", "New Delhi", "Delhi", "India's fastest growing smartwatch and audio brand. D2C electronics.", 350),

        # Additional D2C brands
        ("firstcry.com", "FirstCry", "kids", "Pune", "Maharashtra", "Asia's largest online store for kids and baby products. D2C kids.", 1000),
    ]

    start_time = time.time()
    all_leads = []
    qualified_leads = []
    seen_domains = set()

    logger.info("=" * 70)
    logger.info("ICP-BASED PIPELINE - Discovering COMAI Ideal Customer Profile matches")
    logger.info("=" * 70)
    logger.info("ICP Criteria: Indian, Shopify/WooCommerce, D2C, 5-100 employees,")
    logger.info("No AI chatbot, WhatsApp Business, Strong social presence")
    logger.info("=" * 70)

    for i, (domain, name, category, city, state, desc, products) in enumerate(icp_candidates):
        if domain in seen_domains:
            continue
        seen_domains.add(domain)

        logger.info("[%d/%d] Processing %s...", i+1, len(icp_candidates), name)

        raw = RawEcommerceLead(
            company_name=name,
            website=f"https://{domain}",
            domain=domain,
            platform="shopify",  # Will be verified by detector
            industry=category,
            category=category,
            country="India",
            city=city,
            state=state,
            description=desc,
            product_count=products,
            source="icp_seed_data",
        )

        lead = EnrichedEcommerceLead(raw=raw)

        # Technology detection
        try:
            lead = await tech_enricher.enrich(lead)
        except Exception as e:
            logger.warning("  Tech enrichment failed: %s", e)

        # Contact enrichment
        try:
            lead = await contact_enricher.enrich(lead)
        except Exception as e:
            logger.warning("  Contact enrichment failed: %s", e)

        # Social links
        try:
            social = await social_collector.enrich_social_links(raw.website)
            lead.raw.social_links.update(social)
            lead.instagram_url = social.get("instagram", "")
            lead.facebook_url = social.get("facebook", "")
            lead.linkedin_url = social.get("linkedin", "")
        except Exception as e:
            logger.warning("  Social enrichment failed: %s", e)

        # Score with COMAI scorer
        lead = scorer.score(lead)

        # Quality gate
        lead = quality_gate.evaluate(lead)

        # ICP scoring
        icp_result = icp_scorer.score(lead)

        # Check if qualifies for export
        qualifies, export_reasons = icp_scorer.qualifies_for_export(icp_result, lead)

        # Sales intelligence
        lead = sales_intel.generate(lead)

        # Store ICP results in lead metadata
        lead.raw.metadata["icp_result"] = {
            "icp_match_score": icp_result.icp_match_score,
            "buying_probability": icp_result.buying_probability,
            "growth_score": icp_result.growth_score,
            "dm_accessibility": icp_result.decision_maker_accessibility,
            "support_pain": icp_result.support_pain_score,
            "overall_comai_sales_score": icp_result.overall_comai_sales_score,
            "qualifies": qualifies,
            "export_reasons": export_reasons,
            "reasons": icp_result.reasons,
            "disqualifiers": icp_result.disqualifiers,
        }

        all_leads.append(lead)

        if qualifies:
            qualified_leads.append(lead)
            logger.info("  ✓ QUALIFIED | ICP: %.0f | Buy Prob: %.0f | Overall: %.0f | Score: %.0f",
                        icp_result.icp_match_score, icp_result.buying_probability,
                        icp_result.overall_comai_sales_score, lead.comai_score)
        else:
            logger.info("  ✗ NOT QUALIFIED | ICP: %.0f | Buy Prob: %.0f | Reasons: %s",
                        icp_result.icp_match_score, icp_result.buying_probability,
                        "; ".join(export_reasons[:2]))

    elapsed = time.time() - start_time
    logger.info("=" * 70)
    logger.info("Pipeline completed in %.1f seconds", elapsed)

    # Stats
    logger.info("")
    logger.info("RESULTS SUMMARY:")
    logger.info("  Total leads processed: %d", len(all_leads))
    logger.info("  Qualified for export: %d", len(qualified_leads))
    logger.info("  Qualification rate: %.1f%%", len(qualified_leads)/len(all_leads)*100 if all_leads else 0)

    # Detection stats
    platform_detected = sum(1 for l in all_leads if l.raw.platform and l.raw.platform != "unknown")
    tech_detected = sum(1 for l in all_leads if any([l.shopify_detected, l.woocommerce_detected, l.magento_detected, l.chatbot_detected, l.whatsapp_detected, l.crm_detected]))
    contacts = sum(1 for l in all_leads if l.email or l.phone)
    decision_makers = sum(1 for l in all_leads if l.founder_name)

    logger.info("")
    logger.info("DETECTION RATES:")
    logger.info("  Platform: %d/%d (%.0f%%)", platform_detected, len(all_leads), platform_detected/len(all_leads)*100 if all_leads else 0)
    logger.info("  Technology: %d/%d (%.0f%%)", tech_detected, len(all_leads), tech_detected/len(all_leads)*100 if all_leads else 0)
    logger.info("  Contacts: %d/%d (%.0f%%)", contacts, len(all_leads), contacts/len(all_leads)*100 if all_leads else 0)
    logger.info("  Decision Makers: %d/%d (%.0f%%)", decision_makers, len(all_leads), decision_makers/len(all_leads)*100 if all_leads else 0)

    # Export
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["COMAI ICP-QUALIFIED LEADS"])
    ws_summary.append([])
    ws_summary.append(["Generated", time.strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Total Processed", len(all_leads)])
    ws_summary.append(["ICP Qualified", len(qualified_leads)])
    ws_summary.append(["Qualification Rate", f"{len(qualified_leads)/len(all_leads)*100:.1f}%" if all_leads else "0%"])
    ws_summary.append([])
    ws_summary.append(["ICP CRITERIA:"])
    ws_summary.append(["- Indian businesses only"])
    ws_summary.append(["- Shopify or WooCommerce stores"])
    ws_summary.append(["- Small to mid-sized D2C brands"])
    ws_summary.append(["- 5-100 employees (estimated)"])
    ws_summary.append(["- No AI chatbot detected"])
    ws_summary.append(["- WhatsApp Business present"])
    ws_summary.append(["- Strong Instagram/Facebook presence"])
    ws_summary.append([])
    ws_summary.append(["EXPORT CRITERIA:"])
    ws_summary.append(["- ICP Match Score >= 80"])
    ws_summary.append(["- Buying Probability >= 75"])
    ws_summary.append(["- Valid website"])
    ws_summary.append(["- At least one verified contact"])
    ws_summary.append(["- Clear COMAI fit reason"])

    ws_leads = wb.create_sheet("ICP Qualified Leads")
    headers = [
        "Company", "Website", "Platform", "Industry", "City",
        "Email", "Phone", "Founder", "Role",
        "Shopify", "Chatbot", "WhatsApp", "CRM",
        "ICP Match", "Buy Prob", "Growth", "DM Access", "Support Pain", "Overall Score",
        "COMAI Score", "Priority",
        "Sales Reason", "Pain Points",
        "Call Opener", "Pitch Angle", "Recommended Feature",
    ]
    ws_leads.append(headers)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for cell in ws_leads[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    for lead in qualified_leads:
        icp_data = lead.raw.metadata.get("icp_result", {})
        ws_leads.append([
            lead.raw.company_name,
            lead.raw.website,
            lead.raw.platform,
            lead.raw.industry,
            lead.raw.city,
            lead.email,
            lead.phone,
            lead.founder_name,
            lead.decision_maker_role,
            lead.shopify_detected,
            lead.chatbot_detected,
            lead.whatsapp_detected,
            lead.crm_detected,
            icp_data.get("icp_match_score", 0),
            icp_data.get("buying_probability", 0),
            icp_data.get("growth_score", 0),
            icp_data.get("dm_accessibility", 0),
            icp_data.get("support_pain", 0),
            icp_data.get("overall_comai_sales_score", 0),
            lead.comai_score,
            lead.lead_priority,
            lead.sales_reason,
            "; ".join(lead.pain_points),
            (lead.call_opener or "")[:200],
            lead.pitch_angle,
            lead.recommended_feature,
        ])

    # Auto-width
    for col in ws_leads.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_leads.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output_path = Path(__file__).parent / "comai_icp_qualified_leads.xlsx"
    wb.save(str(output_path))
    logger.info("Exported to: %s", output_path)

    # Print top leads
    if qualified_leads:
        logger.info("")
        logger.info("TOP 5 ICP-QUALIFIED LEADS:")
        logger.info("-" * 70)
        sorted_leads = sorted(qualified_leads, 
                            key=lambda l: l.raw.metadata.get("icp_result", {}).get("overall_comai_sales_score", 0),
                            reverse=True)
        for i, lead in enumerate(sorted_leads[:5], 1):
            icp_data = lead.raw.metadata.get("icp_result", {})
            logger.info("%d. %s", i, lead.raw.company_name)
            logger.info("   ICP Match: %.0f | Buy Prob: %.0f | Overall: %.0f",
                        icp_data.get("icp_match_score", 0),
                        icp_data.get("buying_probability", 0),
                        icp_data.get("overall_comai_sales_score", 0))
            logger.info("   Platform: %s | Chatbot: %s | WhatsApp: %s",
                        lead.raw.platform, lead.chatbot_detected, lead.whatsapp_detected)
            logger.info("   Email: %s | Phone: %s | Founder: %s",
                        bool(lead.email), bool(lead.phone), lead.founder_name)
            logger.info("")

    return len(qualified_leads)


if __name__ == "__main__":
    result = asyncio.run(run_icp_pipeline())
    print(f"\nDone! Generated {result} ICP-qualified leads.")
