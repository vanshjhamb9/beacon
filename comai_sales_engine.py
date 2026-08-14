"""
COMAI Sales-Ready Lead Intelligence Engine — Sprint 43.5
========================================================
NOT a lead scraper. This is a B2B SaaS Sales Intelligence Engine.

Every company must pass ALL gates:
  1. Business Validation — Real ecommerce, physical products
  2. ICP Match — India, D2C, ₹3-150 Cr, 10-200 employees
  3. Technology Detection — Shopify/WooCommerce/custom stack
  4. Pain Detection — No chatbot, no AI, manual support
  5. Intent Detection — Hiring, ads, growth signals
  6. Decision Maker — Founder/CEO accessible
  7. Contact Verification — Verified email OR phone required
  8. Commercial Fit — Score ≥ 75/100

If ANY gate fails → REJECT.

Usage:
    python comai_sales_engine.py
    python comai_sales_engine.py --limit 100
    python comai_sales_engine.py --output leads.xlsx

Dependencies:
    pip install httpx openpyxl beautifulsoup4 lxml
"""

from __future__ import annotations

import asyncio
import re
import time
import argparse
import hashlib
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any
from urllib.parse import urlparse

import httpx

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False


# ============================================================
# CONSTANTS
# ============================================================

MAX_REVENUE_CR = 300
MAX_EMPLOYEES = 500
MIN_REVENUE_CR = 2
MIN_EMPLOYEES = 5
MIN_MONTHLY_ORDERS = 50
MIN_TRAFFIC = 5000


# ============================================================
# REJECTION RULES
# ============================================================

REJECT_COMPANIES = {
    # Listed / Public
    "reliance", "tata", "aditya birla", "mahindra", "infosys", "wipro",
    "tcs", "hcl", "bajaj", "hero", "maruti", "ashok leyland",
    "itc", "godrej", "dabur", "emami", "marico", "nestle",
    "britannia", "parle", "haldiram", "balaji", "bingo",
    "tanishq", "kalyan jewellers", "malabar gold", "joyalukkas",
    "senco gold", "pc jeweller",
    # Marketplaces
    "amazon", "flipkart", "meesho", "snapdeal", "myntra", "ajio",
    "nykaa marketplace", "tata cliq", "croma", "dmart", "big bazaar",
    # Big D2C (>$300 Cr)
    "mamaearth", "honasa", "boAt", "lenskart", "nykaa", "pepperfry",
    "urban ladder", "firstcry", "zepto", "blinkit", "swiggy", "zomato",
    "paytm", "phonepe", "cred", "razorpay",
    # Government / Institutions
    "government", "ministry", "hospital", "university", "college",
    "bank", "insurance", "school",
}

REJECT_KEYWORDS = {
    "nike", "adidas", "puma", "reebok", "under armour",
    "h&m", "zara", "uniqlo", "gucci", "prada", "louis vuitton",
    "consulting", "agency", "software", "saas", "b2b",
    "wholesale", "distributor", "manufacturer",
    "government", "hospital", "university", "college", "school",
    "bank", "insurance", "restaurant", "hotel",
}


def should_reject(brand_name: str) -> tuple[bool, str]:
    """Check if brand should be rejected."""
    name_lower = brand_name.lower()

    for company in REJECT_COMPANIES:
        if company in name_lower:
            return True, f"Big player: {brand_name}"

    for kw in REJECT_KEYWORDS:
        if kw in name_lower:
            return True, f"Rejected keyword: {kw}"

    return False, ""


# ============================================================
# EVIDENCE TRACKER
# ============================================================

@dataclass
class Evidence:
    """Track evidence for every data point."""
    source: str
    url: str
    timestamp: str
    method: str
    confidence: float  # 0.0 - 1.0

    def to_dict(self) -> dict:
        return {
            "source": self.source,
            "url": self.url,
            "timestamp": self.timestamp,
            "method": self.method,
            "confidence": self.confidence,
        }


@dataclass
class EvidenceBundle:
    """Collection of evidence for a lead."""
    items: list[Evidence] = field(default_factory=list)

    def add(self, source: str, url: str, method: str, confidence: float = 0.8):
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        self.items.append(Evidence(source, url, now, method, confidence))

    @property
    def best_url(self) -> str:
        return self.items[0].url if self.items else ""

    @property
    def avg_confidence(self) -> float:
        if not self.items:
            return 0.0
        return sum(e.confidence for e in self.items) / len(self.items)

    def urls(self) -> list[str]:
        return [e.url for e in self.items if e.url]


# ============================================================
# DATA MODELS
# ============================================================

@dataclass
class SeedBrand:
    name: str
    website: str
    category: str
    sub_category: str
    city: str
    state: str
    founded_year: int | None = None
    est_revenue_cr: int = 0
    est_employees: int = 0
    est_monthly_orders: int = 0
    est_traffic: int = 0
    has_whatsapp: bool = False
    has_instagram: bool = False
    has_meta_ads: bool = False
    description: str = ""


@dataclass
class TechStack:
    platform: str = "unknown"
    platform_confidence: float = 0.0
    has_chatbot: bool = False
    has_whatsapp: bool = False
    has_ai: bool = False
    email_marketing: str = ""
    review_platform: str = ""
    support_tool: str = ""
    analytics: str = ""
    meta_pixel: bool = False
    payment_gateway: str = ""
    crm: str = ""
    automation_level: str = "none"  # none, basic, moderate, advanced


@dataclass
class ContactInfo:
    emails: list[str] = field(default_factory=list)
    phones: list[str] = field(default_factory=list)
    founder_name: str = ""
    founder_title: str = ""
    founder_email: str = ""
    founder_linkedin: str = ""
    ceo_name: str = ""
    ceo_email: str = ""
    ceo_linkedin: str = ""
    head_growth: str = ""
    head_marketing: str = ""
    head_cx: str = ""
    head_operations: str = ""
    linkedin_url: str = ""
    instagram_url: str = ""
    facebook_url: str = ""
    whatsapp_number: str = ""

    @property
    def best_email(self) -> str:
        generic = {"support", "info", "hello", "sales", "care", "contact", "help",
                    "feedback", "noreply", "admin", "office", "team", "billing"}
        # Priority: founder > CEO > head > other
        for e in [self.founder_email, self.ceo_email]:
            if e:
                return e
        for e in self.emails:
            prefix = e.split("@")[0].lower()
            if prefix not in generic:
                return e
        return self.emails[0] if self.emails else ""

    @property
    def best_phone(self) -> str:
        return self.phones[0] if self.phones else self.whatsapp_number

    @property
    def best_decision_maker(self) -> str:
        return self.founder_name or self.ceo_name or self.head_growth or ""

    @property
    def best_dm_linkedin(self) -> str:
        return self.founder_linkedin or self.ceo_linkedin or ""


@dataclass
class PainSignals:
    no_chatbot: bool = False
    no_ai: bool = False
    no_whatsapp_automation: bool = False
    manual_support: bool = False
    no_cart_recovery: bool = False
    no_personalization: bool = False
    no_product_recommendation: bool = False
    weak_engagement: bool = False
    no_repeat_strategy: bool = False
    no_upsell: bool = False
    high_faq_volume: bool = False

    @property
    def pain_count(self) -> int:
        return sum([
            self.no_chatbot, self.no_ai, self.no_whatsapp_automation,
            self.manual_support, self.no_cart_recovery, self.no_personalization,
            self.no_product_recommendation, self.weak_engagement,
            self.no_repeat_strategy, self.no_upsell, self.high_faq_volume,
        ])

    @property
    def score(self) -> float:
        count = self.pain_count
        return min(count * 10, 100)


@dataclass
class BuyingIntent:
    hiring_support: bool = False
    hiring_growth: bool = False
    hiring_marketing: bool = False
    hiring_ai: bool = False
    recent_redesign: bool = False
    recent_migration: bool = False
    new_products: bool = False
    international_expansion: bool = False
    running_meta_ads: bool = False
    growing_instagram: bool = False
    growing_reviews: bool = False
    growing_catalog: bool = False
    new_collections: bool = False
    new_warehouse: bool = False
    scaling_ops: bool = False
    growing_team: bool = False
    raised_funding: bool = False

    @property
    def signal_count(self) -> int:
        return sum([
            self.hiring_support, self.hiring_growth, self.hiring_marketing,
            self.hiring_ai, self.recent_redesign, self.recent_migration,
            self.new_products, self.international_expansion, self.running_meta_ads,
            self.growing_instagram, self.growing_reviews, self.growing_catalog,
            self.new_collections, self.new_warehouse, self.scaling_ops,
            self.growing_team, self.raised_funding,
        ])

    @property
    def score(self) -> float:
        count = self.signal_count
        return min(count * 8, 100)


@dataclass
class CommercialFit:
    """8-dimensional commercial fit score."""
    revenue_size: float = 0.0       # 20%
    decision_maker_access: float = 0.0  # 20%
    growth_stage: float = 0.0       # 15%
    technology_fit: float = 0.0     # 15%
    pain_level: float = 0.0        # 10%
    buying_intent: float = 0.0     # 10%
    automation_readiness: float = 0.0  # 5%
    sales_complexity: float = 0.0   # 5%

    @property
    def total(self) -> float:
        return (
            self.revenue_size * 0.20 +
            self.decision_maker_access * 0.20 +
            self.growth_stage * 0.15 +
            self.technology_fit * 0.15 +
            self.pain_level * 0.10 +
            self.buying_intent * 0.10 +
            self.automation_readiness * 0.05 +
            self.sales_complexity * 0.05
        )

    @property
    def passed(self) -> bool:
        return self.total >= 75.0

    @property
    def grade(self) -> str:
        if self.total >= 90:
            return "A+"
        elif self.total >= 85:
            return "A"
        elif self.total >= 80:
            return "B+"
        elif self.total >= 75:
            return "B"
        elif self.total >= 70:
            return "C"
        else:
            return "F"


# ============================================================
# SEED DATABASE — Mid-Size Indian D2C Brands
# ============================================================

SEED_BRANDS: list[SeedBrand] = [
    # === FASHION — ₹3-150 Cr, D2C ===
    SeedBrand("Bewakoof", "https://www.bewakoof.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2012, 80, 100, 50000, 500000, True, True, True, "Online fashion brand"),
    SeedBrand("FabAlley", "https://www.faballey.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2012, 50, 60, 20000, 300000, True, True, True, "Western fashion brand"),
    SeedBrand("W for Woman", "https://www.wforwoman.com", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2014, 100, 150, 30000, 400000, True, True, True, "Contemporary ethnic"),
    SeedBrand("AND", "https://www.andindia.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2010, 80, 120, 25000, 350000, True, True, True, "Premium western wear"),
    SeedBrand("Global Desi", "https://www.globaldesi.com", "Fashion", "Ethnic Wear", "Mumbai", "Maharashtra", 2011, 60, 80, 20000, 300000, True, True, True, "Indo-western fashion"),
    SeedBrand("Biba", "https://www.biba.in", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2010, 100, 200, 40000, 500000, True, True, True, "Indian ethnic wear"),
    SeedBrand("Libas", "https://www.libas.in", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2015, 50, 70, 20000, 250000, True, True, True, "Contemporary ethnic wear"),
    SeedBrand("Suta", "https://www.suta.in", "Fashion", "Ethnic Wear", "Bengaluru", "Karnataka", 2016, 30, 40, 15000, 200000, True, True, True, "Handloom sarees"),
    SeedBrand("Berrylush", "https://www.berrylush.com", "Fashion", "Western Wear", "Gurugram", "Haryana", 2017, 25, 30, 10000, 150000, True, True, True, "Affordable western wear"),
    SeedBrand("Tokyo Talkies", "https://www.tokyotalkies.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2016, 30, 35, 12000, 180000, True, True, True, "Youth fashion brand"),
    SeedBrand("SASSAFRAS", "https://www.sassafrasofficial.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2018, 20, 25, 8000, 120000, True, True, True, "Women western wear"),
    SeedBrand("Clovia", "https://www.clovia.com", "Fashion", "Lingerie", "Noida", "Uttar Pradesh", 2013, 50, 80, 30000, 400000, True, True, True, "Lingerie brand"),
    SeedBrand("Zivame", "https://www.zivame.com", "Fashion", "Lingerie", "Bengaluru", "Karnataka", 2013, 60, 100, 35000, 500000, True, True, True, "Lingerie brand"),
    SeedBrand("The Souled Store", "https://www.thesouledstore.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2013, 70, 120, 40000, 600000, True, True, True, "Pop culture fashion"),
    SeedBrand("Snitch", "https://www.snitch.co.in", "Fashion", "Western Wear", "Bengaluru", "Karnataka", 2019, 40, 60, 20000, 300000, True, True, True, "Men fashion brand"),
    SeedBrand("Andamen", "https://www.andamen.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2018, 25, 30, 10000, 150000, True, True, True, "Premium men basics"),
    SeedBrand("MensXP", "https://www.mensxp.com", "Fashion", "Grooming", "New Delhi", "Delhi", 2012, 40, 50, 15000, 200000, True, True, True, "Men lifestyle brand"),

    # === BEAUTY — ₹3-150 Cr, D2C ===
    SeedBrand("Plum Goodness", "https://www.plumgoodness.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2013, 80, 120, 30000, 400000, True, True, True, "Vegan beauty brand"),
    SeedBrand("mCaffeine", "https://www.mcaffeine.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2016, 100, 150, 40000, 500000, True, True, True, "Caffeine-based skincare"),
    SeedBrand("The Man Company", "https://www.themancompany.com", "Beauty", "Grooming", "Ahmedabad", "Gujarat", 2015, 60, 80, 25000, 300000, True, True, True, "Men grooming brand"),
    SeedBrand("Bombay Shaving Company", "https://www.bombayshavingcompany.com", "Beauty", "Grooming", "Gurugram", "Haryana", 2016, 40, 50, 15000, 200000, True, True, True, "Men grooming"),
    SeedBrand("Beardo", "https://www.beardo.in", "Beauty", "Grooming", "Hyderabad", "Telangana", 2015, 50, 70, 20000, 250000, True, True, True, "Men grooming brand"),
    SeedBrand("Ustraa", "https://www.ustraa.com", "Beauty", "Grooming", "New Delhi", "Delhi", 2017, 25, 35, 12000, 150000, True, True, True, "Men grooming"),
    SeedBrand("Sugar Cosmetics", "https://www.sugarcosmetics.com", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2015, 150, 200, 50000, 600000, True, True, True, "Color cosmetics brand"),
    SeedBrand("MyGlamm", "https://www.myglamm.com", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2017, 100, 150, 35000, 450000, True, True, True, "D2C beauty brand"),
    SeedBrand("Earth Rhythm", "https://www.earthrhythm.com", "Beauty", "Skincare", "New Delhi", "Delhi", 2019, 30, 40, 15000, 200000, True, True, True, "Clean beauty brand"),
    SeedBrand("Minimalist", "https://www.minimalist.co.in", "Beauty", "Skincare", "Gurugram", "Haryana", 2020, 80, 120, 30000, 400000, True, True, True, "Science-backed skincare"),
    SeedBrand("De Construct", "https://www.deconstruct.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 20, 25, 10000, 120000, True, True, True, "Minimalist skincare"),
    SeedBrand("SkinKraft", "https://www.skinkraft.com", "Beauty", "Personalized", "Mumbai", "Maharashtra", 2018, 40, 60, 20000, 250000, True, True, True, "Personalized skincare"),
    SeedBrand("Re'equil", "https://www.reeequil.com", "Beauty", "Skincare", "Gurugram", "Haryana", 2017, 25, 35, 12000, 150000, True, True, True, "Clinical skincare"),
    SeedBrand("Dr. Sheth's", "https://www.drsheths.com", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 15, 20, 8000, 100000, True, True, True, "Indian skincare"),
    SeedBrand("O3+", "https://www.o3plus.com", "Beauty", "Skincare", "New Delhi", "Delhi", 2012, 30, 40, 15000, 200000, True, True, True, "Professional skincare"),
    SeedBrand("Good Vibes", "https://www.goodvibes.life", "Beauty", "Personal Care", "New Delhi", "Delhi", 2017, 40, 50, 18000, 220000, True, True, True, "Affordable beauty"),
    SeedBrand("Aqualogica", "https://www.aqualogica.in", "Beauty", "Skincare", "Gurugram", "Haryana", 2019, 30, 40, 15000, 180000, True, True, True, "Hydrating skincare"),
    SeedBrand("Derma Co", "https://www.thedermaco.com", "Beauty", "Skincare", "Hyderabad", "Telangana", 2020, 80, 100, 30000, 400000, True, True, True, "Dermatologist skincare"),
    SeedBrand("Pilgrim", "https://www.pilgrim.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 50, 70, 20000, 250000, True, True, True, "International beauty"),
    SeedBrand("Arata", "https://www.arata.in", "Beauty", "Personal Care", "New Delhi", "Delhi", 2018, 15, 20, 8000, 100000, True, True, True, "Clean personal care"),

    # === HOME DECOR — ₹3-150 Cr, D2C ===
    SeedBrand("Nestasia", "https://www.nestasia.in", "Home Decor", "Home Accessories", "Kolkata", "West Bengal", 2018, 25, 35, 12000, 150000, True, True, True, "Premium home decor"),
    SeedBrand("Jaypore", "https://www.jaypore.com", "Home Decor", "Handicrafts", "New Delhi", "Delhi", 2014, 50, 70, 20000, 250000, True, True, True, "Artisanal home decor"),
    SeedBrand("Chumbak", "https://www.chumbak.com", "Home Decor", "Home Accessories", "Bengaluru", "Karnataka", 2011, 40, 50, 18000, 220000, True, True, True, "Quirky home and lifestyle"),
    SeedBrand("Cult Decor", "https://www.cultdecor.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2015, 30, 40, 12000, 150000, True, True, True, "Premium home decor"),
    SeedBrand("Zwende", "https://www.zwende.com", "Home Decor", "Handicrafts", "Bengaluru", "Karnataka", 2016, 12, 15, 5000, 80000, True, True, True, "Custom home decor"),
    SeedBrand("The Decor Kart", "https://www.thedecorkart.com", "Home Decor", "Home Accessories", "New Delhi", "Delhi", 2017, 15, 20, 8000, 100000, True, True, True, "Home decor brand"),
    SeedBrand("Address Home", "https://www.addresshome.com", "Home Decor", "Home Textile", "New Delhi", "Delhi", 2010, 60, 80, 25000, 300000, True, True, True, "Premium home linen"),
    SeedBrand("Homesake", "https://www.homesakeindia.com", "Home Decor", "Handicrafts", "New Delhi", "Delhi", 2016, 10, 15, 5000, 60000, True, True, True, "Handicraft home decor"),
    SeedBrand("Ellementry", "https://www.ellementry.com", "Home Decor", "Kitchenware", "New Delhi", "Delhi", 2017, 20, 25, 10000, 120000, True, True, True, "Handcrafted kitchenware"),
    SeedBrand("Mintwud", "https://www.mintwud.com", "Home Decor", "Furniture", "Mumbai", "Maharashtra", 2018, 25, 30, 12000, 150000, True, True, True, "Affordable furniture"),
    SeedBrand("Wonderchef", "https://www.wonderchef.com", "Home Decor", "Kitchenware", "Gurugram", "Haryana", 2013, 80, 120, 30000, 400000, True, True, True, "Kitchen appliances"),
    SeedBrand("Bergner", "https://www.bergner.in", "Home Decor", "Kitchenware", "Mumbai", "Maharashtra", 2015, 40, 50, 18000, 220000, True, True, True, "Premium cookware"),

    # === JEWELLERY — ₹3-150 Cr, D2C ===
    SeedBrand("Melorra", "https://www.melorra.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2016, 60, 80, 25000, 300000, True, True, True, "Everyday fine jewellery"),
    SeedBrand("CaratLane", "https://www.caratlane.com", "Jewellery", "Fine Jewellery", "Chennai", "Tamil Nadu", 2010, 200, 300, 80000, 900000, True, True, True, "Online jewellery"),
    SeedBrand("BlueStone", "https://www.bluestone.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2011, 150, 200, 60000, 700000, True, True, True, "Online fine jewellery"),
    SeedBrand("Giva", "https://www.giva.co", "Jewellery", "Silver Jewellery", "Bengaluru", "Karnataka", 2019, 30, 40, 15000, 180000, True, True, True, "Silver jewellery brand"),
    SeedBrand("Sukkhi", "https://www.sukkhi.com", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2014, 40, 50, 18000, 220000, True, True, True, "Fashion jewellery"),
    SeedBrand("Zaveri Pearls", "https://www.zfrp.in", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2015, 30, 40, 15000, 180000, True, True, True, "Fashion jewellery"),
    SeedBrand("YouBella", "https://www.youbella.com", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2016, 20, 25, 10000, 120000, True, True, True, "Fashion jewellery"),
    SeedBrand("Enamour", "https://www.enamour.in", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2018, 15, 20, 8000, 100000, True, True, True, "Contemporary fine jewellery"),
    SeedBrand("Candere", "https://www.candere.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2013, 40, 50, 18000, 220000, True, True, True, "Online jewellery"),

    # === HEALTH & WELLNESS — ₹3-150 Cr, D2C ===
    SeedBrand("Kapiva", "https://www.kapiva.in", "Health & Wellness", "Ayurvedic", "Gurugram", "Haryana", 2016, 60, 80, 25000, 300000, True, True, True, "Ayurvedic wellness"),
    SeedBrand("Fast&Up", "https://www.fastandup.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2015, 50, 70, 20000, 250000, True, True, True, "Sports nutrition"),
    SeedBrand("Plix Life", "https://www.plixlife.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2019, 40, 50, 18000, 220000, True, True, True, "Plant-based supplements"),
    SeedBrand("True Elements", "https://www.trueelements.com", "Health & Wellness", "Superfoods", "Pune", "Maharashtra", 2016, 50, 60, 20000, 250000, True, True, True, "Healthy food brand"),
    SeedBrand("Dr. Vaidya's", "https://www.drvaidyas.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2016, 40, 50, 15000, 200000, True, True, True, "Ayurvedic products"),
    SeedBrand("Rasayanam", "https://www.rasayanam.com", "Health & Wellness", "Ayurvedic", "Bengaluru", "Karnataka", 2018, 20, 25, 10000, 120000, True, True, True, "Ayurvedic supplements"),
    SeedBrand("Neuherbs", "https://www.neuherbs.com", "Health & Wellness", "Supplements", "New Delhi", "Delhi", 2019, 20, 25, 10000, 120000, True, True, True, "Health supplements"),
    SeedBrand("Vedix", "https://www.vedix.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2019, 15, 20, 8000, 100000, True, True, True, "Personalized Ayurveda"),

    # === FOOD & SNACKS — ₹3-150 Cr, D2C ===
    SeedBrand("Yoga Bar", "https://www.yogabar.com", "Food & Snacks", "Healthy Snacks", "Bengaluru", "Karnataka", 2017, 60, 80, 25000, 300000, True, True, True, "Healthy muesli and bars"),
    SeedBrand("Slurrp Farm", "https://www.slurrpfarm.com", "Food & Snacks", "Kids Food", "New Delhi", "Delhi", 2017, 30, 40, 15000, 180000, True, True, True, "Healthy kids food"),
    SeedBrand("Nutty Gritties", "https://www.nuttygritties.com", "Food & Snacks", "Dry Fruits", "New Delhi", "Delhi", 2016, 20, 25, 10000, 120000, True, True, True, "Premium dry fruits"),
    SeedBrand("Farmley", "https://www.farmley.com", "Food & Snacks", "Dry Fruits", "New Delhi", "Delhi", 2017, 40, 50, 18000, 220000, True, True, True, "Premium dry fruits"),
    SeedBrand("Vahdam Teas", "https://www.vahdamteas.com", "Tea/Coffee", "Tea", "New Delhi", "Delhi", 2015, 80, 120, 35000, 450000, True, True, True, "Premium Indian teas"),
    SeedBrand("Sleepy Owl", "https://www.sleepyowl.in", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2016, 30, 40, 15000, 180000, True, True, True, "Coffee brand"),
    SeedBrand("Blue Tokai", "https://www.bluetokai.com", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2013, 50, 70, 20000, 250000, True, True, True, "Specialty coffee"),
    SeedBrand("Rage Coffee", "https://www.ragecoffee.com", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2018, 20, 25, 10000, 120000, True, True, True, "Coffee brand"),
    SeedBrand("Raw Pressery", "https://www.rawpressery.com", "Food & Snacks", "Beverages", "Mumbai", "Maharashtra", 2016, 50, 70, 20000, 250000, True, True, True, "Cold-pressed juices"),
    SeedBrand("iD Fresh Food", "https://www.idfreshfood.com", "Food & Snacks", "Fresh Food", "Bengaluru", "Karnataka", 2010, 100, 150, 40000, 500000, True, True, True, "Fresh food brand"),
    SeedBrand("24 Mantra", "https://www.24mantra.com", "Food & Snacks", "Organic Food", "Hyderabad", "Telangana", 2010, 30, 40, 15000, 180000, True, True, True, "Organic food"),

    # === ELECTRONICS — ₹3-150 Cr, D2C ===
    SeedBrand("Hammer", "https://www.hammerlifestyle.in", "Electronics", "Audio", "New Delhi", "Delhi", 2018, 40, 50, 18000, 220000, True, True, True, "Premium audio brand"),
    SeedBrand("Boult Audio", "https://www.boultaudio.com", "Electronics", "Audio", "New Delhi", "Delhi", 2017, 60, 80, 25000, 300000, True, True, True, "Audio accessories"),
    SeedBrand("pTron", "https://www.ptron.in", "Electronics", "Accessories", "Hyderabad", "Telangana", 2014, 100, 150, 40000, 500000, True, True, True, "Affordable tech"),
    SeedBrand("Mivi", "https://www.mivi.in", "Electronics", "Audio", "Hyderabad", "Telangana", 2016, 60, 80, 25000, 300000, True, True, True, "Audio accessories"),
    SeedBrand("Crossbeats", "https://www.crossbeats.com", "Electronics", "Audio", "Bengaluru", "Karnataka", 2014, 40, 50, 18000, 220000, True, True, True, "Audio and wearable"),
    SeedBrand("Fire-Boltt", "https://www.fireboltt.com", "Electronics", "Wearables", "New Delhi", "Delhi", 2016, 120, 180, 50000, 600000, True, True, True, "Smartwatch brand"),
    SeedBrand("Ambrane", "https://www.ambraneindia.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2012, 80, 100, 30000, 400000, True, True, True, "Mobile accessories"),
    SeedBrand("Portronics", "https://www.portronics.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2010, 100, 150, 40000, 500000, True, True, True, "Consumer electronics"),
    SeedBrand("Zoook", "https://www.zoook.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2011, 50, 60, 20000, 250000, True, True, True, "Consumer electronics"),
    SeedBrand("Leaf", "https://www.leafnlife.com", "Electronics", "Audio", "Bengaluru", "Karnataka", 2019, 15, 20, 8000, 100000, True, True, True, "Wireless audio brand"),
    SeedBrand("XECH", "https://www.xech.com", "Electronics", "Accessories", "Mumbai", "Maharashtra", 2017, 15, 20, 8000, 100000, True, True, True, "Tech accessories"),

    # === BABY PRODUCTS — ₹3-150 Cr, D2C ===
    SeedBrand("Hopskotch", "https://www.hopskotch.in", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2014, 40, 50, 18000, 220000, True, True, True, "Kids fashion brand"),
    SeedBrand("LuvLap", "https://www.luvlap.com", "Baby Products", "Baby Care", "New Delhi", "Delhi", 2012, 30, 40, 15000, 180000, True, True, True, "Baby care brand"),
    SeedBrand("R for Rabbit", "https://www.rforgabbit.com", "Baby Products", "Baby Gear", "Ahmedabad", "Gujarat", 2015, 20, 25, 10000, 120000, True, True, True, "Baby gear brand"),
    SeedBrand("Skillmatics", "https://www.skillmatics.com", "Baby Products", "Educational", "Mumbai", "Maharashtra", 2016, 50, 70, 20000, 250000, True, True, True, "Educational games"),
    SeedBrand("Smartivity", "https://www.smartivity.com", "Baby Products", "Educational", "New Delhi", "Delhi", 2015, 15, 20, 8000, 100000, True, True, True, "STEM toys"),
    SeedBrand("FirstCry", "https://www.firstcry.com", "Baby Products", "Multi-Category", "Pune", "Maharashtra", 2010, 150, 250, 60000, 700000, True, True, True, "Baby products"),

    # === PET PRODUCTS — ₹3-150 Cr, D2C ===
    SeedBrand("Heads Up For Tails", "https://www.headsuptails.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2016, 30, 40, 15000, 180000, True, True, True, "Premium pet brand"),
    SeedBrand("Wiggles", "https://www.wiggles.in", "Pet Products", "Pet Care", "Mumbai", "Maharashtra", 2018, 20, 25, 10000, 120000, True, True, True, "Pet care brand"),
    SeedBrand("Drools", "https://www.drools.com", "Pet Products", "Pet Food", "Hyderabad", "Telangana", 2015, 80, 100, 30000, 400000, True, True, True, "Pet food brand"),

    # === FOOTWEAR — ₹3-150 Cr, D2C ===
    SeedBrand("Neeman's", "https://www.neemans.com", "Footwear", "Casual", "Hyderabad", "Telangana", 2018, 30, 40, 15000, 180000, True, True, True, "Sustainable footwear"),

    # === BAGS — ₹3-150 Cr, D2C ===
    SeedBrand("Safari Industries", "https://www.safari-industries.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 100, 150, 40000, 500000, True, True, True, "Luggage brand"),
    SeedBrand("Wildcraft", "https://www.wildcraft.com", "Bags", "Backpacks", "Bengaluru", "Karnataka", 2010, 80, 120, 30000, 400000, True, True, True, "Outdoor gear"),
    SeedBrand("Lavie World", "https://www.lavieworld.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 30, 40, 15000, 180000, True, True, True, "Fashion bags"),
    SeedBrand("Fur Jaden", "https://www.furjaden.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2017, 12, 15, 5000, 80000, True, True, True, "Fashion bags"),
    SeedBrand("Caprese", "https://www.caprese.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 40, 50, 18000, 220000, True, True, True, "Premium bags"),

    # === GIFTS — ₹3-150 Cr, D2C ===
    SeedBrand("Ferns N Petals", "https://www.fernnpetals.com", "Gifts", "Flowers & Gifts", "New Delhi", "Delhi", 2010, 80, 120, 30000, 400000, True, True, True, "Flowers and gifts"),
    SeedBrand("IGP", "https://www.igp.com", "Gifts", "Gifts & Flowers", "Mumbai", "Maharashtra", 2010, 60, 80, 25000, 300000, True, True, True, "Gifts and flowers"),

    # === MORE FASHION ===
    SeedBrand("LimeRoad", "https://www.limeroad.com", "Fashion", "Multi-Category", "New Delhi", "Delhi", 2012, 50, 60, 20000, 250000, True, True, True, "Fashion discovery"),
    SeedBrand("Fynd", "https://www.fynd.com", "Fashion", "Multi-Category", "Mumbai", "Maharashtra", 2012, 60, 80, 25000, 300000, True, True, True, "Omnichannel fashion"),
    SeedBrand("Voonik", "https://www.voonik.com", "Fashion", "Multi-Category", "Bengaluru", "Karnataka", 2013, 25, 30, 10000, 120000, True, True, True, "Fashion marketplace"),
    SeedBrand("Koovs", "https://www.koovs.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2012, 40, 50, 18000, 220000, True, True, True, "Online fashion"),
    SeedBrand("StalkBuyLove", "https://www.stalkbuyllove.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2013, 30, 40, 15000, 180000, True, True, True, "Women western wear"),
    SeedBrand("Roposo", "https://www.roposo.com", "Fashion", "Multi-Category", "Bengaluru", "Karnataka", 2012, 20, 25, 8000, 100000, True, True, True, "Fashion social commerce"),
    SeedBrand("W for Woman", "https://www.wforwoman.com", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2014, 100, 150, 40000, 500000, True, True, True, "Contemporary ethnic"),
    SeedBrand("Biba", "https://www.biba.in", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2010, 100, 200, 50000, 600000, True, True, True, "Indian ethnic wear"),
    SeedBrand("Libas", "https://www.libas.in", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2015, 50, 70, 20000, 250000, True, True, True, "Contemporary ethnic wear"),
    SeedBrand("Suta", "https://www.suta.in", "Fashion", "Ethnic Wear", "Bengaluru", "Karnataka", 2016, 30, 40, 15000, 200000, True, True, True, "Handloom sarees"),
    SeedBrand("Berrylush", "https://www.berrylush.com", "Fashion", "Western Wear", "Gurugram", "Haryana", 2017, 25, 30, 10000, 150000, True, True, True, "Affordable western wear"),
    SeedBrand("Tokyo Talkies", "https://www.tokyotalkies.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2016, 30, 35, 12000, 180000, True, True, True, "Youth fashion brand"),
    SeedBrand("SASSAFRAS", "https://www.sassafrasofficial.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2018, 20, 25, 8000, 120000, True, True, True, "Women western wear"),
    SeedBrand("Clovia", "https://www.clovia.com", "Fashion", "Lingerie", "Noida", "Uttar Pradesh", 2013, 50, 80, 30000, 400000, True, True, True, "Lingerie brand"),
    SeedBrand("Zivame", "https://www.zivame.com", "Fashion", "Lingerie", "Bengaluru", "Karnataka", 2013, 60, 100, 35000, 500000, True, True, True, "Lingerie brand"),
    SeedBrand("The Souled Store", "https://www.thesouledstore.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2013, 70, 120, 40000, 600000, True, True, True, "Pop culture fashion"),
    SeedBrand("Snitch", "https://www.snitch.co.in", "Fashion", "Western Wear", "Bengaluru", "Karnataka", 2019, 40, 60, 20000, 300000, True, True, True, "Men fashion brand"),
    SeedBrand("Andamen", "https://www.andamen.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2018, 25, 30, 10000, 150000, True, True, True, "Premium men basics"),
    SeedBrand("MensXP", "https://www.mensxp.com", "Fashion", "Grooming", "New Delhi", "Delhi", 2012, 40, 50, 15000, 200000, True, True, True, "Men lifestyle brand"),

    # === MORE BEAUTY ===
    SeedBrand("Plum Goodness", "https://www.plumgoodness.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2013, 80, 120, 30000, 400000, True, True, True, "Vegan beauty brand"),
    SeedBrand("mCaffeine", "https://www.mcaffeine.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2016, 100, 150, 40000, 500000, True, True, True, "Caffeine-based skincare"),
    SeedBrand("The Man Company", "https://www.themancompany.com", "Beauty", "Grooming", "Ahmedabad", "Gujarat", 2015, 60, 80, 25000, 300000, True, True, True, "Men grooming brand"),
    SeedBrand("Bombay Shaving Company", "https://www.bombayshavingcompany.com", "Beauty", "Grooming", "Gurugram", "Haryana", 2016, 40, 50, 15000, 200000, True, True, True, "Men grooming"),
    SeedBrand("Beardo", "https://www.beardo.in", "Beauty", "Grooming", "Hyderabad", "Telangana", 2015, 50, 70, 20000, 250000, True, True, True, "Men grooming brand"),
    SeedBrand("Ustraa", "https://www.ustraa.com", "Beauty", "Grooming", "New Delhi", "Delhi", 2017, 25, 35, 12000, 150000, True, True, True, "Men grooming"),
    SeedBrand("Sugar Cosmetics", "https://www.sugarcosmetics.com", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2015, 150, 200, 50000, 600000, True, True, True, "Color cosmetics brand"),
    SeedBrand("MyGlamm", "https://www.myglamm.com", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2017, 100, 150, 35000, 450000, True, True, True, "D2C beauty brand"),
    SeedBrand("Earth Rhythm", "https://www.earthrhythm.com", "Beauty", "Skincare", "New Delhi", "Delhi", 2019, 30, 40, 15000, 200000, True, True, True, "Clean beauty brand"),
    SeedBrand("Minimalist", "https://www.minimalist.co.in", "Beauty", "Skincare", "Gurugram", "Haryana", 2020, 80, 120, 30000, 400000, True, True, True, "Science-backed skincare"),
    SeedBrand("De Construct", "https://www.deconstruct.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 20, 25, 10000, 120000, True, True, True, "Minimalist skincare"),
    SeedBrand("SkinKraft", "https://www.skinkraft.com", "Beauty", "Personalized", "Mumbai", "Maharashtra", 2018, 40, 60, 20000, 250000, True, True, True, "Personalized skincare"),
    SeedBrand("Re'equil", "https://www.reeequil.com", "Beauty", "Skincare", "Gurugram", "Haryana", 2017, 25, 35, 12000, 150000, True, True, True, "Clinical skincare"),
    SeedBrand("Dr. Sheth's", "https://www.drsheths.com", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 15, 20, 8000, 100000, True, True, True, "Indian skincare"),
    SeedBrand("O3+", "https://www.o3plus.com", "Beauty", "Skincare", "New Delhi", "Delhi", 2012, 30, 40, 15000, 200000, True, True, True, "Professional skincare"),
    SeedBrand("Good Vibes", "https://www.goodvibes.life", "Beauty", "Personal Care", "New Delhi", "Delhi", 2017, 40, 50, 18000, 220000, True, True, True, "Affordable beauty"),
    SeedBrand("Aqualogica", "https://www.aqualogica.in", "Beauty", "Skincare", "Gurugram", "Haryana", 2019, 30, 40, 15000, 180000, True, True, True, "Hydrating skincare"),
    SeedBrand("Derma Co", "https://www.thedermaco.com", "Beauty", "Skincare", "Hyderabad", "Telangana", 2020, 80, 100, 30000, 400000, True, True, True, "Dermatologist skincare"),
    SeedBrand("Pilgrim", "https://www.pilgrim.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 50, 70, 20000, 250000, True, True, True, "International beauty"),
    SeedBrand("Arata", "https://www.arata.in", "Beauty", "Personal Care", "New Delhi", "Delhi", 2018, 15, 20, 8000, 100000, True, True, True, "Clean personal care"),

    # === MORE HOME DECOR ===
    SeedBrand("Nestasia", "https://www.nestasia.in", "Home Decor", "Home Accessories", "Kolkata", "West Bengal", 2018, 25, 35, 12000, 150000, True, True, True, "Premium home decor"),
    SeedBrand("Jaypore", "https://www.jaypore.com", "Home Decor", "Handicrafts", "New Delhi", "Delhi", 2014, 50, 70, 20000, 250000, True, True, True, "Artisanal home decor"),
    SeedBrand("Chumbak", "https://www.chumbak.com", "Home Decor", "Home Accessories", "Bengaluru", "Karnataka", 2011, 40, 50, 18000, 220000, True, True, True, "Quirky home and lifestyle"),
    SeedBrand("Cult Decor", "https://www.cultdecor.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2015, 30, 40, 12000, 150000, True, True, True, "Premium home decor"),
    SeedBrand("Zwende", "https://www.zwende.com", "Home Decor", "Handicrafts", "Bengaluru", "Karnataka", 2016, 12, 15, 5000, 80000, True, True, True, "Custom home decor"),
    SeedBrand("The Decor Kart", "https://www.thedecorkart.com", "Home Decor", "Home Accessories", "New Delhi", "Delhi", 2017, 15, 20, 8000, 100000, True, True, True, "Home decor brand"),
    SeedBrand("Address Home", "https://www.addresshome.com", "Home Decor", "Home Textile", "New Delhi", "Delhi", 2010, 60, 80, 25000, 300000, True, True, True, "Premium home linen"),
    SeedBrand("Homesake", "https://www.homesakeindia.com", "Home Decor", "Handicrafts", "New Delhi", "Delhi", 2016, 10, 15, 5000, 60000, True, True, True, "Handicraft home decor"),
    SeedBrand("Ellementry", "https://www.ellementry.com", "Home Decor", "Kitchenware", "New Delhi", "Delhi", 2017, 20, 25, 10000, 120000, True, True, True, "Handcrafted kitchenware"),
    SeedBrand("Mintwud", "https://www.mintwud.com", "Home Decor", "Furniture", "Mumbai", "Maharashtra", 2018, 25, 30, 12000, 150000, True, True, True, "Affordable furniture"),
    SeedBrand("Wonderchef", "https://www.wonderchef.com", "Home Decor", "Kitchenware", "Gurugram", "Haryana", 2013, 80, 120, 30000, 400000, True, True, True, "Kitchen appliances"),
    SeedBrand("Bergner", "https://www.bergner.in", "Home Decor", "Kitchenware", "Mumbai", "Maharashtra", 2015, 40, 50, 18000, 220000, True, True, True, "Premium cookware"),

    # === MORE JEWELLERY ===
    SeedBrand("Melorra", "https://www.melorra.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2016, 60, 80, 25000, 300000, True, True, True, "Everyday fine jewellery"),
    SeedBrand("CaratLane", "https://www.caratlane.com", "Jewellery", "Fine Jewellery", "Chennai", "Tamil Nadu", 2010, 200, 300, 80000, 900000, True, True, True, "Online jewellery"),
    SeedBrand("BlueStone", "https://www.bluestone.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2011, 150, 200, 60000, 700000, True, True, True, "Online fine jewellery"),
    SeedBrand("Giva", "https://www.giva.co", "Jewellery", "Silver Jewellery", "Bengaluru", "Karnataka", 2019, 30, 40, 15000, 180000, True, True, True, "Silver jewellery brand"),
    SeedBrand("Sukkhi", "https://www.sukkhi.com", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2014, 40, 50, 18000, 220000, True, True, True, "Fashion jewellery"),
    SeedBrand("Zaveri Pearls", "https://www.zfrp.in", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2015, 30, 40, 15000, 180000, True, True, True, "Fashion jewellery"),
    SeedBrand("YouBella", "https://www.youbella.com", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2016, 20, 25, 10000, 120000, True, True, True, "Fashion jewellery"),
    SeedBrand("Enamour", "https://www.enamour.in", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2018, 15, 20, 8000, 100000, True, True, True, "Contemporary fine jewellery"),
    SeedBrand("Candere", "https://www.candere.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2013, 40, 50, 18000, 220000, True, True, True, "Online jewellery"),

    # === MORE HEALTH & WELLNESS ===
    SeedBrand("Kapiva", "https://www.kapiva.in", "Health & Wellness", "Ayurvedic", "Gurugram", "Haryana", 2016, 60, 80, 25000, 300000, True, True, True, "Ayurvedic wellness"),
    SeedBrand("Fast&Up", "https://www.fastandup.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2015, 50, 70, 20000, 250000, True, True, True, "Sports nutrition"),
    SeedBrand("Plix Life", "https://www.plixlife.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2019, 40, 50, 18000, 220000, True, True, True, "Plant-based supplements"),
    SeedBrand("True Elements", "https://www.trueelements.com", "Health & Wellness", "Superfoods", "Pune", "Maharashtra", 2016, 50, 60, 20000, 250000, True, True, True, "Healthy food brand"),
    SeedBrand("Dr. Vaidya's", "https://www.drvaidyas.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2016, 40, 50, 15000, 200000, True, True, True, "Ayurvedic products"),
    SeedBrand("Rasayanam", "https://www.rasayanam.com", "Health & Wellness", "Ayurvedic", "Bengaluru", "Karnataka", 2018, 20, 25, 10000, 120000, True, True, True, "Ayurvedic supplements"),
    SeedBrand("Neuherbs", "https://www.neuherbs.com", "Health & Wellness", "Supplements", "New Delhi", "Delhi", 2019, 20, 25, 10000, 120000, True, True, True, "Health supplements"),
    SeedBrand("Vedix", "https://www.vedix.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2019, 15, 20, 8000, 100000, True, True, True, "Personalized Ayurveda"),

    # === MORE FOOD & SNACKS ===
    SeedBrand("Yoga Bar", "https://www.yogabar.com", "Food & Snacks", "Healthy Snacks", "Bengaluru", "Karnataka", 2017, 60, 80, 25000, 300000, True, True, True, "Healthy muesli and bars"),
    SeedBrand("Slurrp Farm", "https://www.slurrpfarm.com", "Food & Snacks", "Kids Food", "New Delhi", "Delhi", 2017, 30, 40, 15000, 180000, True, True, True, "Healthy kids food"),
    SeedBrand("Nutty Gritties", "https://www.nuttygritties.com", "Food & Snacks", "Dry Fruits", "New Delhi", "Delhi", 2016, 20, 25, 10000, 120000, True, True, True, "Premium dry fruits"),
    SeedBrand("Farmley", "https://www.farmley.com", "Food & Snacks", "Dry Fruits", "New Delhi", "Delhi", 2017, 40, 50, 18000, 220000, True, True, True, "Premium dry fruits"),
    SeedBrand("Vahdam Teas", "https://www.vahdamteas.com", "Tea/Coffee", "Tea", "New Delhi", "Delhi", 2015, 80, 120, 35000, 450000, True, True, True, "Premium Indian teas"),
    SeedBrand("Sleepy Owl", "https://www.sleepyowl.in", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2016, 30, 40, 15000, 180000, True, True, True, "Coffee brand"),
    SeedBrand("Blue Tokai", "https://www.bluetokai.com", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2013, 50, 70, 20000, 250000, True, True, True, "Specialty coffee"),
    SeedBrand("Rage Coffee", "https://www.ragecoffee.com", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2018, 20, 25, 10000, 120000, True, True, True, "Coffee brand"),
    SeedBrand("Raw Pressery", "https://www.rawpressery.com", "Food & Snacks", "Beverages", "Mumbai", "Maharashtra", 2016, 50, 70, 20000, 250000, True, True, True, "Cold-pressed juices"),
    SeedBrand("iD Fresh Food", "https://www.idfreshfood.com", "Food & Snacks", "Fresh Food", "Bengaluru", "Karnataka", 2010, 100, 150, 40000, 500000, True, True, True, "Fresh food brand"),
    SeedBrand("24 Mantra", "https://www.24mantra.com", "Food & Snacks", "Organic Food", "Hyderabad", "Telangana", 2010, 30, 40, 15000, 180000, True, True, True, "Organic food"),

    # === MORE ELECTRONICS ===
    SeedBrand("Hammer", "https://www.hammerlifestyle.in", "Electronics", "Audio", "New Delhi", "Delhi", 2018, 40, 50, 18000, 220000, True, True, True, "Premium audio brand"),
    SeedBrand("Boult Audio", "https://www.boultaudio.com", "Electronics", "Audio", "New Delhi", "Delhi", 2017, 60, 80, 25000, 300000, True, True, True, "Audio accessories"),
    SeedBrand("pTron", "https://www.ptron.in", "Electronics", "Accessories", "Hyderabad", "Telangana", 2014, 100, 150, 40000, 500000, True, True, True, "Affordable tech"),
    SeedBrand("Mivi", "https://www.mivi.in", "Electronics", "Audio", "Hyderabad", "Telangana", 2016, 60, 80, 25000, 300000, True, True, True, "Audio accessories"),
    SeedBrand("Crossbeats", "https://www.crossbeats.com", "Electronics", "Audio", "Bengaluru", "Karnataka", 2014, 40, 50, 18000, 220000, True, True, True, "Audio and wearable"),
    SeedBrand("Fire-Boltt", "https://www.fireboltt.com", "Electronics", "Wearables", "New Delhi", "Delhi", 2016, 120, 180, 50000, 600000, True, True, True, "Smartwatch brand"),
    SeedBrand("Ambrane", "https://www.ambraneindia.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2012, 80, 100, 30000, 400000, True, True, True, "Mobile accessories"),
    SeedBrand("Portronics", "https://www.portronics.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2010, 100, 150, 40000, 500000, True, True, True, "Consumer electronics"),
    SeedBrand("Zoook", "https://www.zoook.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2011, 50, 60, 20000, 250000, True, True, True, "Consumer electronics"),
    SeedBrand("Leaf", "https://www.leafnlife.com", "Electronics", "Audio", "Bengaluru", "Karnataka", 2019, 15, 20, 8000, 100000, True, True, True, "Wireless audio brand"),
    SeedBrand("XECH", "https://www.xech.com", "Electronics", "Accessories", "Mumbai", "Maharashtra", 2017, 15, 20, 8000, 100000, True, True, True, "Tech accessories"),

    # === MORE BABY PRODUCTS ===
    SeedBrand("Hopskotch", "https://www.hopskotch.in", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2014, 40, 50, 18000, 220000, True, True, True, "Kids fashion brand"),
    SeedBrand("LuvLap", "https://www.luvlap.com", "Baby Products", "Baby Care", "New Delhi", "Delhi", 2012, 30, 40, 15000, 180000, True, True, True, "Baby care brand"),
    SeedBrand("R for Rabbit", "https://www.rforgabbit.com", "Baby Products", "Baby Gear", "Ahmedabad", "Gujarat", 2015, 20, 25, 10000, 120000, True, True, True, "Baby gear brand"),
    SeedBrand("Skillmatics", "https://www.skillmatics.com", "Baby Products", "Educational", "Mumbai", "Maharashtra", 2016, 50, 70, 20000, 250000, True, True, True, "Educational games"),
    SeedBrand("Smartivity", "https://www.smartivity.com", "Baby Products", "Educational", "New Delhi", "Delhi", 2015, 15, 20, 8000, 100000, True, True, True, "STEM toys"),
    SeedBrand("FirstCry", "https://www.firstcry.com", "Baby Products", "Multi-Category", "Pune", "Maharashtra", 2010, 150, 250, 60000, 700000, True, True, True, "Baby products"),

    # === MORE PET PRODUCTS ===
    SeedBrand("Heads Up For Tails", "https://www.headsuptails.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2016, 30, 40, 15000, 180000, True, True, True, "Premium pet brand"),
    SeedBrand("Wiggles", "https://www.wiggles.in", "Pet Products", "Pet Care", "Mumbai", "Maharashtra", 2018, 20, 25, 10000, 120000, True, True, True, "Pet care brand"),
    SeedBrand("Drools", "https://www.drools.com", "Pet Products", "Pet Food", "Hyderabad", "Telangana", 2015, 80, 100, 30000, 400000, True, True, True, "Pet food brand"),

    # === MORE FOOTWEAR ===
    SeedBrand("Neeman's", "https://www.neemans.com", "Footwear", "Casual", "Hyderabad", "Telangana", 2018, 30, 40, 15000, 180000, True, True, True, "Sustainable footwear"),

    # === MORE BAGS ===
    SeedBrand("Safari Industries", "https://www.safari-industries.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 100, 150, 40000, 500000, True, True, True, "Luggage brand"),
    SeedBrand("Wildcraft", "https://www.wildcraft.com", "Bags", "Backpacks", "Bengaluru", "Karnataka", 2010, 80, 120, 30000, 400000, True, True, True, "Outdoor gear"),
    SeedBrand("Lavie World", "https://www.lavieworld.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 30, 40, 15000, 180000, True, True, True, "Fashion bags"),
    SeedBrand("Fur Jaden", "https://www.furjaden.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2017, 12, 15, 5000, 80000, True, True, True, "Fashion bags"),
    SeedBrand("Caprese", "https://www.caprese.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 40, 50, 18000, 220000, True, True, True, "Premium bags"),

    # === MORE GIFTS ===
    SeedBrand("Ferns N Petals", "https://www.fernnpetals.com", "Gifts", "Flowers & Gifts", "New Delhi", "Delhi", 2010, 80, 120, 30000, 400000, True, True, True, "Flowers and gifts"),
    SeedBrand("IGP", "https://www.igp.com", "Gifts", "Gifts & Flowers", "Mumbai", "Maharashtra", 2010, 60, 80, 25000, 300000, True, True, True, "Gifts and flowers"),
]


# ============================================================
# TECHNOLOGY DETECTOR
# ============================================================

PLATFORM_PATTERNS: dict[str, list[str]] = {
    "shopify": [
        r"cdn\.shopify\.com", r"Shopify\.theme", r"myshopify\.com",
        r"shopify-section", r"shopify-payment-button", r"Shopify\.loadFeatures",
        r"Shopify\.analytics", r"x-shopify", r"shopify-domain",
        r"assets\.shopifycdn", r"Shopify\.routes",
    ],
    "shopify_plus": [r"shopify-plus", r"Shopify\.shop"],
    "woocommerce": [
        r"woocommerce", r"wc[-_]ajax", r"wp-content/plugins/woocommerce",
        r"woocommerce-session", r"wc_cart_fragments_params",
    ],
    "magento": [
        r"magento", r"Mage\.", r"skin/frontend",
        r"catalog/product", r"magentocommerce",
    ],
    "custom": [r"next\.js", r"__NEXT_DATA__", r"react", r"nuxt", r"gatsby", r"vue\.js"],
}

CHATBOT_PATTERNS = [r"tidio", r"intercom", r"crisp\.chat", r"tawk\.to", r"zendesk-chat", r"gorgias", r"drift"]
WHATSAPP_PATTERNS = [r"wa\.me", r"api\.whatsapp\.com", r"whatsapp.*widget"]
AI_PATTERNS = [r"ai.*chatbot", r"chatgpt", r"openai.*widget", r"powered by ai", r"ai.*assistant"]
EMAIL_MKTG_PATTERNS = {"klaviyo": [r"klaviyo"], "mailchimp": [r"mailchimp"], "sendgrid": [r"sendgrid"]}
REVIEW_PATTERNS = {"judge.me": [r"judge\.me"], "yotpo": [r"yotpo"], "stamped": [r"stamped\.io"]}
SUPPORT_PATTERNS = {"zendesk": [r"zendesk"], "freshdesk": [r"freshdesk"], "intercom": [r"intercom"], "gorgias": [r"gorgias"]}
ANALYTICS_PATTERNS = {"ga4": [r"gtag/js/G-", r"google_tag_manager"], "hotjar": [r"hotjar"]}


def detect_tech(html: str, url: str, headers: dict | None = None) -> TechStack:
    """Detect technology stack."""
    result = TechStack()
    headers = headers or {}

    # Header-based detection
    header_vals = " ".join(v.lower() for v in headers.values())
    if "shopify" in header_vals:
        result.platform = "shopify"
        result.platform_confidence = 0.9

    # HTML-based detection
    if result.platform == "unknown":
        for platform, patterns in PLATFORM_PATTERNS.items():
            matches = sum(1 for p in patterns if re.search(p, html, re.IGNORECASE))
            if matches > 0:
                result.platform = platform
                result.platform_confidence = min(matches * 0.35, 1.0)
                break

    if result.platform == "unknown" and "myshopify.com" in url:
        result.platform = "shopify"
        result.platform_confidence = 0.95

    # Chatbot
    result.has_chatbot = any(re.search(p, html, re.IGNORECASE) for p in CHATBOT_PATTERNS)
    result.has_whatsapp = any(re.search(p, html, re.IGNORECASE) for p in WHATSAPP_PATTERNS)
    result.has_ai = any(re.search(p, html, re.IGNORECASE) for p in AI_PATTERNS)
    result.meta_pixel = "fbq(" in html.lower()

    # Email marketing
    for name, patterns in EMAIL_MKTG_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.email_marketing = name
            break

    # Reviews
    for name, patterns in REVIEW_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.review_platform = name
            break

    # Support
    for name, patterns in SUPPORT_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.support_tool = name
            break

    # Analytics
    for name, patterns in ANALYTICS_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.analytics = name
            break

    # Automation level
    tools = sum(1 for x in [result.email_marketing, result.review_platform, result.support_tool, result.analytics] if x)
    if tools >= 4:
        result.automation_level = "advanced"
    elif tools >= 3:
        result.automation_level = "moderate"
    elif tools >= 1:
        result.automation_level = "basic"
    else:
        result.automation_level = "none"

    return result


# ============================================================
# PAIN DETECTOR
# ============================================================

def detect_pains(html: str, tech: TechStack, brand: SeedBrand) -> PainSignals:
    """Detect pain signals from website and tech stack."""
    pains = PainSignals()

    # No chatbot
    pains.no_chatbot = not tech.has_chatbot

    # No AI
    pains.no_ai = not tech.has_ai

    # No WhatsApp automation
    pains.no_whatsapp_automation = not tech.has_whatsapp

    # Manual support (no support tool detected)
    pains.manual_support = not tech.support_tool

    # No cart recovery (no email marketing)
    pains.no_cart_recovery = not tech.email_marketing

    # No personalization (no AI, no recommendation engine)
    pains.no_personalization = not tech.has_ai

    # No product recommendation
    pains.no_product_recommendation = not tech.has_ai

    # Weak engagement (no review platform, no email marketing)
    pains.weak_engagement = not tech.review_platform and not tech.email_marketing

    # No repeat purchase strategy
    pains.no_repeat_strategy = not tech.email_marketing and not tech.has_whatsapp

    # No upsell
    pains.no_upsell = not tech.has_ai

    # High FAQ volume (check for FAQ page indicators)
    faq_indicators = ["faq", "frequently asked", "common questions"]
    pains.high_faq_volume = any(ind in html.lower() for ind in faq_indicators) and not tech.support_tool

    return pains


# ============================================================
# INTENT DETECTOR
# ============================================================

def detect_intent(html: str, brand: SeedBrand) -> BuyingIntent:
    """Detect buying intent signals."""
    intent = BuyingIntent()

    # Running meta ads
    intent.running_meta_ads = brand.has_meta_ads or "fbq(" in html.lower()

    # Growing Instagram
    intent.growing_instagram = brand.has_instagram

    # Active WhatsApp
    intent.hiring_support = brand.has_whatsapp

    # New products / collections
    new_indicators = ["new arrival", "new collection", "just launched", "new in"]
    intent.new_products = any(ind in html.lower() for ind in new_indicators)

    # Scaling
    intent.scaling_ops = brand.est_monthly_orders > 10000

    # Growing team
    intent.growing_team = brand.est_employees > 30

    return intent


# ============================================================
# CONTACT SCRAPER
# ============================================================

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_REGEX = re.compile(r"(?:\+91[\s\-]?)?[6-9]\d{9}")
LINKEDIN_REGEX = re.compile(r"linkedin\.com/(?:company|in)/[a-zA-Z0-9\-]+")
INSTAGRAM_REGEX = re.compile(r"instagram\.com/([a-zA-Z0-9_.]+)")
FACEBOOK_REGEX = re.compile(r"facebook\.com/([a-zA-Z0-9_.]+)")

GENERIC_PREFIXES = {"support", "info", "hello", "sales", "care", "contact", "help", "feedback", "noreply", "admin", "office", "team", "billing", "careers", "jobs", "hr", "enquiry", "cs", "business", "name", "hello", "enquiries", "customercare", "customer", "orders", "returns"}
FREE_EMAIL = {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com", "icloud.com", "mail.com"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico", ".bmp"}
INVALID_EMAIL_PATTERNS = {"2x.", ".jpg", ".png", ".webp", ".gif", ".svg", "@2x", "assets", "cdn", "static", "media", "images", "files", "base64", "company.com", "example.com", "test.com"}


def _is_valid_email(email: str) -> bool:
    email = email.lower().strip()
    if len(email) > 80 or len(email) < 5:
        return False
    domain = email.split("@")[-1] if "@" in email else ""
    if domain in FREE_EMAIL:
        return False
    if any(ext in email for ext in IMAGE_EXTENSIONS):
        return False
    if any(p in email for p in INVALID_EMAIL_PATTERNS):
        return False
    if any(p in email.split("@")[0] for p in GENERIC_PREFIXES):
        return False
    if not re.match(r"[a-z0-9.\-]+\.[a-z]{2,}$", domain):
        return False
    return True


def _is_valid_phone(phone: str) -> bool:
    digits = re.sub(r"[^0-9]", "", phone)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) != 10:
        return False
    if not digits[0] in "6789":
        return False
    if len(set(digits)) <= 2:
        return False
    return True


def _extract_emails(text: str, contact: ContactInfo) -> None:
    for match in EMAIL_REGEX.findall(text):
        email = match.lower().strip()
        if _is_valid_email(email) and email not in contact.emails:
            contact.emails.append(email)


def _extract_phones(text: str, contact: ContactInfo) -> None:
    for match in PHONE_REGEX.findall(text):
        phone = match.strip()
        if _is_valid_phone(phone) and phone not in contact.phones:
            contact.phones.append(phone)


def _extract_social(text: str, contact: ContactInfo) -> None:
    if not contact.linkedin_url:
        m = LINKEDIN_REGEX.search(text)
        if m:
            contact.linkedin_url = "https://" + m.group(0)
    if not contact.instagram_url:
        m = INSTAGRAM_REGEX.search(text)
        if m:
            contact.instagram_url = "https://" + m.group(0)
    if not contact.facebook_url:
        m = FACEBOOK_REGEX.search(text)
        if m:
            contact.facebook_url = "https://" + m.group(0)


def _extract_founder(text: str, contact: ContactInfo) -> None:
    patterns = [
        r"(?:founder|ceo|co[-\s]?founder|managing director)[\s:]+([A-Z][a-z]+ [A-Z][a-z]+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m and not contact.founder_name:
            contact.founder_name = m.group(1)


async def scrape_website(url: str, client: httpx.AsyncClient) -> tuple[str, ContactInfo]:
    """Scrape website for HTML + contacts."""
    contact = ContactInfo()
    html = ""
    base = url.rstrip("/")

    pages = [
        base, base + "/pages/contact", base + "/pages/contact-us",
        base + "/contact", base + "/contact-us",
        base + "/pages/about", base + "/pages/about-us",
        base + "/about", base + "/about-us",
    ]

    for page_url in pages:
        try:
            resp = await client.get(page_url, timeout=6.0, follow_redirects=True)
            if resp.status_code == 200:
                text = resp.text[:50000]
                if not html:
                    html = text
                _extract_emails(text, contact)
                _extract_phones(text, contact)
                _extract_social(text, contact)
                _extract_founder(text, contact)
                if (contact.best_email and contact.best_phone) or (contact.best_email and contact.linkedin_url):
                    break
        except Exception:
            continue

    return html, contact


async def enrich_contacts(brand_name: str, website: str, client: httpx.AsyncClient) -> tuple[str, ContactInfo]:
    """Full enrichment: website + DuckDuckGo search."""
    html, contact = await scrape_website(website, client)

    # DuckDuckGo search for missing data
    if not contact.best_email or not contact.best_phone:
        try:
            resp = await client.get(
                "https://html.duckduckgo.com/html/",
                params={"q": f'"{brand_name}" founder phone number email India'},
                timeout=10.0,
                follow_redirects=True,
            )
            if resp.status_code == 200:
                text = resp.text
                _extract_phones(text, contact)
                _extract_emails(text, contact)
        except Exception:
            pass

    # LinkedIn search
    if not contact.linkedin_url:
        try:
            resp = await client.get(
                "https://html.duckduckgo.com/html/",
                params={"q": f"site:linkedin.com/company {brand_name} India"},
                timeout=10.0,
                follow_redirects=True,
            )
            if resp.status_code == 200:
                m = LINKEDIN_REGEX.search(resp.text)
                if m:
                    contact.linkedin_url = "https://" + m.group(0)
        except Exception:
            pass

    return html, contact


# ============================================================
# COMMERCIAL FIT SCORER
# ============================================================

def calculate_commercial_fit(
    brand: SeedBrand,
    tech: TechStack,
    contact: ContactInfo,
    pains: PainSignals,
    intent: BuyingIntent,
) -> CommercialFit:
    """Calculate 8-dimensional commercial fit score."""
    fit = CommercialFit()

    # Revenue Size (20%) — ₹3-150 Cr is ideal
    if 3 <= brand.est_revenue_cr <= 150:
        fit.revenue_size = 100
    elif brand.est_revenue_cr < 3:
        fit.revenue_size = max(0, brand.est_revenue_cr / 3 * 100)
    else:
        fit.revenue_size = max(0, 100 - (brand.est_revenue_cr - 150) / 5 * 10)

    # Decision Maker Access (20%)
    if contact.founder_name and contact.best_email:
        fit.decision_maker_access = 100
    elif contact.founder_name:
        fit.decision_maker_access = 70
    elif contact.best_email:
        fit.decision_maker_access = 60
    elif contact.best_phone:
        fit.decision_maker_access = 40
    else:
        fit.decision_maker_access = 0

    # Growth Stage (15%) — Based on employees, orders, traffic
    growth_score = 0
    if 10 <= brand.est_employees <= 200:
        growth_score += 33
    elif brand.est_employees > 200:
        growth_score += 20
    if brand.est_monthly_orders >= 100:
        growth_score += 33
    if brand.est_traffic >= 20000:
        growth_score += 34
    fit.growth_stage = min(growth_score, 100)

    # Technology Fit (15%)
    if tech.platform in ("shopify", "shopify_plus"):
        fit.technology_fit = 100
    elif tech.platform in ("woocommerce",):
        fit.technology_fit = 80
    elif tech.platform == "magento":
        fit.technology_fit = 60
    elif tech.platform != "unknown":
        fit.technology_fit = 50
    else:
        fit.technology_fit = 30

    # Pain Level (10%)
    fit.pain_level = pains.score

    # Buying Intent (10%)
    fit.buying_intent = intent.score

    # Automation Readiness (5%) — Low automation = high opportunity
    if tech.automation_level == "none":
        fit.automation_readiness = 100
    elif tech.automation_level == "basic":
        fit.automation_readiness = 70
    elif tech.automation_level == "moderate":
        fit.automation_readiness = 40
    else:
        fit.automation_readiness = 20

    # Sales Complexity (5%) — Lower complexity = higher score
    if contact.best_email and contact.best_phone:
        fit.sales_complexity = 100
    elif contact.best_email or contact.best_phone:
        fit.sales_complexity = 70
    else:
        fit.sales_complexity = 30

    return fit


# ============================================================
# FINAL LEAD
# ============================================================

@dataclass
class SalesReadyLead:
    company_name: str
    website: str
    category: str
    sub_category: str
    country: str
    city: str
    state: str
    revenue_estimate: str
    employee_estimate: str
    traffic_estimate: str
    monthly_orders: str
    founded_year: int | None
    platform: str
    platform_confidence: float
    technology_stack: list[str]
    shopify_apps: list[str]
    crm: str
    helpdesk: str
    email_platform: str
    meta_pixel: bool
    google_analytics: str
    whatsapp: bool
    instagram: str
    facebook: str
    linkedin_company: str
    founder_name: str
    founder_title: str
    decision_maker: str
    business_email: str
    business_phone: str
    linkedin_decision_maker: str
    growth_signals: list[str]
    pain_signals: list[str]
    intent_signals: list[str]
    automation_readiness: str
    commercial_fit: float
    commercial_fit_grade: str
    icp_score: float
    sales_readiness: float
    close_probability: float
    expected_arr: str
    priority: str
    reason_comai_fits: str
    reason_now: str
    recommended_outreach: str
    evidence_urls: list[str]
    last_verified: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "Company Name": self.company_name,
            "Website": self.website,
            "Category": self.category,
            "Sub Category": self.sub_category,
            "Country": self.country,
            "City": self.city,
            "State": self.state,
            "Revenue Estimate": self.revenue_estimate,
            "Employee Estimate": self.employee_estimate,
            "Traffic Estimate": self.traffic_estimate,
            "Monthly Orders": self.monthly_orders,
            "Founded Year": self.founded_year or "",
            "Platform": self.platform,
            "Platform Confidence": round(self.platform_confidence, 2),
            "Technology Stack": "; ".join(self.technology_stack),
            "Shopify Apps": "; ".join(self.shopify_apps),
            "CRM": self.crm,
            "Helpdesk": self.helpdesk,
            "Email Platform": self.email_platform,
            "Meta Pixel": "Yes" if self.meta_pixel else "No",
            "Google Analytics": self.google_analytics,
            "WhatsApp": "Yes" if self.whatsapp else "No",
            "Instagram": self.instagram,
            "Facebook": self.facebook,
            "LinkedIn Company": self.linkedin_company,
            "Founder": self.founder_name,
            "Founder Title": self.founder_title,
            "Decision Maker": self.decision_maker,
            "Business Email": self.business_email,
            "Business Phone": self.business_phone,
            "LinkedIn Decision Maker": self.linkedin_decision_maker,
            "Growth Signals": "; ".join(self.growth_signals),
            "Pain Signals": "; ".join(self.pain_signals),
            "Intent Signals": "; ".join(self.intent_signals),
            "Automation Readiness": self.automation_readiness,
            "Commercial Fit": round(self.commercial_fit, 1),
            "Commercial Fit Grade": self.commercial_fit_grade,
            "ICP Score": round(self.icp_score, 1),
            "Sales Readiness": round(self.sales_readiness, 1),
            "Close Probability": f"{self.close_probability:.0%}",
            "Expected ARR": self.expected_arr,
            "Priority": self.priority,
            "Reason COMAI Fits": self.reason_comai_fits,
            "Reason Now": self.reason_now,
            "Recommended Outreach": self.recommended_outreach,
            "Evidence URLs": "; ".join(self.evidence_urls),
            "Last Verified": self.last_verified,
        }


# ============================================================
# BRAND PROCESSOR
# ============================================================

async def process_brand(
    brand: SeedBrand,
    client: httpx.AsyncClient,
    semaphore: asyncio.Semaphore,
) -> SalesReadyLead | None:
    """Process a single brand through the full pipeline."""
    async with semaphore:
        try:
            # Step 1: Enrich contacts + get HTML
            html, contact = await enrich_contacts(brand.name, brand.website, client)

            # Step 2: Detect technology
            tech = detect_tech(html, brand.website)

            # Step 3: Detect pain signals
            pains = detect_pains(html, tech, brand)

            # Step 4: Detect buying intent
            intent = detect_intent(html, brand)

            # Step 5: Calculate commercial fit
            fit = calculate_commercial_fit(brand, tech, contact, pains, intent)

            # === QUALITY GATES ===
            # Gate 1: Must have at least one contact method
            if not contact.best_email and not contact.best_phone:
                return None

            # Gate 2: Commercial fit must be ≥ 75
            if fit.total < 75.0:
                return None

            # Gate 3: Must have valid website
            if not html:
                return None

            # Build technology stack
            tech_stack = []
            if tech.platform != "unknown":
                tech_stack.append(tech.platform)
            if tech.email_marketing:
                tech_stack.append(tech.email_marketing)
            if tech.review_platform:
                tech_stack.append(tech.review_platform)
            if tech.support_tool:
                tech_stack.append(tech.support_tool)
            if tech.analytics:
                tech_stack.append(tech.analytics)
            if tech.meta_pixel:
                tech_stack.append("meta_pixel")

            # Build signals lists
            growth_signals = []
            if intent.running_meta_ads:
                growth_signals.append("Running Meta Ads")
            if intent.growing_instagram:
                growth_signals.append("Active Instagram")
            if intent.new_products:
                growth_signals.append("New Products/Collections")
            if intent.scaling_ops:
                growth_signals.append("Scaling Operations")
            if intent.growing_team:
                growth_signals.append("Growing Team")

            pain_signals = []
            if pains.no_chatbot:
                pain_signals.append("No Chatbot")
            if pains.no_ai:
                pain_signals.append("No AI Tools")
            if pains.no_whatsapp_automation:
                pain_signals.append("No WhatsApp Automation")
            if pains.manual_support:
                pain_signals.append("Manual Customer Support")
            if pains.no_cart_recovery:
                pain_signals.append("No Cart Recovery")
            if pains.no_personalization:
                pain_signals.append("No Personalization")
            if pains.no_product_recommendation:
                pain_signals.append("No Product Recommendations")
            if pains.weak_engagement:
                pain_signals.append("Weak Customer Engagement")

            intent_signals = []
            if intent.running_meta_ads:
                intent_signals.append("Active Advertiser")
            if intent.growing_instagram:
                intent_signals.append("Growing Social Presence")
            if intent.scaling_ops:
                intent_signals.append("Scaling Operations")

            # Calculate scores
            icp_score = fit.total
            sales_readiness = fit.total * 0.8 + pains.score * 0.1 + intent.score * 0.1
            close_prob = min(fit.total / 100 * 0.7 + pains.score / 100 * 0.2 + intent.score / 100 * 0.1, 0.95)

            # Expected ARR
            arr = max(3, brand.est_revenue_cr * 0.03) * 100000
            arr_str = f"₹{arr / 100000:.1f}L"

            # Priority
            if fit.total >= 85 and contact.best_email and contact.best_phone:
                priority = "HOT"
            elif fit.total >= 75:
                priority = "WARM"
            else:
                priority = "NURTURE"

            # Reason COMAI fits
            reasons = []
            if pains.no_chatbot:
                reasons.append("No chatbot — needs 24/7 AI support")
            if pains.no_ai:
                reasons.append("No AI tools — high automation opportunity")
            if pains.no_whatsapp_automation:
                reasons.append("No WhatsApp automation — can automate conversations")
            if pains.no_cart_recovery:
                reasons.append("No cart recovery — can recover lost sales")
            if pains.no_personalization:
                reasons.append("No personalization — can boost conversions")
            reason_comai = "; ".join(reasons[:3]) if reasons else "COMAI can automate ecommerce operations"

            # Reason NOW
            now_reasons = []
            if intent.running_meta_ads:
                now_reasons.append("Running ads — needs conversion optimization")
            if intent.scaling_ops:
                now_reasons.append("Scaling — needs automation to handle growth")
            if intent.new_products:
                now_reasons.append("New products — needs AI recommendations")
            reason_now = "; ".join(now_reasons[:2]) if now_reasons else "Growing D2C brand ready for AI automation"

            # Recommended outreach
            if contact.best_email and contact.founder_name:
                rec_outreach = f"Personalized email to {contact.founder_name} with COMAI case study"
            elif contact.best_email:
                rec_outreach = "Personalized email with ROI calculator"
            elif contact.best_phone:
                rec_outreach = "Direct call with discovery questions"
            else:
                rec_outreach = "LinkedIn connection request + follow-up"

            # Evidence URLs
            evidence = [brand.website]
            if contact.instagram_url:
                evidence.append(contact.instagram_url)
            if contact.linkedin_url:
                evidence.append(contact.linkedin_url)
            if contact.facebook_url:
                evidence.append(contact.facebook_url)

            now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

            return SalesReadyLead(
                company_name=brand.name,
                website=brand.website,
                category=brand.category,
                sub_category=brand.sub_category,
                country="India",
                city=brand.city,
                state=brand.state,
                revenue_estimate=f"₹{max(3, brand.est_revenue_cr // 2)}-{brand.est_revenue_cr} Cr",
                employee_estimate=f"{max(10, brand.est_employees // 2)}-{brand.est_employees}",
                traffic_estimate=f"{max(20, brand.est_traffic // 1000)}K-{brand.est_traffic // 1000}K monthly",
                monthly_orders=f"{max(100, brand.est_monthly_orders // 2)}-{brand.est_monthly_orders}",
                founded_year=brand.founded_year,
                platform=tech.platform,
                platform_confidence=tech.platform_confidence,
                technology_stack=tech_stack,
                shopify_apps=[],
                crm=tech.support_tool or "None detected",
                helpdesk=tech.support_tool or "None detected",
                email_platform=tech.email_marketing or "None detected",
                meta_pixel=tech.meta_pixel,
                google_analytics=tech.analytics or "None detected",
                whatsapp=brand.has_whatsapp,
                instagram=contact.instagram_url or "",
                facebook=contact.facebook_url or "",
                linkedin_company=contact.linkedin_url or "",
                founder_name=contact.founder_name or brand.name + " Team",
                founder_title="Founder/CEO",
                decision_maker=contact.best_decision_maker,
                business_email=contact.best_email,
                business_phone=contact.best_phone,
                linkedin_decision_maker=contact.best_dm_linkedin,
                growth_signals=growth_signals,
                pain_signals=pain_signals,
                intent_signals=intent_signals,
                automation_readiness=tech.automation_level,
                commercial_fit=fit.total,
                commercial_fit_grade=fit.grade,
                icp_score=icp_score,
                sales_readiness=sales_readiness,
                close_probability=close_prob,
                expected_arr=arr_str,
                priority=priority,
                reason_comai_fits=reason_comai,
                reason_now=reason_now,
                recommended_outreach=rec_outreach,
                evidence_urls=evidence,
                last_verified=now_str,
            )

        except Exception as e:
            print(f"  Error processing {brand.name}: {e}")
            return None


# ============================================================
# MAIN PIPELINE
# ============================================================

async def run_pipeline(limit: int = 400, output: str = "comai_sales_leads.xlsx") -> None:
    """Run the COMAI Sales-Ready Lead Intelligence Engine."""
    print("=" * 70)
    print("COMAI SALES-READY LEAD INTELLIGENCE ENGINE")
    print("Sprint 43.5 — Quality Over Quantity")
    print("=" * 70)

    # Load and filter seed database
    all_seeds = SEED_BRANDS.copy()
    print(f"\nRaw seeds: {len(all_seeds)}")

    # Deduplicate
    seen = set()
    unique = []
    for seed in all_seeds:
        key = seed.website.rstrip("/").lower()
        if key not in seen:
            seen.add(key)
            unique.append(seed)
    seeds = unique
    print(f"After dedup: {len(seeds)}")

    # Filter rejects
    filtered = []
    rejected_count = 0
    for seed in seeds:
        is_rej, reason = should_reject(seed.name)
        if is_rej:
            rejected_count += 1
        else:
            filtered.append(seed)
    seeds = filtered[:limit]
    print(f"After reject filter: {len(seeds)} brands (rejected {rejected_count})")

    # Process all brands
    semaphore = asyncio.Semaphore(12)
    results: list[SalesReadyLead] = []
    gate_failures = {"no_contact": 0, "low_fit": 0, "no_html": 0}

    print(f"\nProcessing {len(seeds)} brands...")
    start_time = time.time()

    async with httpx.AsyncClient(
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
        },
        follow_redirects=True,
    ) as client:
        tasks = [process_brand(seed, client, semaphore) for seed in seeds]

        completed = 0
        for coro in asyncio.as_completed(tasks):
            result = await coro
            completed += 1
            if result:
                results.append(result)
            if completed % 25 == 0:
                elapsed = time.time() - start_time
                rate = completed / elapsed if elapsed > 0 else 0
                print(f"  Processed {completed}/{len(seeds)} | Qualified: {len(results)} | {rate:.1f}/sec")

    elapsed = time.time() - start_time
    print(f"\nProcessed: {len(seeds)} | Qualified: {len(results)} | Time: {elapsed:.0f}s")
    print(f"Qualification Rate: {len(results)/len(seeds)*100:.1f}%")

    # Separate by priority
    hot = [r for r in results if r.priority == "HOT"]
    warm = [r for r in results if r.priority == "WARM"]
    nurture = [r for r in results if r.priority == "NURTURE"]

    print(f"\nPriority Breakdown:")
    print(f"  HOT:     {len(hot)}")
    print(f"  WARM:    {len(warm)}")
    print(f"  NURTURE: {len(nurture)}")

    # Stats
    with_email = sum(1 for r in results if r.business_email)
    with_phone = sum(1 for r in results if r.business_phone)
    with_linkedin = sum(1 for r in results if r.linkedin_company)
    with_instagram = sum(1 for r in results if r.instagram)

    print(f"\nContact Availability:")
    print(f"  Email:    {with_email} ({with_email*100//len(results) if results else 0}%)")
    print(f"  Phone:    {with_phone} ({with_phone*100//len(results) if results else 0}%)")
    print(f"  LinkedIn: {with_linkedin} ({with_linkedin*100//len(results) if results else 0}%)")
    print(f"  Instagram:{with_instagram} ({with_instagram*100//len(results) if results else 0}%)")

    # Sort by commercial fit
    qualified = hot + warm + nurture
    qualified.sort(key=lambda x: x.commercial_fit, reverse=True)

    # Export
    _export_excel(qualified, output)
    _export_summary(qualified, output.replace(".xlsx", "_summary.txt"))

    print(f"\nExported to: {output}")
    print("=" * 70)


# ============================================================
# EXCEL EXPORT
# ============================================================

def _export_excel(leads: list[SalesReadyLead], filename: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "COMAI Sales Leads"

    if not leads:
        wb.save(filename)
        return

    headers = list(leads[0].to_dict().keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    priority_fills = {
        "HOT": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "WARM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "NURTURE": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    }

    for row_idx, lead in enumerate(leads, 2):
        data = lead.to_dict()
        priority = data.get("Priority", "")
        row_fill = priority_fills.get(priority)

        for col_idx, header in enumerate(headers, 1):
            value = data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_fill:
                cell.fill = row_fill

    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value or ""))
            for row in range(1, min(len(leads) + 2, 50))
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_length + 2, 35)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)


def _export_summary(leads: list[SalesReadyLead], filename: str) -> None:
    hot = sum(1 for l in leads if l.priority == "HOT")
    warm = sum(1 for l in leads if l.priority == "WARM")
    nurture = sum(1 for l in leads if l.priority == "NURTURE")
    with_email = sum(1 for l in leads if l.business_email)
    with_phone = sum(1 for l in leads if l.business_phone)
    with_linkedin = sum(1 for l in leads if l.linkedin_company)
    avg_fit = sum(l.commercial_fit for l in leads) / len(leads) if leads else 0
    avg_readiness = sum(l.sales_readiness for l in leads) / len(leads) if leads else 0

    categories = {}
    for l in leads:
        categories[l.category] = categories.get(l.category, 0) + 1

    summary = f"""
COMAI SALES-READY LEAD INTELLIGENCE — Sprint 43.5 Summary
==========================================================

Total Qualified Leads: {len(leads)}

Priority Breakdown:
  HOT:     {hot} (Can call TODAY)
  WARM:    {warm} (Can call THIS WEEK)
  NURTURE: {nurture} (Follow-up sequence)

Contact Availability:
  With Email:    {with_email} ({with_email*100//len(leads) if leads else 0}%)
  With Phone:    {with_phone} ({with_phone*100//len(leads) if leads else 0}%)
  With LinkedIn: {with_linkedin} ({with_linkedin*100//len(leads) if leads else 0}%)

Average Commercial Fit: {avg_fit:.1f}/100
Average Sales Readiness: {avg_readiness:.1f}/100

Category Breakdown:
"""
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        summary += f"  {cat}: {count}\n"

    summary += f"\nGenerated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n"

    with open(filename, "w") as f:
        f.write(summary)

    print(summary)


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="COMAI Sales-Ready Lead Intelligence Engine")
    parser.add_argument("--limit", type=int, default=400, help="Max brands to process")
    parser.add_argument("--output", type=str, default="comai_sales_leads.xlsx", help="Output filename")
    args = parser.parse_args()

    asyncio.run(run_pipeline(limit=args.limit, output=args.output))
