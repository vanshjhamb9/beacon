"""Export LinkedIn Sales Validation to XLSX format."""

from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PROJECT_ROOT = Path(__file__).parent
EXPORTS_DIR = PROJECT_ROOT / "exports"


def load_json(filename: str) -> dict[str, Any]:
    """Load JSON file."""
    filepath = EXPORTS_DIR / filename
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def export_linkedin_validation_xlsx() -> None:
    """Export LinkedIn Sales Validation to XLSX."""
    if not HAS_OPENPYXL:
        print("openpyxl not available. Skipping XLSX export.")
        return

    data = load_json("linkedin_sales_validation.json")
    cards = data.get("sales_intelligence_cards", [])

    wb = Workbook()

    # Sheet 1: Sales Intelligence Cards
    ws = wb.active
    ws.title = "Sales Intelligence Cards"

    # Headers
    headers = [
        "Rank", "Company", "Requirement", "Intent Score", "Intent Level",
        "Outsourcing Fit", "Decision Maker", "DM Role", "DM Confidence",
        "LinkedIn", "Why This Person", "Why Now", "Service Match",
        "Recommended Pitch", "Quality Gate", "Outreach State",
        "Connection Request", "Follow-up #1", "Follow-up #2",
        "Likely Objection", "Objection Response", "Recommended CTA"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row, card in enumerate(cards, 2):
        dm = card.get("decision_maker", {})
        msg = card.get("linkedin_message", {})
        
        row_data = [
            row - 1,  # Rank
            card.get("company"),
            card.get("requirement"),
            card.get("intent_score"),
            card.get("intent_level"),
            card.get("outsourcing_fit"),
            dm.get("name"),
            dm.get("role"),
            dm.get("confidence"),
            dm.get("linkedin_url"),
            dm.get("why_this_person"),
            card.get("why_now"),
            card.get("service_match"),
            card.get("recommended_pitch"),
            card.get("quality_gate_status"),
            card.get("outreach_state"),
            msg.get("connection_request"),
            msg.get("follow_up_1"),
            msg.get("follow_up_2"),
            card.get("likely_objection"),
            card.get("objection_response"),
            card.get("recommended_cta"),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Color code by outreach state
        outreach_state = card.get("outreach_state")
        if outreach_state == "LINKEDIN_DRAFT_READY":
            fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = fill

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 25

    # Sheet 2: Tracking Records
    ws_tracking = wb.create_sheet("Tracking Records")
    
    tracking_headers = [
        "Company", "Opportunity ID", "Approved At", "Approved By",
        "Sent At", "Sent By", "Connection Status", "Reply Status",
        "Reply Date", "Reply Type", "Meeting Booked", "Meeting Date",
        "Outcome", "Notes"
    ]
    
    for col, header in enumerate(tracking_headers, 1):
        cell = ws_tracking.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    tracking = data.get("tracking_records", [])
    for row, record in enumerate(tracking, 2):
        track_data = [
            record.get("company"),
            record.get("opportunity_id"),
            record.get("approved_at"),
            record.get("approved_by"),
            record.get("sent_at"),
            record.get("sent_by"),
            record.get("connection_status"),
            record.get("reply_status"),
            record.get("reply_date"),
            record.get("reply_type"),
            record.get("meeting_booked"),
            record.get("meeting_date"),
            record.get("outcome"),
            record.get("notes"),
        ]
        
        for col, value in enumerate(track_data, 1):
            cell = ws_tracking.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths for tracking sheet
    for col in range(1, len(tracking_headers) + 1):
        ws_tracking.column_dimensions[chr(64 + col)].width = 20

    # Save
    output_file = EXPORTS_DIR / "linkedin_sales_validation.xlsx"
    wb.save(output_file)
    print(f"Saved LinkedIn Sales Validation to {output_file}")


def main():
    """Export LinkedIn Sales Validation to XLSX."""
    print("Exporting LinkedIn Sales Validation to XLSX...")
    export_linkedin_validation_xlsx()
    print("Export complete!")


if __name__ == "__main__":
    main()
