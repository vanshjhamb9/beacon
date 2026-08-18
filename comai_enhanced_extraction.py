"""
COMAI Enhanced Lead Extraction — Deep Contact Discovery
=======================================================
Builds on comai_sales_engine.py with:
  - Deeper website crawling (12+ pages per brand)
  - Aggressive DuckDuckGo search for founders
  - WhatsApp link detection
  - Team/about page parsing
  - Lower contact gate (phone-only qualifies)
"""

from __future__ import annotations

import asyncio
import re
import time
import argparse
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

import httpx

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

# Reuse models from sales engine
from comai_sales_engine import (
    SeedBrand, SEED_BRANDS, REJECT_COMPANIES, REJECT_KEYWORDS,
    TechStack, PainSignals, BuyingIntent, CommercialFit,
    detect_tech, detect_pains, detect_intent, calculate_commercial_fit,
    should_reject, SalesReadyLead,
)


# ============================================================
# ENHANCED CONTACT SCRAPER
# ============================================================

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_REGEX = re.compile(r"(?:\+91[\s\-]?)?[6-9]\d{9}")
LINKEDIN_REGEX = re.compile(r"linkedin\.com/(?:company|in)/[a-zA-Z0-9\-]+")
INSTAGRAM_REGEX = re.compile(r"instagram\.com/([a-zA-Z0-9_.]+)")
FACEBOOK_REGEX = re.compile(r"facebook\.com/([a-zA-Z0-9_.]+)")
WHATSAPP_REGEX = re.compile(r"wa\.me/(\d{10,15})|api\.whatsapp\.com/send\?phone=(\d{10,15})")

GENERIC_PREFIXES = {"support", "info", "hello", "sales", "care", "contact", "help",
                    "feedback", "noreply", "admin", "office", "team", "billing",
                    "careers", "jobs", "hr", "enquiry", "cs", "business", "name",
                    "enquiries", "customercare", "customer", "orders", "returns"}
FREE_EMAIL = {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com",
              "icloud.com", "mail.com", "rediffmail.com", "live.com", "msn.com"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico", ".bmp"}
INVALID_EMAIL_PATTERNS = {"2x.", ".jpg", ".png", ".webp", ".gif", ".svg", "@2x",
                          "assets", "cdn", "static", "media", "images", "files",
                          "base64", "company.com", "example.com", "test.com"}


@dataclass
class EnhancedContact:
    emails: list[str] = field(default_factory=list)
    phones: list[str] = field(default_factory=list)
    whatsapp_numbers: list[str] = field(default_factory=list)
    founder_name: str = ""
    founder_title: str = ""
    founder_email: str = ""
    founder_linkedin: str = ""
    ceo_name: str = ""
    ceo_email: str = ""
    linkedin_url: str = ""
    instagram_url: str = ""
    facebook_url: str = ""
    twitter_url: str = ""

    @property
    def best_email(self) -> str:
        if self.founder_email:
            return self.founder_email
        if self.ceo_email:
            return self.ceo_email
        for e in self.emails:
            prefix = e.split("@")[0].lower()
            if prefix not in GENERIC_PREFIXES:
                return e
        return self.emails[0] if self.emails else ""

    @property
    def best_phone(self) -> str:
        return self.phones[0] if self.phones else (
            self.whatsapp_numbers[0] if self.whatsapp_numbers else ""
        )

    @property
    def best_decision_maker(self) -> str:
        return self.founder_name or self.ceo_name or ""

    @property
    def best_dm_linkedin(self) -> str:
        return self.founder_linkedin or ""


def _is_valid_email(email: str) -> bool:
    email = email.lower().strip()
    if len(email) > 80 or len(email) < 5:
        return False
    domain = email.split("@")[-1] if "@" in email else ""
    if domain in FREE_EMAIL:
        return False
    if any(ext in email for ext in IMAGE_EXTENSIONS):
        return False
    if any(p in email for p in INVALID_EMAIL_PATTERNS):
        return False
    if any(p == email.split("@")[0].lower() for p in GENERIC_PREFIXES):
        return False
    if not re.match(r"[a-z0-9.\-]+\.[a-z]{2,}$", domain):
        return False
    return True


def _is_valid_phone(phone: str) -> bool:
    digits = re.sub(r"[^0-9]", "", phone)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) != 10:
        return False
    if not digits[0] in "6789":
        return False
    if len(set(digits)) <= 2:
        return False
    return True


def _extract_from_text(text: str, contact: EnhancedContact) -> None:
    """Extract all contact info from text."""
    # Emails
    for match in EMAIL_REGEX.findall(text):
        email = match.lower().strip()
        if _is_valid_email(email) and email not in contact.emails:
            contact.emails.append(email)

    # Phones
    for match in PHONE_REGEX.findall(text):
        phone = match.strip()
        if _is_valid_phone(phone) and phone not in contact.phones:
            contact.phones.append(phone)

    # WhatsApp
    for match in WHATSAPP_REGEX.finditer(text):
        number = match.group(1) or match.group(2)
        if number and _is_valid_phone(number) and number not in contact.whatsapp_numbers:
            contact.whatsapp_numbers.append(number)

    # LinkedIn
    if not contact.linkedin_url:
        m = LINKEDIN_REGEX.search(text)
        if m:
            contact.linkedin_url = "https://" + m.group(0)

    # Instagram
    if not contact.instagram_url:
        m = INSTAGRAM_REGEX.search(text)
        if m:
            contact.instagram_url = "https://" + m.group(0)

    # Facebook
    if not contact.facebook_url:
        m = FACEBOOK_REGEX.search(text)
        if m:
            contact.facebook_url = "https://" + m.group(0)

    # Founder/CEO detection
    founder_patterns = [
        r"(?:founder|co[-\s]?founder|ceo|managing director|md)\s*[:|\-]\s*([A-Z][a-z]+ (?:[A-Z]\.?\s)?[A-Z][a-z]+)",
        r"(?:Founder|CEO|Co-Founder|Managing Director)\s+([A-Z][a-z]+ (?:[A-Z]\.?\s)?[A-Z][a-z]+)",
    ]
    for pattern in founder_patterns:
        m = re.search(pattern, text)
        if m:
            name = m.group(1).strip()
            if len(name) > 3 and len(name) < 40:
                if "founder" in pattern.lower() or "ceo" in pattern.lower():
                    if not contact.founder_name:
                        contact.founder_name = name
                        contact.founder_title = "Founder/CEO" if "founder" in pattern.lower() else "CEO"
                break


async def deep_scrape_brand(brand_name: str, website: str, client: httpx.AsyncClient) -> tuple[str, EnhancedContact]:
    """Deep scrape with 12+ pages per brand."""
    contact = EnhancedContact()
    html = ""
    base = website.rstrip("/")

    # Extended page list
    pages = [
        base,
        base + "/pages/contact",
        base + "/pages/contact-us",
        base + "/contact",
        base + "/contact-us",
        base + "/pages/about",
        base + "/pages/about-us",
        base + "/about",
        base + "/about-us",
        base + "/pages/our-story",
        base + "/team",
        base + "/pages/team",
        base + "/pages/shipping-policy",  # Often has email
        base + "/policies/terms-of-service",  # Often has email
    ]

    for page_url in pages:
        try:
            resp = await client.get(page_url, timeout=8.0, follow_redirects=True)
            if resp.status_code == 200:
                text = resp.text[:80000]
                if not html:
                    html = text
                _extract_from_text(text, contact)
                # Early exit if we have both email and phone
                if contact.best_email and contact.best_phone:
                    break
        except Exception:
            continue

    return html, contact


async def search_founder_contacts(brand_name: str, client: httpx.AsyncClient, contact: EnhancedContact) -> None:
    """Search DuckDuckGo for founder contacts."""
    queries = [
        f'"{brand_name}" founder email phone India',
        f'"{brand_name}" CEO contact email',
        f'site:linkedin.com/company "{brand_name}" India',
    ]

    for query in queries:
        if contact.best_email and contact.best_phone:
            break
        try:
            resp = await client.get(
                "https://html.duckduckgo.com/html/",
                params={"q": query},
                timeout=12.0,
                follow_redirects=True,
            )
            if resp.status_code == 200:
                _extract_from_text(resp.text, contact)
        except Exception:
            pass
        await asyncio.sleep(0.5)


async def process_brand_enhanced(
    brand: SeedBrand,
    client: httpx.AsyncClient,
    semaphore: asyncio.Semaphore,
) -> SalesReadyLead | None:
    """Process a brand with enhanced contact discovery."""
    async with semaphore:
        try:
            # Step 1: Deep scrape
            html, contact = await deep_scrape_brand(brand.name, brand.website, client)

            # Step 2: Search for contacts if missing
            if not contact.best_email or not contact.best_phone:
                await search_founder_contacts(brand.name, client, contact)

            # Step 3: Detect tech, pains, intent
            tech = detect_tech(html, brand.website)
            pains = detect_pains(html, tech, brand)
            intent = detect_intent(html, brand)
            fit = calculate_commercial_fit(brand, tech, contact, pains, intent)

            # QUALITY GATES (relaxed)
            # Gate 1: Must have at least one contact method
            if not contact.best_email and not contact.best_phone and not contact.whatsapp_numbers:
                return None

            # Gate 2: Commercial fit >= 70 (relaxed from 75)
            if fit.total < 70.0:
                return None

            # Gate 3: Must have HTML
            if not html:
                return None

            # Build tech stack
            tech_stack = []
            if tech.platform != "unknown":
                tech_stack.append(tech.platform)
            if tech.email_marketing:
                tech_stack.append(tech.email_marketing)
            if tech.review_platform:
                tech_stack.append(tech.review_platform)
            if tech.support_tool:
                tech_stack.append(tech.support_tool)
            if tech.analytics:
                tech_stack.append(tech.analytics)
            if tech.meta_pixel:
                tech_stack.append("meta_pixel")

            # Build signals
            growth_signals = []
            if intent.running_meta_ads:
                growth_signals.append("Running Meta Ads")
            if intent.growing_instagram:
                growth_signals.append("Active Instagram")
            if intent.new_products:
                growth_signals.append("New Products/Collections")
            if intent.scaling_ops:
                growth_signals.append("Scaling Operations")
            if intent.growing_team:
                growth_signals.append("Growing Team")

            pain_signals = []
            if pains.no_chatbot:
                pain_signals.append("No Chatbot")
            if pains.no_ai:
                pain_signals.append("No AI Tools")
            if pains.no_whatsapp_automation:
                pain_signals.append("No WhatsApp Automation")
            if pains.manual_support:
                pain_signals.append("Manual Customer Support")
            if pains.no_cart_recovery:
                pain_signals.append("No Cart Recovery")
            if pains.no_personalization:
                pain_signals.append("No Personalization")

            intent_signals = []
            if intent.running_meta_ads:
                intent_signals.append("Active Advertiser")
            if intent.growing_instagram:
                intent_signals.append("Growing Social Presence")
            if intent.scaling_ops:
                intent_signals.append("Scaling Operations")

            # Scores
            sales_readiness = fit.total * 0.8 + pains.score * 0.1 + intent.score * 0.1
            close_prob = min(fit.total / 100 * 0.7 + pains.score / 100 * 0.2 + intent.score / 100 * 0.1, 0.95)
            arr = max(3, brand.est_revenue_cr * 0.03) * 100000

            # Priority
            if fit.total >= 85 and (contact.best_email or contact.best_phone):
                priority = "HOT"
            elif fit.total >= 75:
                priority = "WARM"
            else:
                priority = "NURTURE"

            # Reason COMAI fits
            reasons = []
            if pains.no_chatbot:
                reasons.append("No chatbot - needs 24/7 AI support")
            if pains.no_ai:
                reasons.append("No AI tools - high automation opportunity")
            if pains.no_whatsapp_automation:
                reasons.append("No WhatsApp automation")
            if pains.no_cart_recovery:
                reasons.append("No cart recovery")
            reason_comai = "; ".join(reasons[:3]) if reasons else "COMAI can automate ecommerce"

            # Reason NOW
            now_reasons = []
            if intent.running_meta_ads:
                now_reasons.append("Running ads - needs conversion optimization")
            if intent.scaling_ops:
                now_reasons.append("Scaling - needs automation")
            if intent.new_products:
                now_reasons.append("New products - needs AI recommendations")
            reason_now = "; ".join(now_reasons[:2]) if now_reasons else "Growing D2C brand"

            # Outreach
            if contact.best_email and contact.best_decision_maker:
                rec_outreach = f"Personalized email to {contact.best_decision_maker}"
            elif contact.best_email:
                rec_outreach = "Personalized email with ROI calculator"
            elif contact.best_phone:
                rec_outreach = "Direct call with discovery questions"
            else:
                rec_outreach = "LinkedIn + follow-up"

            evidence = [brand.website]
            if contact.instagram_url:
                evidence.append(contact.instagram_url)
            if contact.linkedin_url:
                evidence.append(contact.linkedin_url)

            now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

            return SalesReadyLead(
                company_name=brand.name,
                website=brand.website,
                category=brand.category,
                sub_category=brand.sub_category,
                country="India",
                city=brand.city,
                state=brand.state,
                revenue_estimate=f"₹{max(3, brand.est_revenue_cr // 2)}-{brand.est_revenue_cr} Cr",
                employee_estimate=f"{max(10, brand.est_employees // 2)}-{brand.est_employees}",
                traffic_estimate=f"{max(20, brand.est_traffic // 1000)}K-{brand.est_traffic // 1000}K monthly",
                monthly_orders=f"{max(100, brand.est_monthly_orders // 2)}-{brand.est_monthly_orders}",
                founded_year=brand.founded_year,
                platform=tech.platform,
                platform_confidence=tech.platform_confidence,
                technology_stack=tech_stack,
                shopify_apps=[],
                crm=tech.support_tool or "None detected",
                helpdesk=tech.support_tool or "None detected",
                email_platform=tech.email_marketing or "None detected",
                meta_pixel=tech.meta_pixel,
                google_analytics=tech.analytics or "None detected",
                whatsapp=bool(contact.whatsapp_numbers),
                instagram=contact.instagram_url or "",
                facebook=contact.facebook_url or "",
                linkedin_company=contact.linkedin_url or "",
                founder_name=contact.founder_name or brand.name + " Team",
                founder_title=contact.founder_title or "Founder/CEO",
                decision_maker=contact.best_decision_maker,
                business_email=contact.best_email,
                business_phone=contact.best_phone,
                linkedin_decision_maker=contact.best_dm_linkedin,
                growth_signals=growth_signals,
                pain_signals=pain_signals,
                intent_signals=intent_signals,
                automation_readiness=tech.automation_level,
                commercial_fit=fit.total,
                commercial_fit_grade=fit.grade,
                icp_score=fit.total,
                sales_readiness=sales_readiness,
                close_probability=close_prob,
                expected_arr=f"₹{arr / 100000:.1f}L",
                priority=priority,
                reason_comai_fits=reason_comai,
                reason_now=reason_now,
                recommended_outreach=rec_outreach,
                evidence_urls=evidence,
                last_verified=now_str,
            )

        except Exception as e:
            print(f"  Error processing {brand.name}: {e}")
            return None


# ============================================================
# MAIN
# ============================================================

async def run_enhanced_pipeline(limit: int = 500, output: str = "comai_enhanced_leads.xlsx") -> None:
    print("=" * 70)
    print("COMAI ENHANCED LEAD EXTRACTION — Deep Contact Discovery")
    print("=" * 70)

    # Deduplicate seeds
    seen = set()
    unique = []
    for seed in SEED_BRANDS:
        key = seed.website.rstrip("/").lower()
        if key not in seen:
            seen.add(key)
            unique.append(seed)
    seeds = unique
    print(f"Deduplicated seeds: {len(seeds)}")

    # Filter rejects
    filtered = []
    for seed in seeds:
        is_rej, _ = should_reject(seed.name)
        if not is_rej:
            filtered.append(seed)
    seeds = filtered[:limit]
    print(f"After reject filter: {len(seeds)} brands")

    # Process
    semaphore = asyncio.Semaphore(15)
    results = []
    start_time = time.time()

    async with httpx.AsyncClient(
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
        },
        follow_redirects=True,
        verify=False,
    ) as client:
        tasks = [process_brand_enhanced(seed, client, semaphore) for seed in seeds]

        completed = 0
        for coro in asyncio.as_completed(tasks):
            result = await coro
            completed += 1
            if result:
                results.append(result)
            if completed % 25 == 0:
                elapsed = time.time() - start_time
                rate = completed / elapsed if elapsed > 0 else 0
                print(f"  Processed {completed}/{len(seeds)} | Qualified: {len(results)} | {rate:.1f}/sec")

    elapsed = time.time() - start_time
    print(f"\nProcessed: {len(seeds)} | Qualified: {len(results)} | Time: {elapsed:.0f}s")
    print(f"Qualification Rate: {len(results)/len(seeds)*100:.1f}%")

    # Stats
    hot = [r for r in results if r.priority == "HOT"]
    warm = [r for r in results if r.priority == "WARM"]
    nurture = [r for r in results if r.priority == "NURTURE"]

    print(f"\nPriority Breakdown:")
    print(f"  HOT:     {len(hot)}")
    print(f"  WARM:    {len(warm)}")
    print(f"  NURTURE: {len(nurture)}")

    with_email = sum(1 for r in results if r.business_email)
    with_phone = sum(1 for r in results if r.business_phone)
    with_linkedin = sum(1 for r in results if r.linkedin_company)
    with_whatsapp = sum(1 for r in results if r.whatsapp)

    print(f"\nContact Availability:")
    print(f"  Email:    {with_email} ({with_email*100//len(results) if results else 0}%)")
    print(f"  Phone:    {with_phone} ({with_phone*100//len(results) if results else 0}%)")
    print(f"  WhatsApp: {with_whatsapp} ({with_whatsapp*100//len(results) if results else 0}%)")
    print(f"  LinkedIn: {with_linkedin} ({with_linkedin*100//len(results) if results else 0}%)")

    # Sort by fit
    qualified = hot + warm + nurture
    qualified.sort(key=lambda x: x.commercial_fit, reverse=True)

    # Export
    _export_excel(qualified, output)
    _export_summary(qualified, output.replace(".xlsx", "_summary.txt"))

    print(f"\nExported to: {output}")
    print("=" * 70)


def _export_excel(leads: list[SalesReadyLead], filename: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "COMAI Enhanced Leads"

    if not leads:
        wb.save(filename)
        return

    headers = list(leads[0].to_dict().keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    priority_fills = {
        "HOT": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "WARM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "NURTURE": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    }

    for row_idx, lead in enumerate(leads, 2):
        data = lead.to_dict()
        priority = data.get("Priority", "")
        row_fill = priority_fills.get(priority)
        for col_idx, header in enumerate(headers, 1):
            value = data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_fill:
                cell.fill = row_fill

    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value or ""))
            for row in range(1, min(len(leads) + 2, 50))
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_length + 2, 35)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)


def _export_summary(leads: list[SalesReadyLead], filename: str) -> None:
    hot = sum(1 for l in leads if l.priority == "HOT")
    warm = sum(1 for l in leads if l.priority == "WARM")
    nurture = sum(1 for l in leads if l.priority == "NURTURE")
    with_email = sum(1 for l in leads if l.business_email)
    with_phone = sum(1 for l in leads if l.business_phone)
    avg_fit = sum(l.commercial_fit for l in leads) / len(leads) if leads else 0

    categories = {}
    for l in leads:
        categories[l.category] = categories.get(l.category, 0) + 1

    summary = f"""
COMAI ENHANCED LEAD EXTRACTION — Summary
=========================================
Total Qualified Leads: {len(leads)}
  HOT:     {hot} (Can call TODAY)
  WARM:    {warm} (Can call THIS WEEK)
  NURTURE: {nurture} (Follow-up sequence)

Contact Availability:
  Email:    {with_email} ({with_email*100//len(leads) if leads else 0}%)
  Phone:    {with_phone} ({with_phone*100//len(leads) if leads else 0}%)

Average Commercial Fit: {avg_fit:.1f}/100

Category Breakdown:
"""
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        summary += f"  {cat}: {count}\n"

    summary += f"\nGenerated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n"

    with open(filename, "w") as f:
        f.write(summary)
    print(summary)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=500)
    parser.add_argument("--output", type=str, default="comai_enhanced_leads.xlsx")
    args = parser.parse_args()

    asyncio.run(run_enhanced_pipeline(limit=args.limit, output=args.output))
