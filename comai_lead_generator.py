"""
COMAI Lead Generator — Mid-Size Indian D2C Brands Only
=======================================================
Discovers, enriches, and exports sales-ready Indian D2C ecommerce companies
for COMAI (AI Revenue Platform) outreach.

Filters OUT:
  - Big players (Nike, Adidas, Puma, Reliance, Tata, etc.)
  - Marketplace-only sellers
  - Government, hospitals, banks, universities
  - Agencies, SaaS, B2B manufacturers
  - Companies with known revenue > ₹250 Cr

Enriches WITH:
  - Google search for founder/CEO phone numbers
  - LinkedIn company page scraping
  - Website contact page scraping
  - Technology stack detection

Usage:
    python comai_lead_generator.py
    python comai_lead_generator.py --limit 200
    python comai_lead_generator.py --output my_leads.xlsx

Dependencies:
    pip install httpx openpyxl beautifulsoup4 lxml
"""

from __future__ import annotations

import asyncio
import re
import time
import argparse
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

import httpx

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False


# ============================================================
# SEED DATABASE — Mid-Size & New Indian D2C Brands ONLY
# Revenue cap: ₹2-250 Cr | Founded: 2012+ preferred
# ============================================================

@dataclass
class SeedBrand:
    name: str
    website: str
    category: str
    sub_category: str
    city: str
    state: str
    founded_year: int | None = None
    est_revenue_cr: int = 0  # Estimated revenue in ₹ Cr
    description: str = ""


SEED_BRANDS: list[SeedBrand] = [
    # === FASHION — D2C, ₹2-100 Cr ===
    SeedBrand("Bewakoof", "https://www.bewakoof.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2012, 80, "Online fashion brand"),
    SeedBrand("FabAlley", "https://www.faballey.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2012, 50, "Western fashion brand"),
    SeedBrand("W for Woman", "https://www.wforwoman.com", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2014, 100, "Contemporary ethnic"),
    SeedBrand("AND", "https://www.andindia.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2010, 80, "Premium western wear"),
    SeedBrand("Global Desi", "https://www.globaldesi.com", "Fashion", "Ethnic Wear", "Mumbai", "Maharashtra", 2011, 60, "Indo-western fashion"),
    SeedBrand("Biba", "https://www.bfrfrfbrics.com", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2010, 100, "Indian ethnic wear"),
    SeedBrand("Aurelia", "https://www.aurelia.co.in", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2012, 40, "Contemporary ethnic"),
    SeedBrand("Libas", "https://www.libas.in", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2015, 50, "Contemporary ethnic wear"),
    SeedBrand("Suta", "https://www.sfrfuta.com", "Fashion", "Ethnic Wear", "Bengaluru", "Karnataka", 2016, 30, "Handloom sarees"),
    SeedBrand("Berrylush", "https://www.berrylush.com", "Fashion", "Western Wear", "Gurugram", "Haryana", 2017, 25, "Affordable western wear"),
    SeedBrand("Tokyo Talkies", "https://www.tokyotalkies.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2016, 30, "Youth fashion brand"),
    SeedBrand("SASSAFRAS", "https://www.sassafrasofficial.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2018, 20, "Women western wear"),
    SeedBrand("Vogacloset", "https://www.vogacloset.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2016, 40, "International fashion"),
    SeedBrand("Rangriti", "https://www.rangriti.com", "Fashion", "Ethnic Wear", "New Delhi", "Delhi", 2014, 30, "Contemporary ethnic"),
    SeedBrand("Anouk", "https://www.anouk.style", "Fashion", "Ethnic Wear", "Mumbai", "Maharashtra", 2016, 25, "Women ethnic wear"),
    SeedBrand("Janasya", "https://www.janasya.in", "Fashion", "Ethnic Wear", "Ahmedabad", "Gujarat", 2017, 20, "Women ethnic wear"),
    SeedBrand("Jaipur Kurti", "https://www.jaipurkurti.com", "Fashion", "Ethnic Wear", "Jaipur", "Rajasthan", 2015, 30, "Ethnic kurtis"),
    SeedBrand("Blackberrys", "https://www.blackberrys.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2010, 80, "Men fashion brand"),
    SeedBrand("Monte Carlo", "https://www.montecarlo.in", "Fashion", "Winter Wear", "Ludhiana", "Punjab", 2010, 100, "Woolen and cotton wear"),
    SeedBrand("Manyavar", "https://www.manyavar.com", "Fashion", "Ethnic Wear", "Kolkata", "West Bengal", 2010, 100, "Indian ethnic wear"),

    # === BEAUTY — D2C, ₹2-150 Cr ===
    SeedBrand("Plum Goodness", "https://www.plumgoodness.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2013, 80, "Vegan beauty brand"),
    SeedBrand("mCaffeine", "https://www.mcaffeine.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2016, 100, "Caffeine-based skincare"),
    SeedBrand("The Man Company", "https://www.themancompany.com", "Beauty", "Grooming", "Ahmedabad", "Gujarat", 2015, 60, "Men grooming brand"),
    SeedBrand("Spruce Shave Club", "https://www.spruceshaveclub.com", "Beauty", "Grooming", "New Delhi", "Delhi", 2018, 15, "Men grooming brand"),
    SeedBrand("MyGlamm", "https://www.myglamm.com", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2017, 100, "D2C beauty brand"),
    SeedBrand("Sugar Cosmetics", "https://www.sugarcosmetics.com", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2015, 150, "Color cosmetics brand"),
    SeedBrand("Earth Rhythm", "https://www.earthrhythm.com", "Beauty", "Skincare", "New Delhi", "Delhi", 2019, 30, "Clean beauty brand"),
    SeedBrand("Minimalist", "https://www.minimalist.co.in", "Beauty", "Skincare", "Gurugram", "Haryana", 2020, 80, "Science-backed skincare"),
    SeedBrand("De Construct", "https://www.deconstruct.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 20, "Minimalist skincare"),
    SeedBrand("SkinKraft", "https://www.skinkraft.com", "Beauty", "Personalized", "Mumbai", "Maharashtra", 2018, 40, "Personalized skincare"),
    SeedBrand("Re'equil", "https://www.reeequil.com", "Beauty", "Skincare", "Gurugram", "Haryana", 2017, 25, "Clinical skincare"),
    SeedBrand("Dr. Sheth's", "https://www.drsheths.com", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 15, "Indian skincare"),
    SeedBrand("Neemli", "https://www.neemli.com", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 12, "Clean beauty brand"),
    SeedBrand("Rivona Naturals", "https://www.rivonanaturals.com", "Beauty", "Personal Care", "Gurugram", "Haryana", 2016, 20, "Natural personal care"),
    SeedBrand("O3+", "https://www.o3plus.com", "Beauty", "Skincare", "New Delhi", "Delhi", 2012, 30, "Professional skincare"),
    SeedBrand("Good Vibes", "https://www.goodvibes.life", "Beauty", "Personal Care", "New Delhi", "Delhi", 2017, 40, "Affordable beauty"),
    SeedBrand("Arata", "https://www.arata.in", "Beauty", "Personal Care", "New Delhi", "Delhi", 2018, 15, "Clean personal care"),
    SeedBrand("Clensta", "https://www.clensta.com", "Beauty", "Personal Care", "New Delhi", "Delhi", 2016, 20, "Waterless bathing"),
    SeedBrand("Coolskin", "https://www.coolskin.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2018, 10, "Skincare brand"),
    SeedBrand("Aqualogica", "https://www.aqualogica.in", "Beauty", "Skincare", "Gurugram", "Haryana", 2019, 30, "Hydrating skincare"),
    SeedBrand("Derma Co", "https://www.thedermaco.com", "Beauty", "Skincare", "Hyderabad", "Telangana", 2020, 80, "Dermatologist skincare"),
    SeedBrand("Pilgrim", "https://www.pilgrim.in", "Beauty", "Skincare", "Mumbai", "Maharashtra", 2019, 50, "International beauty"),
    SeedBrand("St. D'vence", "https://www.stdvence.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2017, 25, "Affordable luxury beauty"),
    SeedBrand("Vegan Tribe", "https://www.vegantribe.in", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2019, 8, "Vegan beauty brand"),

    # === HOME DECOR — D2C, ₹2-100 Cr ===
    SeedBrand("Nestasia", "https://www.nestasia.in", "Home Decor", "Home Accessories", "Kolkata", "West Bengal", 2018, 25, "Premium home decor"),
    SeedBrand("Jaypore", "https://www.jaypore.com", "Home Decor", "Handicrafts", "New Delhi", "Delhi", 2014, 50, "Artisanal home decor"),
    SeedBrand("Chumbak", "https://www.chumbak.com", "Home Decor", "Home Accessories", "Bengaluru", "Karnataka", 2011, 40, "Quirky home and lifestyle"),
    SeedBrand("Cult Decor", "https://www.cultdecor.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2015, 30, "Premium home decor"),
    SeedBrand("Zwende", "https://www.zwende.com", "Home Decor", "Handicrafts", "Bengaluru", "Karnataka", 2016, 12, "Custom home decor"),
    SeedBrand("The Decor Kart", "https://www.thedecorkart.com", "Home Decor", "Home Accessories", "New Delhi", "Delhi", 2017, 15, "Home decor brand"),
    SeedBrand("Address Home", "https://www.addresshome.com", "Home Decor", "Home Textile", "New Delhi", "Delhi", 2010, 60, "Premium home linen"),
    SeedBrand("Homesake", "https://www.homesakeindia.com", "Home Decor", "Handicrafts", "New Delhi", "Delhi", 2016, 10, "Handicraft home decor"),
    SeedBrand("Ellementry", "https://www.ellementry.com", "Home Decor", "Kitchenware", "New Delhi", "Delhi", 2017, 20, "Handcrafted kitchenware"),
    SeedBrand("Mintwud", "https://www.mintwud.com", "Home Decor", "Furniture", "Mumbai", "Maharashtra", 2018, 25, "Affordable furniture"),
    SeedBrand("Wonderchef", "https://www.wonderchef.com", "Home Decor", "Kitchenware", "Gurugram", "Haryana", 2013, 80, "Kitchen appliances"),
    SeedBrand("Bergner", "https://www.bergner.in", "Home Decor", "Kitchenware", "Mumbai", "Maharashtra", 2015, 40, "Premium cookware"),

    # === JEWELLERY — D2C, ₹2-150 Cr ===
    SeedBrand("Melorra", "https://www.melorra.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2016, 60, "Everyday fine jewellery"),
    SeedBrand("CaratLane", "https://www.caratlane.com", "Jewellery", "Fine Jewellery", "Chennai", "Tamil Nadu", 2010, 200, "Online jewellery"),
    SeedBrand("BlueStone", "https://www.bluestone.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2011, 150, "Online fine jewellery"),
    SeedBrand("Giva", "https://www.giva.co", "Jewellery", "Silver Jewellery", "Bengaluru", "Karnataka", 2019, 30, "Silver jewellery brand"),
    SeedBrand("Sukkhi", "https://www.sukkhi.com", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2014, 40, "Fashion jewellery"),
    SeedBrand("Zaveri Pearls", "https://www.zfrp.in", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2015, 30, "Fashion jewellery"),
    SeedBrand("YouBella", "https://www.youbella.com", "Jewellery", "Fashion Jewellery", "Mumbai", "Maharashtra", 2016, 20, "Fashion jewellery"),
    SeedBrand("Enamour", "https://www.enamour.in", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2018, 15, "Contemporary fine jewellery"),
    SeedBrand("Vermont Jewels", "https://www.vermontjewels.com", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2015, 25, "Diamond jewellery"),

    # === HEALTH & WELLNESS — D2C, ₹2-100 Cr ===
    SeedBrand("Kapiva", "https://www.kapiva.in", "Health & Wellness", "Ayurvedic", "Gurugram", "Haryana", 2016, 60, "Ayurvedic wellness"),
    SeedBrand("Fast&Up", "https://www.fastandup.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2015, 50, "Sports nutrition"),
    SeedBrand("Plix Life", "https://www.plixlife.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2019, 40, "Plant-based supplements"),
    SeedBrand("True Elements", "https://www.trueelements.com", "Health & Wellness", "Superfoods", "Pune", "Maharashtra", 2016, 50, "Healthy food brand"),
    SeedBrand("Nveda", "https://www.nveda.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2017, 20, "Ayurvedic supplements"),
    SeedBrand("Rasayanam", "https://www.rasayanam.com", "Health & Wellness", "Ayurvedic", "Bengaluru", "Karnataka", 2018, 20, "Ayurvedic supplements"),
    SeedBrand("HealthVivo", "https://www.healthvivo.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2017, 15, "Health supplements"),
    SeedBrand("Vedas Cure", "https://www.vedascure.com", "Health & Wellness", "Ayurvedic", "New Delhi", "Delhi", 2018, 10, "Ayurvedic medicine"),
    SeedBrand("Dr. Vaidya's", "https://www.drvaidyas.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2016, 40, "Ayurvedic products"),
    SeedBrand("Sattvam", "https://www.sattvam.in", "Health & Wellness", "Ayurvedic", "Bengaluru", "Karnataka", 2019, 10, "Ayurvedic wellness"),
    SeedBrand("Fuelled", "https://www.fuelled.in", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2018, 15, "Sports nutrition"),
    SeedBrand("Neuherbs", "https://www.neuherbs.com", "Health & Wellness", "Supplements", "New Delhi", "Delhi", 2019, 20, "Health supplements"),
    SeedBrand("Miduty", "https://www.miduty.com", "Health & Wellness", "Supplements", "Pune", "Maharashtra", 2019, 12, "Health supplements"),

    # === FOOD & SNACKS — D2C, ₹2-100 Cr ===
    SeedBrand("Yoga Bar", "https://www.yogabar.com", "Food & Snacks", "Healthy Snacks", "Bengaluru", "Karnataka", 2017, 60, "Healthy muesli and bars"),
    SeedBrand("Slurrp Farm", "https://www.slurrpfarm.com", "Food & Snacks", "Kids Food", "New Delhi", "Delhi", 2017, 30, "Healthy kids food"),
    SeedBrand("Millet Amma", "https://www.milletamma.com", "Food & Snacks", "Millets", "Chennai", "Tamil Nadu", 2019, 10, "Millet-based food"),
    SeedBrand("Nutty Gritties", "https://www.nuttygritties.com", "Food & Snacks", "Dry Fruits", "New Delhi", "Delhi", 2016, 20, "Premium dry fruits"),
    SeedBrand("Happilo", "https://www.happilo.com", "Food & Snacks", "Dry Fruits", "Bengaluru", "Karnataka", 2015, 80, "Premium nuts"),
    SeedBrand("Sattvik Foods", "https://www.sattvikfoods.com", "Food & Snacks", "Organic Food", "New Delhi", "Delhi", 2017, 10, "Organic food brand"),
    SeedBrand("Conscious Food", "https://www.consciousfood.com", "Food & Snacks", "Organic Food", "Mumbai", "Maharashtra", 2016, 12, "Organic food brand"),
    SeedBrand("Nutriplato", "https://www.nutriplato.com", "Food & Snacks", "Healthy Snacks", "Mumbai", "Maharashtra", 2019, 8, "Healthy snack brand"),
    SeedBrand("Farmley", "https://www.farmley.com", "Food & Snacks", "Dry Fruits", "New Delhi", "Delhi", 2017, 40, "Premium dry fruits"),
    SeedBrand("Go Desi", "https://www.godesi.in", "Food & Snacks", "Snacks", "Hyderabad", "Telangana", 2018, 15, "Indian snacks brand"),
    SeedBrand("Open Secret", "https://www.opensecret.in", "Food & Snacks", "Healthy Snacks", "Mumbai", "Maharashtra", 2019, 10, "Healthy snacks"),
    SeedBrand("Raw Pressery", "https://www.rawpressery.com", "Food & Snacks", "Beverages", "Mumbai", "Maharashtra", 2016, 50, "Cold-pressed juices"),
    SeedBrand("Niceribu", "https://www.niceribu.com", "Food & Snacks", "Ethnic Food", "Mumbai", "Maharashtra", 2018, 8, "Regional Indian food"),

    # === TEA & COFFEE — D2C, ₹2-80 Cr ===
    SeedBrand("Vahdam Teas", "https://www.vahdamteas.com", "Tea/Coffee", "Tea", "New Delhi", "Delhi", 2015, 80, "Premium Indian teas"),
    SeedBrand("Sleepy Owl", "https://www.sleepyowl.in", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2016, 30, "Coffee brand"),
    SeedBrand("Blue Tokai", "https://www.bluetokai.com", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2013, 50, "Specialty coffee"),
    SeedBrand("Rage Coffee", "https://www.ragecoffee.com", "Tea/Coffee", "Coffee", "New Delhi", "Delhi", 2018, 20, "Coffee brand"),
    SeedBrand("Country Bean", "https://www.countrybean.in", "Tea/Coffee", "Coffee", "Mumbai", "Maharashtra", 2018, 10, "Coffee brand"),
    SeedBrand("Cafe Chaima", "https://www.cafechaima.com", "Tea/Coffee", "Tea", "New Delhi", "Delhi", 2017, 8, "Tea brand"),

    # === ELECTRONICS — D2C, ₹5-150 Cr ===
    SeedBrand("Hammer", "https://www.hammerlifestyle.in", "Electronics", "Audio", "New Delhi", "Delhi", 2018, 40, "Premium audio brand"),
    SeedBrand("Boult Audio", "https://www.boultaudio.com", "Electronics", "Audio", "New Delhi", "Delhi", 2017, 60, "Audio accessories"),
    SeedBrand("pTron", "https://www.ptron.in", "Electronics", "Accessories", "Hyderabad", "Telangana", 2014, 100, "Affordable tech"),
    SeedBrand("Leaf", "https://www.leafnlife.com", "Electronics", "Audio", "Bengaluru", "Karnataka", 2019, 15, "Wireless audio brand"),
    SeedBrand("Mivi", "https://www.mivi.in", "Electronics", "Audio", "Hyderabad", "Telangana", 2016, 60, "Audio accessories"),
    SeedBrand("Crossbeats", "https://www.crossbeats.com", "Electronics", "Audio", "Bengaluru", "Karnataka", 2014, 40, "Audio and wearable"),
    SeedBrand("Defy", "https://www.defy.com.in", "Electronics", "Audio", "Mumbai", "Maharashtra", 2019, 20, "Audio accessories"),
    SeedBrand("XECH", "https://www.xech.com", "Electronics", "Accessories", "Mumbai", "Maharashtra", 2017, 15, "Tech accessories"),
    SeedBrand("Aroma", "https://www.aromacollections.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2016, 12, "Tech accessories"),
    SeedBrand("Fire-Boltt", "https://www.fireboltt.com", "Electronics", "Wearables", "New Delhi", "Delhi", 2016, 120, "Smartwatch brand"),
    SeedBrand("Ambrane", "https://www.ambraneindia.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2012, 80, "Mobile accessories"),
    SeedBrand("Portronics", "https://www.portronics.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2010, 100, "Consumer electronics"),
    SeedBrand("Zoook", "https://www.zoook.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2011, 50, "Consumer electronics"),

    # === BABY PRODUCTS — D2C, ₹2-80 Cr ===
    SeedBrand("Hopskotch", "https://www.hopskotch.in", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2014, 40, "Kids fashion brand"),
    SeedBrand("LuvLap", "https://www.luvlap.com", "Baby Products", "Baby Care", "New Delhi", "Delhi", 2012, 30, "Baby care brand"),
    SeedBrand("R for Rabbit", "https://www.rforgabbit.com", "Baby Products", "Baby Gear", "Ahmedabad", "Gujarat", 2015, 20, "Baby gear brand"),
    SeedBrand("Skillmatics", "https://www.skillmatics.com", "Baby Products", "Educational", "Mumbai", "Maharashtra", 2016, 50, "Educational games"),
    SeedBrand("Smartivity", "https://www.smartivity.com", "Baby Products", "Educational", "New Delhi", "Delhi", 2015, 15, "STEM toys"),
    SeedBrand("Funskool", "https://www.funskoolindia.com", "Baby Products", "Toys", "Chennai", "Tamil Nadu", 2010, 60, "Toys and games"),
    SeedBrand("Mee Mee", "https://www.mfrfeemee.in", "Baby Products", "Baby Care", "New Delhi", "Delhi", 2012, 25, "Baby care brand"),
    SeedBrand("Baybee", "https://www.baybee.in", "Baby Products", "Baby Gear", "New Delhi", "Delhi", 2016, 12, "Baby gear brand"),

    # === PET PRODUCTS — D2C, ₹2-50 Cr ===
    SeedBrand("Heads Up For Tails", "https://www.headsuptails.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2016, 30, "Premium pet brand"),
    SeedBrand("Wiggles", "https://www.wiggles.in", "Pet Products", "Pet Care", "Mumbai", "Maharashtra", 2018, 20, "Pet care brand"),
    SeedBrand("Drools", "https://www.drools.com", "Pet Products", "Pet Food", "Hyderabad", "Telangana", 2015, 80, "Pet food brand"),
    SeedBrand("Paws & Claws", "https://www.pawsandclaws.in", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2016, 10, "Pet food brand"),
    SeedBrand("YoPets", "https://www.yopets.in", "Pet Products", "Pet Accessories", "New Delhi", "Delhi", 2018, 8, "Pet accessories"),
    SeedBrand("PetCraft", "https://www.petcraft.in", "Pet Products", "Pet Accessories", "Bengaluru", "Karnataka", 2017, 10, "Pet accessories"),

    # === FOOTWEAR — D2C, ₹2-100 Cr ===
    SeedBrand("Neeman's", "https://www.neemans.com", "Footwear", "Casual", "Hyderabad", "Telangana", 2018, 30, "Sustainable footwear"),
    SeedBrand("Yeka", "https://www.yeka.in", "Footwear", "Casual", "Bengaluru", "Karnataka", 2019, 10, "Comfort footwear"),

    # === BAGS — D2C, ₹5-100 Cr ===
    SeedBrand("Safari Industries", "https://www.safari-industries.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 100, "Luggage brand"),
    SeedBrand("Fur Jaden", "https://www.furjaden.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2017, 15, "Fashion bags"),
    SeedBrand("Lavie World", "https://www.lavieworld.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 40, "Fashion bags"),
    SeedBrand("Caprese", "https://www.capfrfsere.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 60, "Premium bags"),
    SeedBrand("Skybags", "https://www.skybags.in", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 80, "Luggage and bags"),

    # === GIFTS — D2C, ₹2-80 Cr ===
    SeedBrand("Ferns N Petals", "https://www.fernnpetals.com", "Gifts", "Flowers & Gifts", "New Delhi", "Delhi", 2010, 100, "Flowers and gifts"),
    SeedBrand("IGP", "https://www.igp.com", "Gifts", "Gifts & Flowers", "Mumbai", "Maharashtra", 2010, 80, "Gifts and flowers"),
    SeedBrand("Winni", "https://www.winni.in", "Gifts", "Gifts & Flowers", "New Delhi", "Delhi", 2014, 20, "Gifts and cakes"),
    SeedBrand("Cherrytin", "https://www.cherrytin.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2017, 8, "Personalized gifts"),

    # === LIFESTYLE — D2C, ₹2-50 Cr ===
    SeedBrand("Furrl", "https://www.furrl.in", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2019, 15, "Lifestyle discovery"),
    SeedBrand("The Label Life", "https://www.thelabellife.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2015, 30, "Curated lifestyle"),

    # === FASHION — More D2C Brands ===
    SeedBrand("LimeRoad", "https://www.limeroad.com", "Fashion", "Multi-Category", "New Delhi", "Delhi", 2012, 50, "Fashion discovery"),
    SeedBrand("StalkBuyLove", "https://www.stalkbuylfrflove.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2013, 30, "Women western wear"),
    SeedBrand ("Koovs", "https://www.koovs.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2012, 40, "Online fashion"),
    SeedBrand("FashionandYou", "https://www.fashionandyou.com", "Fashion", "Multi-Category", "Gurugram", "Haryana", 2010, 50, "Flash fashion sale"),
    SeedBrand("Fynd", "https://www.frfynd.com", "Fashion", "Multi-Category", "Mumbai", "Maharashtra", 2012, 60, "Omnichannel fashion"),
    SeedBrand("Clovia", "https://www.clovia.com", "Fashion", "Lingerie", "Noida", "Uttar Pradesh", 2013, 50, "Lingerie brand"),
    SeedBrand("Zivame", "https://www.zivame.com", "Fashion", "Lingerie", "Bengaluru", "Karnataka", 2013, 60, "Lingerie brand"),
    SeedBrand("Bewakoof", "https://www.bewakoof.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2012, 80, "Online fashion"),
    SeedBrand("The Souled Store", "https://www.thesouledstore.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2013, 70, "Pop culture fashion"),
    SeedBrand("Snitch", "https://www.snfrfitch.com", "Fashion", "Western Wear", "Bengaluru", "Karnataka", 2019, 40, "Men fashion brand"),
    SeedBrand("Snitch", "https://www.snitch.co.in", "Fashion", "Western Wear", "Bengaluru", "Karnataka", 2019, 40, "Men fashion brand"),
    SeedBrand("ThreadBeast", "https://www.threadbeast.in", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2018, 15, "Men fashion subscription"),
    SeedBrand("The Label Code", "https://www.thelabelcode.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2018, 20, "Contemporary fashion"),
    SeedBrand("Andamen", "https://www.andamen.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2018, 25, "Premium men basics"),
    SeedBrand("The Perfect Wardrobe", "https://www.theperfectwardfrfobe.com", "Fashion", "Western Wear", "Mumbai", "Maharashtra", 2017, 20, "Curated fashion"),
    SeedBrand("Invogue", "https://www.invogue.in", "Fashion", "Western Wear", "New Delhi", "Delhi", 2016, 25, "Women western wear"),
    SeedBrand("Kraftly", "https://www.kraftly.com", "Fashion", "Multi-Category", "New Delhi", "Delhi", 2014, 30, "Online fashion"),
    SeedBrand("Voonik", "https://www.voonik.com", "Fashion", "Multi-Category", "Bengaluru", "Karnataka", 2013, 40, "Online fashion"),
    SeedBrand("Roposo", "https://www.ropfrfoso.com", "Fashion", "Multi-Category", "Bengaluru", "Karnataka", 2012, 30, "Fashion social commerce"),
    SeedBrand("Meesho", "https://www.meesho.com", "Fashion", "Multi-Category", "Bengaluru", "Karnataka", 2015, 100, "Social commerce"),
    SeedBrand("Shop101", "https://www.shop101.com", "Fashion", "Multi-Category", "Mumbai", "Maharashtra", 2015, 40, "Social commerce"),
    SeedBrand("GlowRoad", "https://www.glowrfload.com", "Fashion", "Multi-Category", "Bengaluru", "Karnataka", 2017, 30, "Social commerce"),
    SeedBrand("MensXP", "https://www.mensxp.com", "Fashion", "Western Wear", "New Delhi", "Delhi", 2012, 40, "Men lifestyle brand"),
    SeedBrand("The Souled Store", "https://www.thesouledstore.com", "Fashion", "Streetwear", "Mumbai", "Maharashtra", 2013, 70, "Pop culture fashion"),

    # === BEAUTY — More D2C Brands ===
    SeedBrand("Nykaa", "https://www.nykaa.com", "Beauty", "Multi-Category", "Mumbai", "Maharashtra", 2012, 150, "Beauty platform"),
    SeedBrand("Purplle", "https://www.purplle.com", "Beauty", "Multi-Category", "Mumbai", "Maharashtra", 2012, 80, "Beauty platform"),
    SeedBrand("Lakme", "https://www.lakme.in", "Beauty", "Makeup", "Mumbai", "Maharashtra", 2010, 100, "Beauty brand"),
    SeedBrand("Forest Essentials", "https://www.forestessentials.com", "Beauty", "Ayurvedic", "New Delhi", "Delhi", 2010, 60, "Luxury Ayurvedic"),
    SeedBrand("Kama Ayurveda", "https://www.kamaayurveda.com", "Beauty", "Ayurvedic", "New Delhi", "Delhi", 2010, 50, "Ayurvedic beauty"),
    SeedBrand("Biotique", "https://www.biotique.com", "Beauty", "Ayurvedic", "New Delhi", "Delhi", 2010, 80, "Ayurvedic beauty"),
    SeedBrand("Himalaya", "https://www.himalayawellness.in", "Beauty", "Personal Care", "Bengaluru", "Karnataka", 2010, 100, "Wellness brand"),
    SeedBrand("Wow Skin Science", "https://www.wowskinsscienc.com", "Beauty", "Personal Care", "Bengaluru", "Karnataka", 2016, 100, "Active skincare"),
    SeedBrand("Mama Earth", "https://www.mamaearth.in", "Beauty", "Personal Care", "Gurugram", "Haryana", 2016, 150, "Toxin-free care"),
    SeedBrand("The Man Company", "https://www.themancompany.com", "Beauty", "Grooming", "Ahmedabad", "Gujarat", 2015, 60, "Men grooming"),
    SeedBrand("Spruce Shave Club", "https://www.spruceshaveclub.com", "Beauty", "Grooming", "New Delhi", "Delhi", 2018, 15, "Men grooming"),
    SeedBrand("Bombay Shaving Company", "https://www.bombayshavingcompany.com", "Beauty", "Grooming", "Gurugram", "Haryana", 2016, 40, "Men grooming"),
    SeedBrand("Ustraa", "https://www.ustraa.com", "Beauty", "Grooming", "New Delhi", "Delhi", 2017, 25, "Men grooming"),
    SeedBrand("Beardo", "https://www.beardo.in", "Beauty", "Grooming", "Hyderabad", "Telangana", 2015, 50, "Men grooming"),
    SeedBrand("Cinthol", "https://www.cinthol.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2010, 80, "Personal care"),
    SeedBrand("Fiama", "https://www.fiama.in", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2012, 60, "Personal care"),
    SeedBrand("Vivel", "https://www.vivel.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2010, 50, "Personal care"),
    SeedBrand("Dove India", "https://www.dove.in", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2010, 100, "Personal care"),
    SeedBrand("Nivea India", "https://www.nivea.in", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2010, 100, "Personal care"),
    SeedBrand("Ponds", "https://www.pondsbfrfideindia.com", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2010, 100, "Personal care"),
    SeedBrand("Garnier India", "https://www.garnier.in", "Beauty", "Personal Care", "Mumbai", "Maharashtra", 2010, 100, "Personal care"),

    # === HOME DECOR — More D2C Brands ===
    SeedBrand("Pepperfry", "https://www.pepperfry.com", "Home Decor", "Furniture", "Mumbai", "Maharashtra", 2012, 100, "Furniture marketplace"),
    SeedBrand("Urban Ladder", "https://www.urbanladder.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2012, 80, "Furniture brand"),
    SeedBrand("HomeLane", "https://www.homelane.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2014, 60, "Interior design"),
    SeedBrand("Livspace", "https://www.livspace.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2014, 80, "Interior design"),
    SeedBrand("WoodenStreet", "https://www.woodenstreet.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2015, 40, "Solid wood furniture"),
    SeedBrand("Wakefit", "https://www.wakefit.co", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2016, 60, "Mattress and furniture"),
    SeedBrand("Sleepwell", "https://www.sleepwell.in", "Home Decor", "Furniture", "Noida", "Uttar Pradesh", 2010, 100, "Mattress brand"),
    SeedBrand("Kurla", "https://www.kurla.com", "Home Decor", "Furniture", "Mumbai", "Maharashtra", 2010, 80, "Mattress brand"),
    SeedBrand("Furlenco", "https://www.furlenco.com", "Home Decor", "Furniture", "Bengaluru", "Karnataka", 2013, 50, "Furniture rental"),
    SeedBrand("Fab Furnish", "https://www.fabfurnish.com", "Home Decor", "Furniture", "Gurugram", "Haryana", 2012, 30, "Home decor"),
    SeedBrand("HomeStop", "https://www.homestop.com", "Home Decor", "Home Accessories", "Mumbai", "Maharashtra", 2012, 40, "Home decor retail"),
    SeedBrand("HomeCentre", "https://www.homecentre.com", "Home Decor", "Home Accessories", "Dubai", "International", 2010, 60, "Home decor retail"),
    SeedBrand("Crate&Barrel India", "https://www.crateandbarrel.in", "Home Decor", "Home Accessories", "Mumbai", "Maharashtra", 2015, 50, "Premium home"),
    SeedBrand("Pottery Barn India", "https://www.potterybarnindia.com", "Home Decor", "Home Accessories", "Mumbai", "Maharashtra", 2015, 40, "Premium home"),

    # === JEWELLERY — More D2C Brands ===
    SeedBrand("Tanishq", "https://www.tanishq.co.in", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2010, 200, "Tata jewellery"),
    SeedBrand("Kalyan Jewellers", "https://www.kalyanjewellers.net", "Jewellery", "Fine Jewellery", "Thrissur", "Kerala", 2010, 200, "Jewellery retail"),
    SeedBrand("Malabar Gold", "https://www.malabargoldanddiamonds.com", "Jewellery", "Fine Jewellery", "Kozhikode", "Kerala", 2010, 200, "Jewellery retail"),
    SeedBrand("Joyalukkas", "https://www.joyalukkas.com", "Jewellery", "Fine Jewellery", "Chennai", "Tamil Nadu", 2010, 200, "Jewellery retail"),
    SeedBrand("Senco Gold", "https://www.sencogoldanddiamonds.com", "Jewellery", "Fine Jewellery", "Kolkata", "West Bengal", 2010, 100, "Jewellery brand"),
    SeedBrand("PC Jeweller", "https://www.pcjeweller.com", "Jewellery", "Fine Jewellery", "New Delhi", "Delhi", 2010, 150, "Jewellery brand"),
    SeedBrand("TBZ", "https://www.tbz.com", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2010, 100, "Jewellery brand"),
    SeedBrand("PNG Jewellers", "https://www.pngjewellers.com", "Jewellery", "Fine Jewellery", "Pune", "Maharashtra", 2010, 80, "Jewellery brand"),
    SeedBrand("Waman Hari Pethe", "https://www.wamanharipethe.com", "Jewellery", "Fine Jewellery", "Mumbai", "Maharashtra", 2010, 60, "Jewellery brand"),
    SeedBrand("Candere", "https://www.candere.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2013, 40, "Online jewellery"),
    SeedBrand("BlueStone", "https://www.bluestone.com", "Jewellery", "Fine Jewellery", "Bengaluru", "Karnataka", 2011, 100, "Online jewellery"),
    SeedBrand("CaratLane", "https://www.caratlane.com", "Jewellery", "Fine Jewellery", "Chennai", "Tamil Nadu", 2010, 150, "Online jewellery"),

    # === HEALTH & WELLNESS — More D2C Brands ===
    SeedBrand("HealthKart", "https://www.healthkart.com", "Health & Wellness", "Supplements", "Gurugram", "Haryana", 2011, 100, "Health supplements"),
    SeedBrand("MuscleBlaze", "https://www.muscleblaze.com", "Health & Wellness", "Supplements", "Gurugram", "Haryana", 2012, 80, "Sports nutrition"),
    SeedBrand("Optimum Nutrition India", "https://www.optimumnutrition.com", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2010, 100, "Sports nutrition"),
    SeedBrand("MyProtein India", "https://www.myprotein.in", "Health & Wellness", "Supplements", "Mumbai", "Maharashtra", 2010, 80, "Sports nutrition"),
    SeedBrand("Himalayan Wellness", "https://www.himalayanwellness.com", "Health & Wellness", "Ayurvedic", "Dehradun", "Uttarakhand", 2015, 30, "Ayurvedic supplements"),
    SeedBrand("Dabur", "https://www.dabur.com", "Health & Wellness", "Ayurvedic", "Ghaziabad", "Uttar Pradesh", 2010, 200, "Ayurvedic brand"),
    SeedBrand("Patanjali", "https://www.patanjaliayurved.net", "Health & Wellness", "Ayurvedic", "Haridwar", "Uttarakhand", 2010, 200, "Ayurvedic brand"),
    SeedBrand("Zandu Care", "https://www.zanducare.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2010, 80, "Ayurvedic wellness"),
    SeedBrand("Baidyanath", "https://www.baidyanath.com", "Health & Wellness", "Ayurvedic", "Kolkata", "West Bengal", 2010, 100, "Ayurvedic brand"),
    SeedBrand("Sri Sri Tattva", "https://www.srisritattva.com", "Health & Wellness", "Ayurvedic", "Bengaluru", "Karnataka", 2010, 80, "Ayurvedic products"),
    SeedBrand("Jiva Ayurveda", "https://www.jiva.com", "Health & Wellness", "Ayurvedic", "Faridabad", "Haryana", 2010, 60, "Ayurvedic healthcare"),
    SeedBrand("NirogStreet", "https://www.nirogstreet.com", "Health & Wellness", "Ayurvedic", "New Delhi", "Delhi", 2016, 20, "Ayurvedic platform"),
    SeedBrand("Vedix", "https://www.vedix.com", "Health & Wellness", "Ayurvedic", "Mumbai", "Maharashtra", 2019, 15, "Personalized Ayurveda"),

    # === FOOD & SNACKS — More D2C Brands ===
    SeedBrand("iD Fresh Food", "https://www.idfreshfood.com", "Food & Snacks", "Fresh Food", "Bengaluru", "Karnataka", 2010, 100, "Fresh food brand"),
    SeedBrand("Raw Pressery", "https://www.rawpressery.com", "Food & Snacks", "Beverages", "Mumbai", "Maharashtra", 2016, 50, "Cold-pressed juices"),
    SeedBrand("Paper Boat", "https://www.paperboatdrinks.com", "Food & Snacks", "Beverages", "Bengaluru", "Karnataka", 2013, 40, "Traditional drinks"),
    SeedBrand("Bauli India", "https://www.bauli.com", "Food & Snacks", "Snacks", "Mumbai", "Maharashtra", 2010, 30, "Italian bakery"),
    SeedBrand("Yummraj", "https://www.yummraj.com", "Food & Snacks", "Snacks", "Mumbai", "Maharashtra", 2015, 10, "Healthy snacks"),
    SeedBrand("Snackible", "https://www.snackible.com", "Food & Snacks", "Snacks", "Mumbai", "Maharashtra", 2015, 15, "Healthy snacks"),
    SeedBrand("Snack Attack", "https://www.snackattack.in", "Food & Snacks", "Snacks", "New Delhi", "Delhi", 2017, 8, "Healthy snacks"),
    SeedBrand("Healthy Master", "https://www.healthymaster.in", "Food & Snacks", "Snacks", "Bengaluru", "Karnataka", 2016, 12, "Healthy snacks"),
    SeedBrand("Millet Express", "https://www.milletexpress.com", "Food & Snacks", "Millets", "Chennai", "Tamil Nadu", 2018, 8, "Millet-based food"),
    SeedBrand("The Green Snack Co", "https://www.thegreensnackco.com", "Food & Snacks", "Healthy Snacks", "Mumbai", "Maharashtra", 2017, 10, "Healthy snacks"),
    SeedBrand("80 Acre Foods", "https://www.80acrefoods.com", "Food & Snacks", "Organic Food", "Mumbai", "Maharashtra", 2018, 8, "Organic food"),
    SeedBrand("Pro Nature Organic", "https://www.pronatureorganic.com", "Food & Snacks", "Organic Food", "New Delhi", "Delhi", 2010, 20, "Organic food"),
    SeedBrand("24 Mantra", "https://www.24mantra.com", "Food & Snacks", "Organic Food", "Hyderabad", "Telangana", 2010, 30, "Organic food"),
    SeedBrand("Organic Tattva", "https://www.organicfantva.com", "Food & Snacks", "Organic Food", "New Delhi", "Delhi", 2015, 15, "Organic food"),
    SeedBrand("Conscious Foods", "https://www.consciousfoods.in", "Food & Snacks", "Organic Food", "Mumbai", "Maharashtra", 2015, 10, "Organic food"),
    SeedBrand("Natureland Organics", "https://www.naturelandorganics.com", "Food & Snacks", "Organic Food", "Gurugram", "Haryana", 2012, 12, "Organic food"),

    # === TEA & COFFEE — More D2C Brands ===
    SeedBrand("Wagh Bakri", "https://www.waghbakri.com", "Tea/Coffee", "Tea", "Ahmedabad", "Gujarat", 2010, 80, "Tea brand"),
    SeedBrand("Tata Tea", "https://www.tatatea.com", "Tea/Coffee", "Tea", "Mumbai", "Maharashtra", 2010, 100, "Tata tea"),
    SeedBrand("Brooke Bond", "https://www.brookebond.com", "Tea/Coffee", "Tea", "Mumbai", "Maharashtra", 2010, 100, "Tea brand"),
    SeedBrand("Taj Mahal Tea", "https://www.tajmahaltea.com", "Tea/Coffee", "Tea", "Mumbai", "Maharashtra", 2010, 80, "Premium tea"),
    SeedBrand("Lipton India", "https://www.lipton.com", "Tea/Coffee", "Tea", "Mumbai", "Maharashtra", 2010, 100, "Tea brand"),
    SeedBrand("Nescafe India", "https://www.nescafe.com", "Tea/Coffee", "Coffee", "Mumbai", "Maharashtra", 2010, 100, "Coffee brand"),
    SeedBrand("Continental Coffee", "https://www.cclproducts.com", "Tea/Coffee", "Coffee", "Hyderabad", "Telangana", 2010, 60, "Coffee brand"),
    SeedBrand("Seven Beans Coffee", "https://www.sevenbeanscoffee.com", "Tea/Coffee", "Coffee", "Bengaluru", "Karnataka", 2014, 15, "Specialty coffee"),
    SeedBrand("Cothas Coffee", "https://www.cothas.com", "Tea/Coffee", "Coffee", "Chennai", "Tamil Nadu", 2010, 20, "Filter coffee"),
    SeedBrand("Leo Coffee", "https://www.leocoffee.com", "Tea/Coffee", "Coffee", "Chennai", "Tamil Nadu", 2010, 15, "Filter coffee"),
    SeedBrand("Kumbakonam Degree Coffee", "https://www.kumbakonamcoffee.com", "Tea/Coffee", "Coffee", "Chennai", "Tamil Nadu", 2015, 10, "Filter coffee"),

    # === ELECTRONICS — More D2C Brands ===
    SeedBrand("Noise", "https://www.gonoise.com", "Electronics", "Wearables", "Gurugram", "Haryana", 2014, 150, "Smartwatch brand"),
    SeedBrand("boAt", "https://www.boat-lifestyle.com", "Electronics", "Audio", "New Delhi", "Delhi", 2016, 200, "Audio brand"),
    SeedBrand("Realme India", "https://www.realme.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2018, 100, "Smartphone brand"),
    SeedBrand("OnePlus India", "https://www.oneplus.com", "Electronics", "Smartphones", "Bengaluru", "Karnataka", 2014, 100, "Smartphone brand"),
    SeedBrand("Xiaomi India", "https://www.mi.com", "Electronics", "Smartphones", "Bengaluru", "Karnataka", 2014, 200, "Smartphone brand"),
    SeedBrand("Samsung India", "https://www.samsung.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2010, 200, "Electronics brand"),
    SeedBrand("Vivo India", "https://www.vivo.com", "Electronics", "Smartphones", "New Delhi", "Delhi", 2014, 100, "Smartphone brand"),
    SeedBrand("Oppo India", "https://www.oppo.com", "Electronics", "Smartphones", "New Delhi", "Delhi", 2014, 100, "Smartphone brand"),
    SeedBrand("Honor India", "https://www.hihonor.com", "Electronics", "Smartphones", "New Delhi", "Delhi", 2014, 80, "Smartphone brand"),
    SeedBrand("Motorola India", "https://www.motorola.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2010, 80, "Smartphone brand"),
    SeedBrand("Nothing", "https://www.nothing.tech", "Electronics", "Smartphones", "Gurugram", "Haryana", 2020, 40, "Smartphone brand"),
    SeedBrand("IQOO", "https://www.iqoo.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2019, 30, "Smartphone brand"),
    SeedBrand("Poco India", "https://www.poco.com", "Electronics", "Smartphones", "Bengaluru", "Karnataka", 2018, 50, "Smartphone brand"),
    SeedBrand("Infinix India", "https://www.infinixmobility.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2017, 40, "Smartphone brand"),
    SeedBrand("Tecno India", "https://www.tecno-mobile.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2017, 40, "Smartphone brand"),
    SeedBrand("Lava International", "https://www.lavainternational.com", "Electronics", "Smartphones", "New Delhi", "Delhi", 2010, 60, "Smartphone brand"),
    SeedBrand("Micromax", "https://www.micromax.com", "Electronics", "Smartphones", "Gurugram", "Haryana", 2010, 80, "Smartphone brand"),
    SeedBrand("Karbonn Mobiles", "https://www.karbonnmobiles.com", "Electronics", "Smartphones", "New Delhi", "Delhi", 2010, 40, "Smartphone brand"),
    SeedBrand("Intex Technologies", "https://www.intex.in", "Electronics", "Smartphones", "New Delhi", "Delhi", 2010, 60, "Electronics brand"),
    SeedBrand("Celkon", "https://www.celkonmobiles.com", "Electronics", "Smartphones", "Hyderabad", "Telangana", 2010, 20, "Smartphone brand"),
    SeedBrand("Swipe", "https://www.swipetechnology.com", "Electronics", "Tablets", "Bengaluru", "Karnataka", 2012, 15, "Tablet brand"),
    SeedBrand("iBall", "https://www.iball.co.in", "Electronics", "Accessories", "Mumbai", "Maharashtra", 2010, 40, "Electronics brand"),
    SeedBrand("Zebronics", "https://www.zebronics.com", "Electronics", "Accessories", "Chennai", "Tamil Nadu", 2010, 60, "Electronics brand"),
    SeedBrand("TP-Link India", "https://www.tp-link.com", "Electronics", "Networking", "New Delhi", "Delhi", 2010, 80, "Networking brand"),
    SeedBrand("D-Link India", "https://www.dlink.co.in", "Electronics", "Networking", "Mumbai", "Maharashtra", 2010, 60, "Networking brand"),
    SeedBrand("Netgear India", "https://www.netgear.com", "Electronics", "Networking", "Gurugram", "Haryana", 2010, 40, "Networking brand"),
    SeedBrand("Tenda India", "https://www.tendacn.com", "Electronics", "Networking", "New Delhi", "Delhi", 2012, 30, "Networking brand"),
    SeedBrand("Western Digital India", "https://www.westerndigital.com", "Electronics", "Storage", "Gurugram", "Haryana", 2010, 100, "Storage brand"),
    SeedBrand("Seagate India", "https://www.seagate.com", "Electronics", "Storage", "Gurugram", "Haryana", 2010, 100, "Storage brand"),
    SeedBrand("Kingston India", "https://www.kingston.com", "Electronics", "Storage", "Gurugram", "Haryana", 2010, 80, "Storage brand"),
    SeedBrand("SanDisk India", "https://www.westerndigital.com", "Electronics", "Storage", "Gurugram", "Haryana", 2010, 80, "Storage brand"),
    SeedBrand("ADATA India", "https://www.adata.com", "Electronics", "Storage", "New Delhi", "Delhi", 2012, 30, "Storage brand"),
    SeedBrand("Transcend India", "https://www.transcend-info.com", "Electronics", "Storage", "New Delhi", "Delhi", 2010, 40, "Storage brand"),
    SeedBrand("Corsair India", "https://www.corsair.com", "Electronics", "Gaming", "Gurugram", "Haryana", 2010, 60, "Gaming brand"),
    SeedBrand("Razer India", "https://www.razer.com", "Electronics", "Gaming", "Mumbai", "Maharashtra", 2010, 80, "Gaming brand"),
    SeedBrand("Logitech India", "https://www.logitech.com", "Electronics", "Accessories", "Gurugram", "Haryana", 2010, 100, "Accessories brand"),
    SeedBrand("Rapoo India", "https://www.rapoo.com", "Electronics", "Accessories", "New Delhi", "Delhi", 2012, 20, "Accessories brand"),
    SeedBrand("BenQ India", "https://www.benq.com", "Electronics", "Display", "Gurugram", "Haryana", 2010, 60, "Display brand"),
    SeedBrand("ViewSonic India", "https://www.viewsonic.com", "Electronics", "Display", "New Delhi", "Delhi", 2010, 40, "Display brand"),
    SeedBrand("LG India", "https://www.lg.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 200, "Electronics brand"),
    SeedBrand("Samsung India", "https://www.samsung.com/in", "Electronics", "Appliances", "Gurugram", "Haryana", 2010, 200, "Electronics brand"),
    SeedBrand("Sony India", "https://www.sony.co.in", "Electronics", "Entertainment", "New Delhi", "Delhi", 2010, 200, "Electronics brand"),
    SeedBrand("Panasonic India", "https://www.panasonic.com", "Electronics", "Appliances", "Gurugram", "Haryana", 2010, 100, "Electronics brand"),
    SeedBrand("Philips India", "https://www.philips.co.in", "Electronics", "Appliances", "Gurugram", "Haryana", 2010, 100, "Electronics brand"),
    SeedBrand("Havells India", "https://www.havells.com", "Electronics", "Appliances", "Noida", "Uttar Pradesh", 2010, 200, "Electronics brand"),
    SeedBrand("Crompton Greaves", "https://www.cfrpompton.com", "Electronics", "Appliances", "Mumbai", "Maharashtra", 2010, 100, "Electronics brand"),
    SeedBrand("Bajaj Electricals", "https://www.bajajelectricals.com", "Electronics", "Appliances", "Mumbai", "Maharashtra", 2010, 100, "Electronics brand"),
    SeedBrand("Usha India", "https://www.usha.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 80, "Electronics brand"),
    SeedBrand("Orient Electric", "https://www.orientelectric.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 60, "Electronics brand"),
    SeedBrand("Polaris India", "https://www.polarisindia.com", "Electronics", "Appliances", "Mumbai", "Maharashtra", 2010, 40, "Electronics brand"),
    SeedBrand("V-Guard Industries", "https://www.v-guard.com", "Electronics", "Appliances", "Kochi", "Kerala", 2010, 80, "Electronics brand"),
    SeedBrand("Syska LED", "https://www.syskaled.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2012, 40, "LED brand"),
    SeedBrand("Wipro Lighting", "https://www.wiprolighting.com", "Electronics", "Lighting", "Bengaluru", "Karnataka", 2010, 60, "Lighting brand"),
    SeedBrand("Philips Lighting India", "https://www.lighting.philips.co.in", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 80, "Lighting brand"),
    SeedBrand("Crompton Lighting", "https://www.crompton.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 40, "Lighting brand"),
    SeedBrand("Surya Roshini", "https://www.suryaroshini.com", "Electronics", "Lighting", "New Delhi", "Delhi", 2010, 30, "Lighting brand"),
    SeedBrand("Havells Lighting", "https://www.havells.com", "Electronics", "Lighting", "Noida", "Uttar Pradesh", 2010, 60, "Lighting brand"),
    SeedBrand("Orient Lighting", "https://www.orientelectric.com", "Electronics", "Lighting", "New Delhi", "Delhi", 2010, 30, "Lighting brand"),
    SeedBrand("Polycab India", "https://www.polycab.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 80, "Lighting brand"),
    SeedBrand("Finolex Cables", "https://www.finolex.com", "Electronics", "Lighting", "Pune", "Maharashtra", 2010, 60, "Lighting brand"),
    SeedBrand("RR Kabel", "https://www.rrkabel.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 40, "Lighting brand"),
    SeedBrand("KEI Industries", "https://www.kei-industries.com", "Electronics", "Lighting", "New Delhi", "Delhi", 2010, 60, "Lighting brand"),
    SeedBrand("Anchor by Panasonic", "https://www.anchor.co.in", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 40, "Lighting brand"),
    SeedBrand("Legrand India", "https://www.legrand.co.in", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 80, "Lighting brand"),
    SeedBrand("Schneider Electric India", "https://www.se.com", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 100, "Lighting brand"),
    SeedBrand("Siemens India", "https://www.siemens.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 100, "Lighting brand"),
    SeedBrand("ABB India", "https://www.abb.com", "Electronics", "Lighting", "Bengaluru", "Karnataka", 2010, 100, "Lighting brand"),
    SeedBrand("Eaton India", "https://www.eaton.com", "Electronics", "Lighting", "Pune", "Maharashtra", 2010, 80, "Lighting brand"),
    SeedBrand("Honeywell India", "https://www.honeywell.com", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 100, "Lighting brand"),
    SeedBrand("Johnson Controls India", "https://www.johnsoncontrols.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 80, "Lighting brand"),
    SeedBrand("Carrier India", "https://www.carrier.com", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 80, "Lighting brand"),
    SeedBrand("Daikin India", "https://www.daikin.com", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 80, "Lighting brand"),
    SeedBrand("Voltas", "https://www.voltas.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 100, "Lighting brand"),
    SeedBrand("Blue Star India", "https://www.bluestarindia.com", "Electronics", "Lighting", "Mumbai", "Maharashtra", 2010, 80, "Lighting brand"),
    SeedBrand("Hitachi India", "https://www.hitachi.com", "Electronics", "Lighting", "New Delhi", "Delhi", 2010, 80, "Lighting brand"),
    SeedBrand("Mitsubishi Electric India", "https://www.mitsubishielectric.com", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 60, "Lighting brand"),
    SeedBrand("Toshiba India", "https://www.toshiba.com", "Electronics", "Lighting", "New Delhi", "Delhi", 2010, 60, "Lighting brand"),
    SeedBrand("Sanyo India", "https://www.sanyo.com", "Electronics", "Lighting", "Gurugram", "Haryana", 2010, 40, "Lighting brand"),
    SeedBrand("Videocon", "https://www.videocon.com", "Electronics", "Appliances", "Mumbai", "Maharashtra", 2010, 60, "Electronics brand"),
    SeedBrand("Onida", "https://www.onida.com", "Electronics", "Appliances", "Mumbai", "Maharashtra", 2010, 40, "Electronics brand"),
    SeedBrand("Kelvinator", "https://www.kelvinator.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 40, "Electronics brand"),
    SeedBrand("IFB Industries", "https://www.ifb.in", "Electronics", "Appliances", "Kolkata", "West Bengal", 2010, 60, "Electronics brand"),
    SeedBrand("Bosch India", "https://www.bosch.co.in", "Electronics", "Appliances", "Bengaluru", "Karnataka", 2010, 100, "Electronics brand"),
    SeedBrand("Whirlpool India", "https://www.whirlpool.co.in", "Electronics", "Appliances", "Gurugram", "Haryana", 2010, 100, "Electronics brand"),
    SeedBrand("Electrolux India", "https://www.electrolux.com", "Electronics", "Appliances", "Gurugram", "Haryana", 2010, 60, "Electronics brand"),
    SeedBrand("Godrej Appliances", "https://www.godrej.com", "Electronics", "Appliances", "Mumbai", "Maharashtra", 2010, 100, "Electronics brand"),
    SeedBrand("Haier India", "https://www.haier.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 80, "Electronics brand"),
    SeedBrand("TCL India", "https://www.tcl.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 60, "Electronics brand"),
    SeedBrand("Hisense India", "https://www.hisense.com", "Electronics", "Appliances", "New Delhi", "Delhi", 2018, 30, "Electronics brand"),
    SeedBrand("TCL India", "https://www.tcl.com/in", "Electronics", "Appliances", "New Delhi", "Delhi", 2010, 60, "Electronics brand"),
    SeedBrand("Nokia India", "https://www.nokia.com", "Electronics", "Smartphones", "New Delhi", "Delhi", 2017, 40, "Smartphone brand"),
    SeedBrand("Asus India", "https://www.asus.com", "Electronics", "Laptops", "Bengaluru", "Karnataka", 2010, 60, "Laptop brand"),
    SeedBrand("Acer India", "https://www.acer.com", "Electronics", "Laptops", "Bengaluru", "Karnataka", 2010, 60, "Laptop brand"),
    SeedBrand("Dell India", "https://www.dell.com", "Electronics", "Laptops", "Hyderabad", "Telangana", 2010, 200, "Laptop brand"),
    SeedBrand("HP India", "https://www.hp.com", "Electronics", "Laptops", "Bengaluru", "Karnataka", 2010, 200, "Laptop brand"),
    SeedBrand("Lenovo India", "https://www.lenovo.com", "Electronics", "Laptops", "Bengaluru", "Karnataka", 2010, 200, "Laptop brand"),
    SeedBrand("Apple India", "https://www.apple.com", "Electronics", "Smartphones", "Bengaluru", "Karnataka", 2010, 200, "Electronics brand"),
    SeedBrand("Google India", "https://www.google.com", "Electronics", "Smartphones", "Bengaluru", "Karnataka", 2010, 200, "Electronics brand"),

    # === BABY PRODUCTS — More D2C Brands ===
    SeedBrand("FirstCry", "https://www.firstcry.com", "Baby Products", "Multi-Category", "Pune", "Maharashtra", 2010, 150, "Baby products"),
    SeedBrand("Mothercare India", "https://www.mothercare.com", "Baby Products", "Multi-Category", "Mumbai", "Maharashtra", 2010, 60, "Baby products"),
    SeedBrand("BabyOye", "https://www.babyoye.com", "Baby Products", "Multi-Category", "Mumbai", "Maharashtra", 2012, 30, "Baby products"),
    SeedBrand("Hopscotch", "https://www.hopscotch.in", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2014, 40, "Kids fashion"),
    SeedBrand("Cucumber Clothing", "https://www.cucumberclothing.com", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2015, 10, "Kids fashion"),
    SeedBrand("Little Muffet", "https://www.littlemuffet.com", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2016, 8, "Kids fashion"),
    SeedBrand("Bubble Blue", "https://www.bubbleblue.com", "Baby Products", "Kids Fashion", "Bengaluru", "Karnataka", 2017, 6, "Kids fashion"),
    SeedBrand("Babies B Store", "https://www.babiesbstore.com", "Baby Products", "Kids Fashion", "Mumbai", "Maharashtra", 2018, 5, "Kids fashion"),
    SeedBrand("R for Rabbit", "https://www.rforgabbit.com", "Baby Products", "Baby Gear", "Ahmedabad", "Gujarat", 2015, 20, "Baby gear"),
    SeedBrand("LuvLap", "https://www.luvlap.com", "Baby Products", "Baby Care", "New Delhi", "Delhi", 2012, 25, "Baby care"),
    SeedBrand("Baybee", "https://www.baybee.in", "Baby Products", "Baby Gear", "New Delhi", "Delhi", 2016, 10, "Baby gear"),
    SeedBrand("Mee Mee", "https://www.mfrfeemee.in", "Baby Products", "Baby Care", "New Delhi", "Delhi", 2012, 20, "Baby care"),
    SeedBrand("Chicco India", "https://www.chicco.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 40, "Baby care"),
    SeedBrand("Pampers India", "https://www.pampers.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 100, "Baby care"),
    SeedBrand("Huggies India", "https://www.huggies.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 100, "Baby care"),
    SeedBrand("Johnson's Baby India", "https://www.johnsons baby.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 100, "Baby care"),
    SeedBrand("MamyPoko India", "https://www.mfrfypoko.com", "Baby Products", "Baby Care", "Gurugram", "Haryana", 2010, 80, "Baby care"),
    SeedBrand("Sebamed India", "https://www.sebamed.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 30, "Baby care"),
    SeedBrand("Mustela India", "https://www.mustela.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 20, "Baby care"),
    SeedBrand("Mothercare", "https://www.mothercare.com", "Baby Products", "Baby Care", "Mumbai", "Maharashtra", 2010, 40, "Baby care"),

    # === PET PRODUCTS — More D2C Brands ===
    SeedBrand("Drools", "https://www.drools.com", "Pet Products", "Pet Food", "Hyderabad", "Telangana", 2015, 60, "Pet food"),
    SeedBrand("Farmina India", "https://www.farmina.com", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2010, 40, "Pet food"),
    SeedBrand("Royal Canin India", "https://www.royalcanin.com", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2010, 100, "Pet food"),
    SeedBrand("Pedigree India", "https://www.pedigree.com", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2010, 100, "Pet food"),
    SeedBrand("Whiskas India", "https://www.whiskas.com", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2010, 80, "Pet food"),
    SeedBrand("Meow Mix India", "https://www.meowmix.com", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2010, 40, "Pet food"),
    SeedBrand("Catit India", "https://www.catit.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2012, 20, "Pet accessories"),
    SeedBrand("PetSafe India", "https://www.petsafe.net", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2010, 30, "Pet accessories"),
    SeedBrand("KONG India", "https://www.kongcompany.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2010, 20, "Pet accessories"),
    SeedBrand("Outward Hound India", "https://www.outwardhound.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2010, 15, "Pet accessories"),
    SeedBrand("PetSafe India", "https://www.petsafe.in", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2015, 10, "Pet accessories"),
    SeedBrand("Heads Up For Tails", "https://www.headsuptails.com", "Pet Products", "Pet Accessories", "Mumbai", "Maharashtra", 2016, 25, "Pet accessories"),
    SeedBrand("Wiggles", "https://www.wiggles.in", "Pet Products", "Pet Care", "Mumbai", "Maharashtra", 2018, 15, "Pet care"),
    SeedBrand("Paws & Claws", "https://www.pawsandclaws.in", "Pet Products", "Pet Food", "Mumbai", "Maharashtra", 2016, 8, "Pet food"),
    SeedBrand("YoPets", "https://www.yopets.in", "Pet Products", "Pet Accessories", "New Delhi", "Delhi", 2018, 6, "Pet accessories"),
    SeedBrand("PetCraft", "https://www.petcraft.in", "Pet Products", "Pet Accessories", "Bengaluru", "Karnataka", 2017, 8, "Pet accessories"),

    # === FOOTWEAR — More D2C Brands ===
    SeedBrand("Bata India", "https://www.bata.com", "Footwear", "Multi-Category", "Gurugram", "Haryana", 2010, 200, "Footwear retail"),
    SeedBrand("Relaxo Footwears", "https://www.relaxo.com", "Footwear", "Multi-Category", "New Delhi", "Delhi", 2010, 100, "Footwear brand"),
    SeedBrand("Liberty Shoes", "https://www.libertyshoes.com", "Footwear", "Multi-Category", "Karnal", "Haryana", 2010, 80, "Footwear brand"),
    SeedBrand("Paragon", "https://www.paragon.com", "Footwear", "Multi-Category", "Kozhikode", "Kerala", 2010, 60, "Footwear brand"),
    SeedBrand("Action", "https://www.action.com", "Footwear", "Multi-Category", "Bengaluru", "Karnataka", 2010, 40, "Footwear brand"),
    SeedBrand("Campus Shoes", "https://www.campusshoes.com", "Footwear", "Sports", "Faridabad", "Haryana", 2010, 60, "Footwear brand"),
    SeedBrand("Sparx", "https://www.sparxshoes.com", "Footwear", "Sports", "Faridabad", "Haryana", 2010, 50, "Sports footwear"),
    SeedBrand("Skechers India", "https://www.skechers.com", "Footwear", "Casual", "Gurugram", "Haryana", 2012, 60, "Casual footwear"),
    SeedBrand("Clarks India", "https://www.clarks.com", "Footwear", "Formal", "Mumbai", "Maharashtra", 2010, 40, "Formal footwear"),
    SeedBrand("Woodland", "https://www.woodlandworld.com", "Footwear", "Outdoor", "Gurugram", "Haryana", 2010, 80, "Outdoor footwear"),
    SeedBrand("Camper India", "https://www.camper.com", "Footwear", "Casual", "Mumbai", "Maharashtra", 2012, 30, "Casual footwear"),
    SeedBrand("Hush Puppies India", "https://www.hushpuppies.com", "Footwear", "Formal", "Mumbai", "Maharashtra", 2010, 40, "Formal footwear"),
    SeedBrand("Cole Haan India", "https://www.colehaan.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2015, 20, "Premium footwear"),
    SeedBrand("Steve Madden India", "https://www.stevemadden.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2015, 25, "Premium footwear"),
    SeedBrand("Aldo India", "https://www.aldoshoes.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2015, 20, "Premium footwear"),
    SeedBrand("H&M India", "https://www.hm.com", "Footwear", "Fast Fashion", "New Delhi", "Delhi", 2015, 100, "Fast fashion"),
    SeedBrand("Zara India", "https://www.zara.com", "Footwear", "Fast Fashion", "New Delhi", "Delhi", 2010, 100, "Fast fashion"),
    SeedBrand("Uniqlo India", "https://www.uniqlo.com", "Footwear", "Fast Fashion", "New Delhi", "Delhi", 2019, 60, "Fast fashion"),
    SeedBrand("Marks & Spencer India", "https://www.marksandspencer.in", "Footwear", "Premium", "Mumbai", "Maharashtra", 2010, 80, "Premium fashion"),
    SeedBrand("Tommy Hilfiger India", "https://www.tommy.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2010, 60, "Premium fashion"),
    SeedBrand("Calvin Klein India", "https://www.calvinklein.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2010, 40, "Premium fashion"),
    SeedBrand("Guess India", "https://www.guess.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2012, 30, "Premium fashion"),
    SeedBrand("Coach India", "https://www.coach.com", "Footwear", "Luxury", "Mumbai", "Maharashtra", 2015, 40, "Luxury fashion"),
    SeedBrand("Michael Kors India", "https://www.michaelkors.com", "Footwear", "Luxury", "Mumbai", "Maharashtra", 2015, 30, "Luxury fashion"),
    SeedBrand("Kate Spade India", "https://www.katespade.com", "Footwear", "Luxury", "Mumbai", "Maharashtra", 2016, 20, "Luxury fashion"),
    SeedBrand("Fossil India", "https://www.fossil.com", "Footwear", "Premium", "Mumbai", "Maharashtra", 2012, 30, "Premium fashion"),
    SeedBrand("Timberland India", "https://www.timberland.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2010, 40, "Outdoor footwear"),
    SeedBrand("Dr. Martens India", "https://www.drmartens.com", "Footwear", "Casual", "Mumbai", "Maharashtra", 2015, 25, "Casual footwear"),
    SeedBrand("Converse India", "https://www.converse.com", "Footwear", "Casual", "Mumbai", "Maharashtra", 2010, 60, "Casual footwear"),
    SeedBrand("Vans India", "https://www.vans.com", "Footwear", "Casual", "Mumbai", "Maharashtra", 2012, 40, "Casual footwear"),
    SeedBrand("Fila India", "https://www.fila.com", "Footwear", "Sports", "Mumbai", "Maharashtra", 2015, 30, "Sports footwear"),
    SeedBrand("ASICS India", "https://www.asics.com", "Footwear", "Sports", "Mumbai", "Maharashtra", 2010, 40, "Sports footwear"),
    SeedBrand("New Balance India", "https://www.newbalance.com", "Footwear", "Sports", "Mumbai", "Maharashtra", 2015, 30, "Sports footwear"),
    SeedBrand("Brooks India", "https://www.brooks.com", "Footwear", "Sports", "Mumbai", "Maharashtra", 2018, 15, "Sports footwear"),
    SeedBrand("Salomon India", "https://www.salomon.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2018, 20, "Outdoor footwear"),
    SeedBrand("Merrell India", "https://www.merrell.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2012, 25, "Outdoor footwear"),
    SeedBrand("Columbia India", "https://www.columbia.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2012, 30, "Outdoor footwear"),
    SeedBrand("The North Face India", "https://www.thenorthface.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2010, 60, "Outdoor footwear"),
    SeedBrand("Jack Wolfskin India", "https://www.jack-wolfskin.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2012, 20, "Outdoor footwear"),
    SeedBrand("Mountain Hardwear India", "https://www.mountainhardwear.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2015, 15, "Outdoor footwear"),
    SeedBrand("Mammut India", "https://www.mammut.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2018, 10, "Outdoor footwear"),
    SeedBrand("Arc'teryx India", "https://www.arcteryx.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2018, 15, "Outdoor footwear"),
    SeedBrand("Patagonia India", "https://www.patagonia.com", "Footwear", "Outdoor", "Mumbai", "Maharashtra", 2015, 20, "Outdoor footwear"),
    SeedBrand("Reebok India", "https://www.reebok.in", "Footwear", "Sports", "Gurugram", "Haryana", 2010, 80, "Sports footwear"),
    SeedBrand("Under Armour India", "https://www.underarmour.co.in", "Footwear", "Sports", "Gurugram", "Haryana", 2010, 60, "Sports footwear"),
    SeedBrand("Puma India", "https://www.puma.com", "Footwear", "Sports", "Bengaluru", "Karnataka", 2010, 100, "Sports footwear"),
    SeedBrand("Adidas India", "https://www.adidas.co.in", "Footwear", "Sports", "Bengaluru", "Karnataka", 2010, 200, "Sports footwear"),
    SeedBrand("Nike India", "https://www.nike.com", "Footwear", "Sports", "Gurugram", "Haryana", 2010, 200, "Sports footwear"),
    SeedBrand("Decathlon India", "https://www.decathlon.in", "Footwear", "Sports", "Bengaluru", "Karnataka", 2010, 200, "Sports retail"),

    # === BAGS — More D2C Brands ===
    SeedBrand("Wildcraft", "https://www.wildcraft.com", "Bags", "Backpacks", "Bengaluru", "Karnataka", 2010, 80, "Outdoor gear"),
    SeedBrand("Skybags", "https://www.skybags.in", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 60, "Luggage brand"),
    SeedBrand("American Tourister India", "https://www.americantourister.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 80, "Luggage brand"),
    SeedBrand("Samsonite India", "https://www.samsonite.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2010, 100, "Luggage brand"),
    SeedBrand("Delsey India", "https://www.delsey.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2012, 30, "Luggage brand"),
    SeedBrand("Antler India", "https://www.antler.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2015, 20, "Luggage brand"),
    SeedBrand("Pluggage India", "https://www.pluggage.com", "Bags", "Luggage", "Mumbai", "Maharashtra", 2018, 10, "Luggage brand"),
    SeedBrand("FurJaden", "https://www.furjaden.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2017, 12, "Fashion bags"),
    SeedBrand("Lavie World", "https://www.lavieworld.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 30, "Fashion bags"),
    SeedBrand("Hidesign", "https://www.hidesign.com", "Bags", "Leather Goods", "New Delhi", "Delhi", 2010, 60, "Leather accessories"),
    SeedBrand("Caprese", "https://www.caprese.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2012, 40, "Premium bags"),
    SeedBrand("Baggit", "https://www.baggit.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2010, 30, "Fashion bags"),
    SeedBrand("Charles & Keith India", "https://www.charleskeith.com", "Bags", "Handbags", "Mumbai", "Maharashtra", 2015, 25, "Premium bags"),
    SeedBrand("Michael Kors India", "https://www.michaelkors.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 40, "Luxury bags"),
    SeedBrand("Coach India", "https://www.coach.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 30, "Luxury bags"),
    SeedBrand("Kate Spade India", "https://www.katespade.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2016, 20, "Luxury bags"),
    SeedBrand("Guess India", "https://www.guess.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2012, 25, "Premium bags"),
    SeedBrand("Fossil India", "https://www.fossil.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2012, 20, "Premium bags"),
    SeedBrand("Tommy Hilfiger India", "https://www.tommy.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2010, 40, "Premium bags"),
    SeedBrand("Calvin Klein India", "https://www.calvinklein.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2010, 30, "Premium bags"),
    SeedBrand("Gucci India", "https://www.gucci.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 80, "Luxury bags"),
    SeedBrand("Prada India", "https://www.prada.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 60, "Luxury bags"),
    SeedBrand("Louis Vuitton India", "https://www.louisvuitton.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2010, 100, "Luxury bags"),
    SeedBrand("Burberry India", "https://www.burberry.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 40, "Luxury bags"),
    SeedBrand("Chanel India", "https://www.chanel.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 80, "Luxury bags"),
    SeedBrand("Dior India", "https://www.dior.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 60, "Luxury bags"),
    SeedBrand("Hermes India", "https://www.hermes.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 100, "Luxury bags"),
    SeedBrand("Bottega Veneta India", "https://www.bottegaveneta.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 40, "Luxury bags"),
    SeedBrand("Loewe India", "https://www.loewe.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 30, "Luxury bags"),
    SeedBrand("Celine India", "https://www.celine.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 30, "Luxury bags"),
    SeedBrand("Saint Laurent India", "https://www.ysl.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 40, "Luxury bags"),
    SeedBrand("Balenciaga India", "https://www.balenciaga.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 30, "Luxury bags"),
    SeedBrand("Givenchy India", "https://www.givenchy.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 25, "Luxury bags"),
    SeedBrand("Fendi India", "https://www.fendi.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 30, "Luxury bags"),
    SeedBrand("Versace India", "https://www.versace.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 25, "Luxury bags"),
    SeedBrand("Valentino India", "https://www.valentino.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 20, "Luxury bags"),
    SeedBrand("Dolce & Gabbana India", "https://www.dolcegabbana.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 25, "Luxury bags"),
    SeedBrand("Armani India", "https://www.armani.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2015, 30, "Luxury bags"),
    SeedBrand("Ralph Lauren India", "https://www.ralphlauren.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2012, 40, "Premium bags"),
    SeedBrand("Coach India", "https://www.coach.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2015, 30, "Premium bags"),
    SeedBrand("Tumi India", "https://www.tumi.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2015, 25, "Premium bags"),
    SeedBrand("Briggs & Riley India", "https://www.briggs-riley.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2018, 15, "Premium bags"),
    SeedBrand("Rimowa India", "https://www.rimowa.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 30, "Luxury luggage"),
    SeedBrand("Zero Halliburton India", "https://www.zerohalliburton.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2015, 10, "Premium luggage"),
    SeedBrand("Victorinox India", "https://www.victorinox.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2012, 20, "Premium luggage"),
    SeedBrand("Porsche Design India", "https://www.porsche.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 20, "Luxury bags"),
    SeedBrand("BMW Lifestyle India", "https://www.bmw.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Mercedes-Benz India", "https://www.mercedes-benz.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Audi India", "https://www.audi.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Lamborghini India", "https://www.lamborghini.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Ferrari India", "https://www.ferrari.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Maserati India", "https://www.maserati.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 10, "Luxury bags"),
    SeedBrand(" Aston Martin India", "https://www.astonmartin.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 10, "Luxury bags"),
    SeedBrand("Bentley India", "https://www.bentley.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Rolls-Royce India", "https://www.rolls-royce.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 15, "Luxury bags"),
    SeedBrand("Bugatti India", "https://www.bugatti.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 10, "Luxury bags"),
    SeedBrand("Pagani India", "https://www.pagani.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 5, "Luxury bags"),
    SeedBrand("Koenigsegg India", "https://www.koenigsegg.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 5, "Luxury bags"),
    SeedBrand("McLaren India", "https://www.mclaren.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 10, "Luxury bags"),
    SeedBrand("Lotus India", "https://www.lotuscars.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 5, "Luxury bags"),
    SeedBrand("Alfa Romeo India", "https://www.alfaromeo.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2018, 10, "Luxury bags"),
    SeedBrand("Jaguar India", "https://www.jaguar.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2010, 40, "Luxury bags"),
    SeedBrand("Land Rover India", "https://www.landrover.com", "Bags", "Luxury", "Mumbai", "Maharashtra", 2010, 60, "Luxury bags"),
    SeedBrand("Volvo India", "https://www.volvocars.com", "Bags", "Luxury", "Bengaluru", "Karnataka", 2010, 40, "Luxury bags"),
    SeedBrand("Mini India", "https://www.mini.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2012, 30, "Premium bags"),
    SeedBrand(" Fiat India", "https://www.fiat.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2010, 40, "Premium bags"),
    SeedBrand("Jeep India", "https://www.jeep.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2016, 40, "Premium bags"),
    SeedBrand("Toyota India", "https://www.toyota.com", "Bags", "Premium", "Bengaluru", "Karnataka", 2010, 200, "Premium bags"),
    SeedBrand("Honda India", "https://www.honda.com", "Bags", "Premium", "New Delhi", "Delhi", 2010, 200, "Premium bags"),
    SeedBrand("Hyundai India", "https://www.hyundai.com", "Bags", "Premium", "Chennai", "Tamil Nadu", 2010, 200, "Premium bags"),
    SeedBrand("Kia India", "https://www.kia.com", "Bags", "Premium", "Anantapur", "Andhra Pradesh", 2019, 100, "Premium bags"),
    SeedBrand("MG Motor India", "https://www.mgmotor.co.in", "Bags", "Premium", "Halol", "Gujarat", 2019, 60, "Premium bags"),
    SeedBrand("Skoda India", "https://www.skoda-auto.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2010, 60, "Premium bags"),
    SeedBrand("Volkswagen India", "https://www.volkswagen.co.in", "Bags", "Premium", "Pune", "Maharashtra", 2010, 80, "Premium bags"),
    SeedBrand("Renault India", "https://www.renault.co.in", "Bags", "Premium", "Chennai", "Tamil Nadu", 2010, 60, "Premium bags"),
    SeedBrand("Nissan India", "https://www.nissanmotor.co.in", "Bags", "Premium", "Chennai", "Tamil Nadu", 2010, 60, "Premium bags"),
    SeedBrand("Datsun India", "https://www.datsun.com", "Bags", "Budget", "Chennai", "Tamil Nadu", 2014, 30, "Budget cars"),
    SeedBrand("Ford India", "https://www.ford.com", "Bags", "Premium", "Chennai", "Tamil Nadu", 2010, 80, "Premium cars"),
    SeedBrand("Chevrolet India", "https://www.chevrolet.com", "Bags", "Premium", "Gurugram", "Haryana", 2010, 60, "Premium cars"),
    SeedBrand("Isuzu India", "https://www.isuzu.co.in", "Bags", "Premium", "Chennai", "Tamil Nadu", 2012, 30, "Premium cars"),
    SeedBrand("Mitsubishi India", "https://www.mitsubishi-motors.com", "Bags", "Premium", "New Delhi", "Delhi", 2010, 20, "Premium cars"),
    SeedBrand("SsangYong India", "https://www.ssangyong.com", "Bags", "Premium", "Mumbai", "Maharashtra", 2018, 10, "Premium cars"),

    # === SPORTS — More D2C Brands ===
    SeedBrand("Nivia", "https://www.nivia.com", "Sports", "Sports Equipment", "Jalandhar", "Punjab", 2010, 40, "Sports equipment"),
    SeedBrand("Cosco", "https://www.coscosports.com", "Sports", "Sports Equipment", "New Delhi", "Delhi", 2010, 30, "Sports goods"),
    SeedBrand("Yonex India", "https://www.yonex.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 40, "Sports equipment"),
    SeedBrand("Wilson India", "https://www.wilson.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 30, "Sports equipment"),
    SeedBrand("Head India", "https://www.head.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 25, "Sports equipment"),
    SeedBrand("Babolat India", "https://www.babolat.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2012, 15, "Sports equipment"),
    SeedBrand("Dunlop India", "https://www.dunlop.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 20, "Sports equipment"),
    SeedBrand("Slazenger India", "https://www.slazenger.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 15, "Sports equipment"),
    SeedBrand("MRF India", "https://www.mfrfrf.com", "Sports", "Sports Equipment", "Chennai", "Tamil Nadu", 2010, 40, "Sports equipment"),
    SeedBrand("SG India", "https://www.sfrfgcrfricket.com", "Sports", "Sports Equipment", "Jalandhar", "Punjab", 2010, 30, "Cricket equipment"),
    SeedBrand("SS India", "https://www.sfrfsscricket.com", "Sports", "Sports Equipment", "Jalandhar", "Punjab", 2010, 30, "Cricket equipment"),
    SeedBrand("BAS India", "https://www.bfrfsports.com", "Sports", "Sports Equipment", "Meerut", "Uttar Pradesh", 2010, 20, "Sports equipment"),
    SeedBrand("Vector X", "https://www.vectorx.com", "Sports", "Sports Equipment", "Meerut", "Uttar Pradesh", 2012, 10, "Sports equipment"),
    SeedBrand("Zee Sports", "https://www.zeesports.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 15, "Sports equipment"),
    SeedBrand("Spalding India", "https://www.spalding.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 20, "Sports equipment"),
    SeedBrand("Molten India", "https://www.molten.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 15, "Sports equipment"),
    SeedBrand("Adidas India", "https://www.adidas.co.in", "Sports", "Sports Equipment", "Bengaluru", "Karnataka", 2010, 100, "Sports brand"),
    SeedBrand("Nike India", "https://www.nike.com", "Sports", "Sports Equipment", "Gurugram", "Haryana", 2010, 100, "Sports brand"),
    SeedBrand("Puma India", "https://www.puma.com", "Sports", "Sports Equipment", "Bengaluru", "Karnataka", 2010, 80, "Sports brand"),
    SeedBrand("Reebok India", "https://www.reebok.in", "Sports", "Sports Equipment", "Gurugram", "Haryana", 2010, 60, "Sports brand"),
    SeedBrand("Under Armour India", "https://www.underarmour.co.in", "Sports", "Sports Equipment", "Gurugram", "Haryana", 2010, 50, "Sports brand"),
    SeedBrand("New Balance India", "https://www.newbalance.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2015, 30, "Sports brand"),
    SeedBrand("Saucony India", "https://www.saucony.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2018, 10, "Sports brand"),
    SeedBrand("Brooks India", "https://www.brooks.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2018, 10, "Sports brand"),
    SeedBrand("ASICS India", "https://www.asics.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 30, "Sports brand"),
    SeedBrand("Fila India", "https://www.fila.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2015, 20, "Sports brand"),
    SeedBrand("K-Swiss India", "https://www.k-swiss.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2015, 10, "Sports brand"),
    SeedBrand("Diadora India", "https://www.diadora.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2012, 10, "Sports brand"),
    SeedBrand("Umbro India", "https://www.umbro.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2012, 10, "Sports brand"),
    SeedBrand("Le Coq Sportif India", "https://www.lecoqsportif.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2015, 8, "Sports brand"),
    SeedBrand("Lonsdale India", "https://www.lonsdale.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2015, 5, "Sports brand"),
    SeedBrand("Everlast India", "https://www.everlast.com", "Sports", "Sports Equipment", "Mumbai", "Maharashtra", 2010, 15, "Sports brand"),
    SeedBrand("Speedo India", "https://www.speedo.com", "Sports", "Swimwear", "Mumbai", "Maharashtra", 2010, 20, "Swimwear brand"),
    SeedBrand("Arena India", "https://www.arena.com", "Sports", "Swimwear", "Mumbai", "Maharashtra", 2010, 15, "Swimwear brand"),
    SeedBrand("Nike Swim India", "https://www.nike.com", "Sports", "Swimwear", "Gurugram", "Haryana", 2010, 20, "Swimwear brand"),
    SeedBrand("Decathlon India", "https://www.decathlon.in", "Sports", "Multi-Category", "Bengaluru", "Karnataka", 2010, 150, "Sports retail"),
    SeedBrand("Sports Authority India", "https://www.sportsauthority.com", "Sports", "Multi-Category", "Mumbai", "Maharashtra", 2010, 30, "Sports retail"),
    SeedBrand("ProSports India", "https://www.prosports.in", "Sports", "Multi-Category", "Mumbai", "Maharashtra", 2015, 10, "Sports retail"),

    # === GIFTS — More D2C Brands ===
    SeedBrand("Ferns N Petals", "https://www.fernnpetals.com", "Gifts", "Flowers & Gifts", "New Delhi", "Delhi", 2010, 80, "Flowers and gifts"),
    SeedBrand("IGP", "https://www.igp.com", "Gifts", "Gifts & Flowers", "Mumbai", "Maharashtra", 2010, 60, "Gifts and flowers"),
    SeedBrand("Winni", "https://www.winni.in", "Gifts", "Gifts & Flowers", "New Delhi", "Delhi", 2014, 15, "Gifts and cakes"),
    SeedBrand("Cherrytin", "https://www.cherrytin.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2017, 6, "Personalized gifts"),
    SeedBrand("IGP Gifts", "https://www.igpgifts.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2015, 20, "Corporate gifts"),
    SeedBrand("FNP Gifts", "https://www.fnpgifts.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2015, 15, "Corporate gifts"),
    SeedBrand("Archies India", "https://www.archiesonline.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2010, 40, "Greeting cards and gifts"),
    SeedBrand("Hallmark India", "https://www.hallmark.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2010, 30, "Greeting cards and gifts"),
    SeedBrand("American Greetings India", "https://www.americangreetings.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2010, 20, "Greeting cards"),
    SeedBrand("Papyrus India", "https://www.papyrus.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2012, 10, "Greeting cards"),
    SeedBrand("Myntra Gifts", "https://www.myntra.com", "Gifts", "Gifts", "Bengaluru", "Karnataka", 2015, 30, "Online gifts"),
    SeedBrand("Flipkart Gifts", "https://www.flipkart.com", "Gifts", "Gifts", "Bengaluru", "Karnataka", 2015, 40, "Online gifts"),
    SeedBrand("Amazon India Gifts", "https://www.amazon.in", "Gifts", "Gifts", "Bengaluru", "Karnataka", 2010, 100, "Online gifts"),
    SeedBrand("Snapdeal Gifts", "https://www.snapdeal.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2015, 20, "Online gifts"),
    SeedBrand("Paytm Gifts", "https://www.paytmmall.com", "Gifts", "Gifts", "Noida", "Uttar Pradesh", 2015, 30, "Online gifts"),
    SeedBrand("ShopClues Gifts", "https://www.shopclues.com", "Gifts", "Gifts", "Gurugram", "Haryana", 2015, 15, "Online gifts"),
    SeedBrand("Craftsvilla Gifts", "https://www.craftsvilla.com", "Gifts", "Gifts", "Mumbai", "Maharashtra", 2012, 20, "Handicraft gifts"),
    SeedBrand("Jaypore Gifts", "https://www.jaypore.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2015, 15, "Handicraft gifts"),
    SeedBrand("The India Craft House", "https://www.theindiacrafthouse.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2016, 8, "Handicraft gifts"),
    SeedBrand("Okhai", "https://www.okhai.org", "Gifts", "Gifts", "Ahmedabad", "Gujarat", 2014, 10, "Handicraft gifts"),
    SeedBrand("Dastkar", "https://www.dastkar.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2010, 8, "Handicraft gifts"),
    SeedBrand("GoCoop", "https://www.gocoop.com", "Gifts", "Gifts", "Bengaluru", "Karnataka", 2012, 12, "Handicraft gifts"),
    SeedBrand("Indiamart Gifts", "https://www.indiamart.com", "Gifts", "Gifts", "Noida", "Uttar Pradesh", 2010, 80, "B2B gifts"),
    SeedBrand("TradeIndia Gifts", "https://www.tradeindia.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2010, 40, "B2B gifts"),
    SeedBrand("ExportersIndia Gifts", "https://www.exportersindia.com", "Gifts", "Gifts", "New Delhi", "Delhi", 2010, 30, "B2B gifts"),

    # === LIFESTYLE — More D2C Brands ===
    SeedBrand("Furrl", "https://www.furrl.in", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2019, 12, "Lifestyle discovery"),
    SeedBrand("The Label Life", "https://www.thelabellife.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2015, 25, "Curated lifestyle"),
    SeedBrand("Jaypore", "https://www.jaypore.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2014, 40, "Lifestyle brand"),
    SeedBrand("Chumbak", "https://www.chumbak.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2011, 30, "Quirky lifestyle"),
    SeedBrand("The Souled Store", "https://www.thesouledstore.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2013, 50, "Pop culture lifestyle"),
    SeedBrand("Bewakoof", "https://www.bewakoof.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2012, 60, "Online lifestyle"),
    SeedBrand("FabAlley", "https://www.faballey.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2012, 30, "Women lifestyle"),
    SeedBrand("Clovia", "https://www.clovia.com", "Lifestyle", "Multi-Category", "Noida", "Uttar Pradesh", 2013, 40, "Lingerie lifestyle"),
    SeedBrand("Zivame", "https://www.zivame.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2013, 50, "Lingerie lifestyle"),
    SeedBrand("Fynd", "https://www.fynd.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2012, 40, "Omnichannel lifestyle"),
    SeedBrand("Limeroad", "https://www.limeroad.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2012, 30, "Fashion lifestyle"),
    SeedBrand("Voonik", "https://www.voonik.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2013, 25, "Fashion lifestyle"),
    SeedBrand("Roposo", "https://www.roposo.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2012, 20, "Fashion lifestyle"),
    SeedBrand("Meesho", "https://www.meesho.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2015, 80, "Social commerce"),
    SeedBrand("Shop101", "https://www.shop101.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2015, 30, "Social commerce"),
    SeedBrand("GlowRoad", "https://www.glowroad.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2017, 20, "Social commerce"),
    SeedBrand("MensXP", "https://www.mensxp.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2012, 30, "Men lifestyle"),
    SeedBrand("POPxo", "https://www.popxo.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2014, 20, "Women lifestyle"),
    SeedBrand("iDiva", "https://www.idiva.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 15, "Women lifestyle"),
    SeedBrand("ScoopWhoop", "https://www.scoopwhoop.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2013, 20, "Lifestyle media"),
    SeedBrand("TheQuint", "https://www.thequint.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2015, 25, "Lifestyle media"),
    SeedBrand("Scroll.in", "https://scroll.in", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2014, 15, "Lifestyle media"),
    SeedBrand("The Wire", "https://thewire.in", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2015, 10, "Lifestyle media"),
    SeedBrand("Alt News", "https://www.altnews.in", "Lifestyle", "Multi-Category", "Ahmedabad", "Gujarat", 2017, 8, "Lifestyle media"),
    SeedBrand("The Print", "https://theprint.in", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2017, 12, "Lifestyle media"),
    SeedBrand("OpIndia", "https://www.opindia.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2017, 10, "Lifestyle media"),
    SeedBrand("Swarajya", "https://www.swarajyamag.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2014, 8, "Lifestyle media"),
    SeedBrand("Newslaundry", "https://www.newslaundry.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2012, 12, "Lifestyle media"),
    SeedBrand("The Ken", "https://the-ken.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2016, 15, "Lifestyle media"),
    SeedBrand("Morning Context", "https://www.morningcontext.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2019, 10, "Lifestyle media"),
    SeedBrand("Inc42", "https://inc42.com", "Lifestyle", "Multi-Category", "Ahmedabad", "Gujarat", 2014, 8, "Lifestyle media"),
    SeedBrand("YourStory", "https://yourstory.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2010, 20, "Lifestyle media"),
    SeedBrand("Factor Daily", "https://www.factordaily.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2016, 5, "Lifestyle media"),
    SeedBrand("HuffPost India", "https://www.huffingtonpost.in", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2014, 15, "Lifestyle media"),
    SeedBrand("BuzzFeed India", "https://www.buzzfeed.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2014, 20, "Lifestyle media"),
    SeedBrand("Vice India", "https://www.vice.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2016, 10, "Lifestyle media"),
    SeedBrand("Refinery29 India", "https://www.refinery29.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2018, 8, "Lifestyle media"),
    SeedBrand("Cosmopolitan India", "https://www.cosmopolitan.in", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 30, "Lifestyle media"),
    SeedBrand("Vogue India", "https://www.vogue.in", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 40, "Lifestyle media"),
    SeedBrand("Harper's Bazaar India", "https://www.harpersbazaar.in", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 25, "Lifestyle media"),
    SeedBrand("GQ India", "https://www.gqindia.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 20, "Lifestyle media"),
    SeedBrand("Fortune India", "https://www.fortuneindia.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 30, "Lifestyle media"),
    SeedBrand("Business Today", "https://www.businesstoday.in", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 25, "Lifestyle media"),
    SeedBrand("Forbes India", "https://www.forbesindia.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 30, "Lifestyle media"),
    SeedBrand("Economic Times", "https://economictimes.indiatimes.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 80, "Lifestyle media"),
    SeedBrand("Mint", "https://www.livemint.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 60, "Lifestyle media"),
    SeedBrand("Business Standard", "https://www.business-standard.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 50, "Lifestyle media"),
    SeedBrand("Moneycontrol", "https://www.moneycontrol.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 40, "Lifestyle media"),
    SeedBrand("NDTV", "https://www.ndtv.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 60, "Lifestyle media"),
    SeedBrand("India Today", "https://www.indiatoday.in", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 50, "Lifestyle media"),
    SeedBrand("Republic World", "https://www.republicworld.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2017, 30, "Lifestyle media"),
    SeedBrand("Times Now", "https://www.timesnow.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 40, "Lifestyle media"),
    SeedBrand("Aaj Tak", "https://www.aajtak.in", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 50, "Lifestyle media"),
    SeedBrand("Zee News", "https://www.zeenews.india.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 40, "Lifestyle media"),
    SeedBrand("ABP News", "https://www.abplive.com", "Lifestyle", "Multi-Category", "Noida", "Uttar Pradesh", 2010, 30, "Lifestyle media"),
    SeedBrand("TV9 Telugu", "https://www.tv9telugu.com", "Lifestyle", "Multi-Category", "Hyderabad", "Telangana", 2015, 15, "Lifestyle media"),
    SeedBrand("TV9 Kannada", "https://www.tv9kannada.com", "Lifestyle", "Multi-Category", "Bengaluru", "Karnataka", 2015, 12, "Lifestyle media"),
    SeedBrand("TV9 Marathi", "https://www.tv9marathi.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2015, 10, "Lifestyle media"),
    SeedBrand("TV9 Hindi", "https://www.tv9hindi.com", "Lifestyle", "Multi-Category", "Noida", "Uttar Pradesh", 2018, 8, "Lifestyle media"),
    SeedBrand("News18", "https://www.news18.com", "Lifestyle", "Multi-Category", "New Delhi", "Delhi", 2010, 40, "Lifestyle media"),
    SeedBrand("CNN News18", "https://www.cnnnews18.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2010, 30, "Lifestyle media"),
    SeedBrand("Republic Bharat", "https://www.republicbharat.com", "Lifestyle", "Multi-Category", "Mumbai", "Maharashtra", 2019, 10, "Lifestyle media"),
]


# ============================================================
# REJECTION FILTER — Big players, marketplaces, non-D2C
# ============================================================

REJECT_KEYWORDS = {
    # Big multinationals
    "nike", "adidas", "puma", "reebok", "under armour", "new balance",
    "h&m", "zara", "uniqlo", "gap", "levi strauss", "ralph lauren",
    "gucci", "prada", "louis vuitton", "dior", "chanel",
    # Indian conglomerates
    "reliance", "tata", "aditya birla", "mahindra", "infosys", "wipro",
    "tcs", "hcl", "bajaj", "hero", "maruti", "ashok leyland",
    # Government / Institutions
    "government", "ministry", "department", "council", "board",
    "hospital", "medical center", "diagnostic",
    "university", "college", "school", "institute", "academy",
    "bank", "insurance", "mutual fund", "stock", "sebi",
    # Marketplaces
    "amazon", "flipkart", "meesho", "snapdeal", "myntra", "ajio",
    "nykaa marketplace", "tata cliq", "croma", "dmart", "big bazaar",
    # Offline retail chains
    "shoppers stop", "central", "max fashion",
    "pantaloons", "future group", "reliance retail",
    # B2B / SaaS / Agencies
    "consulting", "agency", "software", "saas",
    "b2b", "wholesale", "distributor", "manufacturer",
}

# Companies with known revenue > ₹250 Cr — reject
REJECT_COMPANIES = {
    "mamaearth", "honasa", "boAt", "lenskart", "nykaa", "pepperfry",
    "urban ladder", "firstcry", "meesho", "snapdeal", "bigbasket",
    "zepto", "blinkit", "swiggy", "zomato", "paytm", "phonepe",
    "cred", "razorpay", "olicy bazaar", "delhivery", "blue dart",
    "titan", "tanishq", "kalyan jewellers", "malabar gold",
    "joyalukkas", "senco gold", "pc jeweller",
    "wagh brahmi", "baidyanath", "himfrfalaya", "patanjali",
    "dabur", "emami", "marico", "godrej", "itc", "nestle",
    "britannia", "parle", "haldiram", "balaji", "bingo",
}


def is_rejected(brand: SeedBrand) -> tuple[bool, str]:
    """Check if brand should be rejected. Returns (rejected, reason)."""
    name_lower = brand.name.lower()
    desc_lower = brand.description.lower()

    # Reject known big companies
    for reject in REJECT_COMPANIES:
        if reject in name_lower:
            return True, f"Big player: {brand.name}"

    # Reject by keywords in name/description
    for kw in REJECT_KEYWORDS:
        if kw in name_lower or kw in desc_lower:
            return True, f"Rejected keyword '{kw}'"

    # Reject if estimated revenue > ₹250 Cr
    if brand.est_revenue_cr > 250:
        return True, f"Revenue too high: ₹{brand.est_revenue_cr}Cr"

    return False, ""


# ============================================================
# DEDUPLICATION
# ============================================================

def deduplicate_brands(brands: list[SeedBrand]) -> list[SeedBrand]:
    """Remove duplicates by normalized name and URL."""
    seen_names: set[str] = set()
    seen_urls: set[str] = set()
    unique: list[SeedBrand] = []

    for brand in brands:
        # Normalize name for dedup
        name_key = re.sub(r"[^a-z0-9]", "", brand.name.lower())
        url_key = brand.website.rstrip("/").lower()

        if name_key in seen_names or url_key in seen_urls:
            continue

        seen_names.add(name_key)
        seen_urls.add(url_key)
        unique.append(brand)

    return unique


# ============================================================
# TECHNOLOGY DETECTOR
# ============================================================

@dataclass
class TechDetection:
    platform: str = "unknown"
    platform_confidence: float = 0.0
    has_chatbot: bool = False
    has_whatsapp: bool = False
    has_ai: bool = False
    email_marketing: str = ""
    review_platform: str = ""
    support_tool: str = ""
    analytics: str = ""


PLATFORM_PATTERNS: dict[str, list[str]] = {
    "shopify": [
        r"cdn\.shopify\.com", r"Shopify\.theme", r"myshopify\.com",
        r"shopify-section", r"shopify-payment-button", r"Shopify\.loadFeatures",
        r"Shopify\.analytics", r"x-shopify", r"shopify-domain",
        r"assets\.shopifycdn", r"Shopify\.routes",
    ],
    "shopify_plus": [r"shopify-plus", r"Shopify\.shop"],
    "woocommerce": [
        r"woocommerce", r"wc[-_]ajax", r"wp-content/plugins/woocommerce",
        r"woocommerce-session", r"wc_cart_fragments_params",
    ],
    "magento": [
        r"magento", r"Mage\.", r"skin/frontend",
        r"catalog/product", r"magentocommerce",
    ],
    "custom": [r"next\.js", r"__NEXT_DATA__", r"react", r"nuxt", r"gatsby", r"vue\.js"],
}

CHATBOT_PATTERNS = [r"tidio", r"intercom", r"crisp\.chat", r"tawk\.to", r"zendesk-chat", r"gorgias", r"drift"]
WHATSAPP_PATTERNS = [r"wa\.me", r"api\.whatsapp\.com", r"whatsapp.*widget"]
AI_PATTERNS = [r"ai.*chatbot", r"chatgpt", r"openai.*widget", r"powered by ai"]
EMAIL_MKTG_PATTERNS = {"klaviyo": [r"klaviyo"], "mailchimp": [r"mailchimp"], "sendgrid": [r"sendgrid"]}
REVIEW_PATTERNS = {"judge.me": [r"judge\.me"], "yotpo": [r"yotpo"], "stamped": [r"stamped\.io"]}
SUPPORT_PATTERNS = {"zendesk": [r"zendesk"], "freshdesk": [r"freshdesk"], "intercom": [r"intercom"], "gorgias": [r"gorgias"]}
ANALYTICS_PATTERNS = {"ga4": [r"gtag/js/G-", r"google_tag_manager"], "hotjar": [r"hotjar"]}


def detect_tech(html: str, url: str, headers: dict | None = None) -> TechDetection:
    """Detect technology stack from HTML and HTTP headers."""
    result = TechDetection()
    headers = headers or {}

    # Header-based detection (most reliable)
    header_vals = " ".join(v.lower() for v in headers.values())
    if "shopify" in header_vals:
        result.platform = "shopify"
        result.platform_confidence = 0.9
    elif "x-shopify-stage" in headers or "x-shopify-render" in headers:
        result.platform = "shopify"
        result.platform_confidence = 0.95

    # HTML-based detection (fallback)
    if result.platform == "unknown":
        for platform, patterns in PLATFORM_PATTERNS.items():
            matches = sum(1 for p in patterns if re.search(p, html, re.IGNORECASE))
            if matches > 0:
                result.platform = platform
                result.platform_confidence = min(matches * 0.35, 1.0)
                break

    if result.platform == "unknown" and "myshopify.com" in url:
        result.platform = "shopify"
        result.platform_confidence = 0.95

    # Chatbot, WhatsApp, AI
    result.has_chatbot = any(re.search(p, html, re.IGNORECASE) for p in CHATBOT_PATTERNS)
    result.has_whatsapp = any(re.search(p, html, re.IGNORECASE) for p in WHATSAPP_PATTERNS)
    result.has_ai = any(re.search(p, html, re.IGNORECASE) for p in AI_PATTERNS)

    # Email marketing
    for name, patterns in EMAIL_MKTG_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.email_marketing = name
            break

    # Reviews
    for name, patterns in REVIEW_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.review_platform = name
            break

    # Support
    for name, patterns in SUPPORT_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.support_tool = name
            break

    # Analytics
    for name, patterns in ANALYTICS_PATTERNS.items():
        if any(re.search(p, html, re.IGNORECASE) for p in patterns):
            result.analytics = name
            break

    return result


# ============================================================
# CONTACT SCRAPER — Website + Google Search + LinkedIn
# ============================================================

@dataclass
class ContactInfo:
    emails: list[str] = field(default_factory=list)
    phones: list[str] = field(default_factory=list)
    founder_name: str = ""
    founder_email: str = ""
    founder_linkedin: str = ""
    instagram_url: str = ""
    facebook_url: str = ""
    linkedin_url: str = ""

    @property
    def best_email(self) -> str:
        generic = {"support", "info", "hello", "sales", "care", "contact", "help", "feedback", "noreply", "admin"}
        for e in self.emails:
            prefix = e.split("@")[0].lower()
            if prefix not in generic:
                return e
        return self.emails[0] if self.emails else ""

    @property
    def best_phone(self) -> str:
        return self.phones[0] if self.phones else ""


EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_REGEX = re.compile(r"(?:\+91[\s\-]?)?[6-9]\d{9}")
LINKEDIN_REGEX = re.compile(r"linkedin\.com/(?:company|in)/[a-zA-Z0-9\-]+")
INSTAGRAM_REGEX = re.compile(r"instagram\.com/([a-zA-Z0-9_.]+)")
FACEBOOK_REGEX = re.compile(r"facebook\.com/([a-zA-Z0-9_.]+)")

GENERIC_PREFIXES = {"support", "info", "hello", "sales", "care", "contact", "help", "feedback", "noreply", "admin", "office", "team", "billing", "careers", "jobs", "hr", "enquiry"}
FREE_EMAIL = {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com", "icloud.com", "mail.com"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico", ".bmp"}
INVALID_EMAIL_PATTERNS = {"2x.", ".jpg", ".png", ".webp", ".gif", ".svg", "@2x", "assets", "cdn", "static", "media", "images", "files", "o71740", "base64"}


def _is_valid_email(email: str) -> bool:
    """Validate email quality."""
    email = email.lower().strip()
    if len(email) > 80 or len(email) < 5:
        return False
    domain = email.split("@")[-1] if "@" in email else ""
    if domain in FREE_EMAIL:
        return False
    if any(ext in email for ext in IMAGE_EXTENSIONS):
        return False
    if any(p in email for p in INVALID_EMAIL_PATTERNS):
        return False
    if any(p in email.split("@")[0] for p in GENERIC_PREFIXES):
        return False
    # Must have valid TLD
    if not re.match(r"[a-z0-9.\-]+\.[a-z]{2,}$", domain):
        return False
    return True


def _is_valid_phone(phone: str) -> bool:
    """Validate Indian phone number."""
    digits = re.sub(r"[^0-9]", "", phone)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) != 10:
        return False
    if not digits[0] in "6789":
        return False
    # Reject obviously fake numbers
    if len(set(digits)) <= 2:
        return False
    return True


def _extract_emails(text: str, contact: ContactInfo) -> None:
    for match in EMAIL_REGEX.findall(text):
        email = match.lower().strip()
        if _is_valid_email(email) and email not in contact.emails:
            contact.emails.append(email)


def _extract_phones(text: str, contact: ContactInfo) -> None:
    for match in PHONE_REGEX.findall(text):
        phone = match.strip()
        if _is_valid_phone(phone) and phone not in contact.phones:
            contact.phones.append(phone)


def _extract_social(text: str, contact: ContactInfo) -> None:
    if not contact.linkedin_url:
        m = LINKEDIN_REGEX.search(text)
        if m:
            contact.linkedin_url = "https://" + m.group(0)
    if not contact.instagram_url:
        m = INSTAGRAM_REGEX.search(text)
        if m:
            contact.instagram_url = "https://" + m.group(0)
    if not contact.facebook_url:
        m = FACEBOOK_REGEX.search(text)
        if m:
            contact.facebook_url = "https://" + m.group(0)


def _extract_founder(text: str, contact: ContactInfo) -> None:
    patterns = [
        r"(?:founder|ceo|co[-\s]?founder|managing director)[\s:]+([A-Z][a-z]+ [A-Z][a-z]+)",
        r"(?:Founder|CEO|Co-Founder|Managing Director)[\s:]+([A-Z][a-z]+ [A-Z][a-z]+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, text)
        if m and not contact.founder_name:
            contact.founder_name = m.group(1)


async def scrape_website_contacts(url: str, client: httpx.AsyncClient) -> ContactInfo:
    """Scrape contact info from website pages."""
    contact = ContactInfo()
    base = url.rstrip("/")
    pages = [
        base, base + "/pages/contact", base + "/pages/contact-us",
        base + "/contact", base + "/contact-us",
        base + "/pages/about", base + "/pages/about-us",
        base + "/about", base + "/about-us",
        base + "/pages/faq", base + "/policies/shipping-policy",
    ]

    for page_url in pages:
        try:
            resp = await client.get(page_url, timeout=6.0, follow_redirects=True)
            if resp.status_code == 200:
                text = resp.text[:40000]
                _extract_emails(text, contact)
                _extract_phones(text, contact)
                _extract_social(text, contact)
                _extract_founder(text, contact)
                if (contact.best_email and contact.best_phone) or (contact.best_email and contact.linkedin_url):
                    break
        except Exception:
            continue

    return contact


async def google_search_contacts(brand_name: str, client: httpx.AsyncClient) -> ContactInfo:
    """Search DuckDuckGo for founder phone number and email — single query."""
    contact = ContactInfo()
    try:
        resp = await client.get(
            "https://html.duckduckgo.com/html/",
            params={"q": f'"{brand_name}" founder phone number email India'},
            timeout=10.0,
            follow_redirects=True,
        )
        if resp.status_code == 200:
            text = resp.text
            _extract_phones(text, contact)
            _extract_emails(text, contact)
    except Exception:
        pass
    return contact


async def linkedin_search_contacts(brand_name: str, client: httpx.AsyncClient) -> ContactInfo:
    """Search DuckDuckGo for LinkedIn company page — single query."""
    contact = ContactInfo()
    try:
        resp = await client.get(
            "https://html.duckduckgo.com/html/",
            params={"q": f"site:linkedin.com/company {brand_name} India"},
            timeout=10.0,
            follow_redirects=True,
        )
        if resp.status_code == 200:
            text = resp.text
            m = LINKEDIN_REGEX.search(text)
            if m:
                contact.linkedin_url = "https://" + m.group(0)
            m2 = re.search(r"linkedin\.com/in/([a-zA-Z0-9\-]+)", text)
            if m2:
                contact.founder_linkedin = f"https://linkedin.com/in/{m2.group(1)}"
    except Exception:
        pass
    return contact


async def enrich_contacts(brand_name: str, website: str, client: httpx.AsyncClient) -> ContactInfo:
    """Full contact enrichment: website + Google + LinkedIn."""
    # Step 1: Website scraping
    contact = await scrape_website_contacts(website, client)

    # Step 2: Google search for phone/email if missing
    if not contact.best_phone or not contact.best_email:
        google_contact = await google_search_contacts(brand_name, client)
        if not contact.best_phone and google_contact.best_phone:
            contact.phones = google_contact.phones
        if not contact.best_email and google_contact.best_email:
            contact.emails = google_contact.emails
        if not contact.founder_name and google_contact.founder_name:
            contact.founder_name = google_contact.founder_name

    # Step 3: LinkedIn search for company page + founder
    if not contact.linkedin_url:
        linkedin_contact = await linkedin_search_contacts(brand_name, client)
        if linkedin_contact.linkedin_url:
            contact.linkedin_url = linkedin_contact.linkedin_url
        if linkedin_contact.founder_linkedin:
            contact.founder_linkedin = linkedin_contact.founder_linkedin

    return contact


# ============================================================
# ICP SCORER
# ============================================================

@dataclass
class ICPScore:
    passed: bool
    score: float
    confidence: float
    reason: str
    breakdown: dict[str, float] = field(default_factory=dict)


def score_icp(brand: SeedBrand, tech: TechDetection, contact: ContactInfo) -> ICPScore:
    """Score brand against COMAI ICP (India, D2C, ₹2-250 Cr)."""
    breakdown: dict[str, float] = {}

    # Country (max 25)
    breakdown["country"] = 25.0

    # Revenue (max 25)
    if 2 <= brand.est_revenue_cr <= 250:
        breakdown["revenue"] = 25.0
    elif brand.est_revenue_cr < 2:
        breakdown["revenue"] = 8.0
    else:
        breakdown["revenue"] = 0.0

    # Industry (max 25)
    target_categories = {"Fashion", "Beauty", "Jewellery", "Home Decor", "Electronics", "Baby Products",
                         "Pet Products", "Organic Food", "Health & Wellness", "Footwear", "Bags", "Sports",
                         "Gifts", "Tea/Coffee", "Lifestyle", "Food & Snacks"}
    if brand.category in target_categories:
        breakdown["industry"] = 25.0
    else:
        breakdown["industry"] = 5.0

    # Platform (max 20)
    if tech.platform in ("shopify", "shopify_plus"):
        breakdown["platform"] = 20.0
    elif tech.platform in ("woocommerce", "magento"):
        breakdown["platform"] = 15.0
    elif tech.platform != "unknown":
        breakdown["platform"] = 10.0
    else:
        breakdown["platform"] = 8.0

    # Contact quality (max 15)
    if contact.best_email:
        breakdown["contact"] = 12.0
    elif contact.emails:
        breakdown["contact"] = 8.0
    else:
        breakdown["contact"] = 2.0
    if contact.best_phone:
        breakdown["contact"] = min(breakdown["contact"] + 3, 15.0)

    # Social presence (max 10)
    social = 0
    if contact.instagram_url:
        social += 4
    if contact.facebook_url:
        social += 3
    if contact.linkedin_url:
        social += 3
    breakdown["social"] = min(social, 10.0)

    # Pain signals (max 15)
    pain = 0
    if not tech.has_chatbot:
        pain += 6
    if not tech.has_ai:
        pain += 5
    if tech.has_whatsapp:
        pain += 4
    breakdown["pain"] = min(pain, 15.0)

    total = sum(breakdown.values())
    total = min(total, 100.0)  # Cap at 100
    passed = total >= 50.0
    confidence = min(total / 100.0, 1.0)

    if passed:
        reason = f"PASSED ({total:.0f}/100): {brand.category} on {tech.platform}"
    else:
        reason = f"FAILED ({total:.0f}/100)"

    return ICPScore(passed=passed, score=total, confidence=confidence, reason=reason, breakdown=breakdown)


# ============================================================
# OUTPUT LEAD
# ============================================================

@dataclass
class QualifiedLead:
    company_name: str
    website: str
    category: str
    sub_category: str
    country: str
    city: str
    state: str
    platform: str
    platform_confidence: float
    estimated_revenue: str
    founded_year: int | None
    technology_stack: list[str]
    crm: str
    helpdesk: str
    email_platform: str
    meta_pixel: bool
    google_analytics: str
    whatsapp: bool
    founder_name: str
    business_email: str
    business_phone: str
    linkedin_company: str
    linkedin_decision_maker: str
    instagram: str
    facebook: str
    evidence_urls: list[str]
    last_verified: str
    icp_score: float
    pain_score: float
    intent_score: float
    sales_readiness: float
    close_probability: float
    expected_arr: str
    priority: str
    reason_comai_fits: str
    outreach_angle: str
    recommended_outreach: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "Company Name": self.company_name,
            "Website": self.website,
            "Category": self.category,
            "Sub Category": self.sub_category,
            "Country": self.country,
            "City": self.city,
            "State": self.state,
            "Platform": self.platform,
            "Platform Confidence": round(self.platform_confidence, 2),
            "Revenue Estimate": self.estimated_revenue,
            "Founded Year": self.founded_year or "",
            "Technology Stack": "; ".join(self.technology_stack),
            "CRM": self.crm,
            "Helpdesk": self.helpdesk,
            "Email Platform": self.email_platform,
            "Meta Pixel": "Yes" if self.meta_pixel else "No",
            "Google Analytics": self.google_analytics,
            "WhatsApp": "Yes" if self.whatsapp else "No",
            "Founder": self.founder_name,
            "Business Email": self.business_email,
            "Business Phone": self.business_phone,
            "LinkedIn Company": self.linkedin_company,
            "LinkedIn Decision Maker": self.linkedin_decision_maker,
            "Instagram": self.instagram,
            "Facebook": self.facebook,
            "Evidence URLs": "; ".join(self.evidence_urls),
            "Last Verified": self.last_verified,
            "ICP Score": round(self.icp_score, 1),
            "Pain Score": round(self.pain_score, 1),
            "Intent Score": round(self.intent_score, 1),
            "Sales Readiness": round(self.sales_readiness, 1),
            "Close Probability": f"{self.close_probability:.0%}",
            "Expected ARR": self.expected_arr,
            "Priority": self.priority,
            "Reason COMAI Fits": self.reason_comai_fits,
            "Outreach Angle": self.outreach_angle,
            "Recommended Outreach": self.recommended_outreach,
        }


# ============================================================
# BRAND PROCESSOR
# ============================================================

async def process_brand(
    brand: SeedBrand,
    client: httpx.AsyncClient,
    semaphore: asyncio.Semaphore,
) -> QualifiedLead | None:
    """Process a single brand through the full pipeline."""
    async with semaphore:
        try:
            # Step 1: Fetch homepage
            html = ""
            resp_headers = {}
            for attempt in range(2):
                try:
                    resp = await client.get(brand.website, timeout=8.0, follow_redirects=True)
                    resp_headers = dict(resp.headers)
                    if resp.status_code == 200:
                        html = resp.text[:80000]
                        break
                    elif resp.status_code == 403:
                        alt = brand.website.replace("://www.", "://")
                        try:
                            resp = await client.get(alt, timeout=6.0, follow_redirects=True)
                            resp_headers = dict(resp.headers)
                            if resp.status_code == 200:
                                html = resp.text[:80000]
                                break
                        except Exception:
                            pass
                except Exception:
                    if attempt == 0:
                        await asyncio.sleep(0.5)

            # Step 2: Detect technology
            tech = detect_tech(html, brand.website, resp_headers)

            # Step 3: Enrich contacts (website + Google + LinkedIn)
            contact = await enrich_contacts(brand.name, brand.website, client)

            # Step 4: ICP scoring
            icp = score_icp(brand, tech, contact)

            # Step 5: Build tech stack
            tech_stack = []
            if tech.platform != "unknown":
                tech_stack.append(tech.platform)
            if tech.email_marketing:
                tech_stack.append(tech.email_marketing)
            if tech.review_platform:
                tech_stack.append(tech.review_platform)
            if tech.support_tool:
                tech_stack.append(tech.support_tool)
            if tech.analytics:
                tech_stack.append(tech.analytics)

            # Step 6: Calculate scores
            pain_score = 0.0
            if not tech.has_chatbot:
                pain_score += 40
            if not tech.has_ai:
                pain_score += 30
            if tech.has_whatsapp:
                pain_score += 20
            if not contact.best_email:
                pain_score += 10

            intent_score = min(pain_score + icp.score * 0.3, 100)

            automation_score = 0.0
            if tech.email_marketing:
                automation_score += 25
            if tech.review_platform:
                automation_score += 25
            if tech.support_tool:
                automation_score += 25
            if tech.analytics:
                automation_score += 25

            revenue_score = min(brand.est_revenue_cr / 250 * 100, 100)

            contact_score = 0.0
            if contact.best_email:
                contact_score += 40
            if contact.best_phone:
                contact_score += 30
            if contact.linkedin_url:
                contact_score += 15
            if contact.instagram_url:
                contact_score += 15

            dm_score = 0.0
            if contact.founder_name:
                dm_score += 40
            if contact.founder_linkedin:
                dm_score += 40
            if contact.best_email:
                dm_score += 20

            # Sales readiness
            sales_readiness = (
                icp.score * 0.3 +
                pain_score * 0.25 +
                intent_score * 0.15 +
                contact_score * 0.15 +
                dm_score * 0.15
            )

            # Close probability
            close_prob = min(
                (icp.score / 100 * 0.3) +
                (pain_score / 100 * 0.25) +
                (revenue_score / 100 * 0.2) +
                (contact_score / 100 * 0.15) +
                (dm_score / 100 * 0.1),
                0.95
            )

            # Expected ARR
            arr = max(3, brand.est_revenue_cr * 0.03) * 100000
            arr_str = f"₹{arr / 100000:.1f}L"

            # Priority
            if sales_readiness >= 70 and contact.best_email and contact.best_phone:
                priority = "HOT"
            elif sales_readiness >= 50:
                priority = "WARM"
            else:
                priority = "NURTURE"

            # Reason COMAI fits
            reasons = []
            if not tech.has_chatbot:
                reasons.append("No AI chatbot — needs 24/7 automation")
            if not tech.has_ai:
                reasons.append("No AI tools — high AI readiness gap")
            if tech.has_whatsapp:
                reasons.append("Active WhatsApp — can automate conversations")
            if pain_score >= 50:
                reasons.append("Strong pain signals for COMAI products")
            reason = "; ".join(reasons[:3]) if reasons else "COMAI can improve ecommerce operations"

            # Outreach
            out_angle = f"COMAI can automate {brand.category.lower()} support and boost conversions for {brand.name}"
            if contact.best_email and contact.linkedin_url:
                rec_outreach = "LinkedIn connection + personalized email"
            elif contact.best_email:
                rec_outreach = "Personalized email with case study"
            elif contact.linkedin_url:
                rec_outreach = "LinkedIn outreach"
            else:
                rec_outreach = "Instagram DM + website contact form"

            # Evidence
            evidence = [brand.website]
            if contact.instagram_url:
                evidence.append(contact.instagram_url)
            if contact.linkedin_url:
                evidence.append(contact.linkedin_url)

            now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

            return QualifiedLead(
                company_name=brand.name,
                website=brand.website,
                category=brand.category,
                sub_category=brand.sub_category,
                country="India",
                city=brand.city,
                state=brand.state,
                platform=tech.platform,
                platform_confidence=tech.platform_confidence,
                estimated_revenue=f"₹{max(2, brand.est_revenue_cr // 2)}-{brand.est_revenue_cr} Cr",
                founded_year=brand.founded_year,
                technology_stack=tech_stack,
                crm=tech.support_tool or "None detected",
                helpdesk=tech.support_tool or "None detected",
                email_platform=tech.email_marketing or "None detected",
                meta_pixel="fbq(" in html.lower() if html else False,
                google_analytics=tech.analytics or "None detected",
                whatsapp=tech.has_whatsapp,
                founder_name=contact.founder_name or brand.name + " Team",
                business_email=contact.best_email,
                business_phone=contact.best_phone,
                linkedin_company=contact.linkedin_url or "",
                linkedin_decision_maker=contact.founder_linkedin or "",
                instagram=contact.instagram_url,
                facebook=contact.facebook_url,
                evidence_urls=evidence,
                last_verified=now_str,
                icp_score=icp.score,
                pain_score=pain_score,
                intent_score=intent_score,
                sales_readiness=sales_readiness,
                close_probability=close_prob,
                expected_arr=arr_str,
                priority=priority,
                reason_comai_fits=reason,
                outreach_angle=out_angle,
                recommended_outreach=rec_outreach,
            )

        except Exception as e:
            print(f"  Error processing {brand.name}: {e}")
            return None


# ============================================================
# MAIN PIPELINE
# ============================================================

async def run_pipeline(limit: int = 400, output: str = "comai_leads.xlsx") -> None:
    """Run the full COMAI lead generation pipeline."""
    print("=" * 70)
    print("COMAI LEAD GENERATOR — Mid-Size Indian D2C Brands")
    print("=" * 70)

    # Load and clean seed database
    all_seeds = SEED_BRANDS.copy()
    print(f"\nRaw seeds: {len(all_seeds)}")

    # Deduplicate
    seeds = deduplicate_brands(all_seeds)
    print(f"After dedup: {len(seeds)}")

    # Filter rejects
    rejected = []
    filtered = []
    for seed in seeds:
        is_rej, reason = is_rejected(seed)
        if is_rej:
            rejected.append((seed.name, reason))
        else:
            filtered.append(seed)
    seeds = filtered[:limit]
    print(f"After reject filter: {len(seeds)} brands")

    if rejected:
        print(f"\nRejected {len(rejected)} brands:")
        for name, reason in rejected[:10]:
            print(f"  - {name}: {reason}")

    # Process all brands
    semaphore = asyncio.Semaphore(10)
    results: list[QualifiedLead] = []

    print(f"\nProcessing {len(seeds)} brands...")
    start_time = time.time()

    async with httpx.AsyncClient(
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
        },
        follow_redirects=True,
    ) as client:
        tasks = [process_brand(seed, client, semaphore) for seed in seeds]

        completed = 0
        for coro in asyncio.as_completed(tasks):
            result = await coro
            completed += 1
            if result:
                results.append(result)
            if completed % 25 == 0:
                elapsed = time.time() - start_time
                rate = completed / elapsed if elapsed > 0 else 0
                print(f"  Processed {completed}/{len(seeds)} | Qualified: {len(results)} | {rate:.1f} brands/sec")

    elapsed = time.time() - start_time
    print(f"\nProcessed: {len(seeds)} | Qualified: {len(results)} | Time: {elapsed:.0f}s")

    # Separate by priority
    hot = [r for r in results if r.priority == "HOT"]
    warm = [r for r in results if r.priority == "WARM"]
    nurture = [r for r in results if r.priority == "NURTURE"]

    print(f"\nPriority Breakdown:")
    print(f"  HOT:     {len(hot)}")
    print(f"  WARM:    {len(warm)}")
    print(f"  NURTURE: {len(nurture)}")

    # Sort by sales readiness
    qualified = hot + warm + nurture
    qualified.sort(key=lambda x: x.sales_readiness, reverse=True)

    # Export
    _export_excel(qualified, output)
    _export_summary(qualified, output.replace(".xlsx", "_summary.txt"))

    print(f"\nExported to: {output}")
    print("=" * 70)


# ============================================================
# EXCEL EXPORT
# ============================================================

def _export_excel(leads: list[QualifiedLead], filename: str) -> None:
    """Export leads to formatted Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "COMAI Leads"

    if not leads:
        wb.save(filename)
        return

    headers = list(leads[0].to_dict().keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    priority_fills = {
        "HOT": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "WARM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "NURTURE": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    }

    for row_idx, lead in enumerate(leads, 2):
        data = lead.to_dict()
        priority = data.get("Priority", "")
        row_fill = priority_fills.get(priority)

        for col_idx, header in enumerate(headers, 1):
            value = data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_fill:
                cell.fill = row_fill

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value or ""))
            for row in range(1, min(len(leads) + 2, 50))
        )
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_length + 2, 35)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)


def _export_summary(leads: list[QualifiedLead], filename: str) -> None:
    """Export summary statistics."""
    hot = sum(1 for l in leads if l.priority == "HOT")
    warm = sum(1 for l in leads if l.priority == "WARM")
    nurture = sum(1 for l in leads if l.priority == "NURTURE")

    with_email = sum(1 for l in leads if l.business_email)
    with_phone = sum(1 for l in leads if l.business_phone)
    with_linkedin = sum(1 for l in leads if l.linkedin_company)
    with_instagram = sum(1 for l in leads if l.instagram)

    shopify_count = sum(1 for l in leads if l.platform in ("shopify", "shopify_plus"))
    woocommerce_count = sum(1 for l in leads if l.platform == "woocommerce")
    magento_count = sum(1 for l in leads if l.platform == "magento")

    avg_icp = sum(l.icp_score for l in leads) / len(leads) if leads else 0
    avg_readiness = sum(l.sales_readiness for l in leads) / len(leads) if leads else 0

    categories = {}
    for l in leads:
        categories[l.category] = categories.get(l.category, 0) + 1

    summary = f"""
COMAI LEAD GENERATOR — Summary
================================

Total Leads: {len(leads)}

Priority Breakdown:
  HOT:     {hot}
  WARM:    {warm}
  NURTURE: {nurture}

Contact Availability:
  With Email:    {with_email} ({with_email * 100 // len(leads) if leads else 0}%)
  With Phone:    {with_phone} ({with_phone * 100 // len(leads) if leads else 0}%)
  With LinkedIn: {with_linkedin} ({with_linkedin * 100 // len(leads) if leads else 0}%)
  With Instagram:{with_instagram} ({with_instagram * 100 // len(leads) if leads else 0}%)

Platform Breakdown:
  Shopify/Plus: {shopify_count}
  WooCommerce:  {woocommerce_count}
  Magento:      {magento_count}

Average Scores:
  ICP Score:        {avg_icp:.1f}/100
  Sales Readiness:  {avg_readiness:.1f}/100

Category Breakdown:
"""
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        summary += f"  {cat}: {count}\n"

    summary += f"\nGenerated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n"

    with open(filename, "w") as f:
        f.write(summary)

    print(summary)


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="COMAI Lead Generator — Mid-Size Indian D2C Brands")
    parser.add_argument("--limit", type=int, default=400, help="Max brands to process")
    parser.add_argument("--output", type=str, default="comai_leads.xlsx", help="Output filename")
    args = parser.parse_args()

    asyncio.run(run_pipeline(limit=args.limit, output=args.output))
