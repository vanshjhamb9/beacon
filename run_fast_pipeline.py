"""Fast Recovery Pipeline - Uses seed data + direct collection, no search."""

import asyncio
import logging
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / "packages"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("fast_pipeline")


async def run_fast_pipeline():
    from packages.ecommerce_leads.models import RawEcommerceLead, EnrichedEcommerceLead
    from packages.ecommerce_leads.enrichment.technology_enrichment import TechnologyEnricher
    from packages.ecommerce_leads.enrichment.contact_enrichment import ContactEnricher
    from packages.ecommerce_leads.collectors.social_collector import SocialCollector
    from packages.ecommerce_leads.scoring.ecommerce_score import EcommerceScorer
    from packages.ecommerce_leads.scoring.quality_gate import QualityGate
    from packages.ecommerce_leads.scoring.sales_intelligence import SalesIntelligenceGenerator
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    tech_enricher = TechnologyEnricher(timeout=20.0)
    contact_enricher = ContactEnricher(timeout=20.0)
    social_collector = SocialCollector(timeout=20.0)
    scorer = EcommerceScorer()
    quality_gate = QualityGate()
    sales_intel = SalesIntelligenceGenerator()

    # Use curated seed data of known Indian D2C brands
    seed_brands = [
        ("mamaearth.in", "Mamaearth", "beauty", "Mumbai", "Maharashtra", "Indian D2C beauty brand offering natural personal care products.", 500),
        ("beardo.in", "Beardo", "grooming", "Mumbai", "Maharashtra", "Indian men's grooming brand selling beard care, hair care, and skincare products.", 200),
        ("mcaffeine.com", "mCaffeine", "skincare", "Mumbai", "Maharashtra", "India's first caffeinated personal care brand for millennials.", 300),
        ("sugarcosmetics.com", "Sugar Cosmetics", "cosmetics", "Mumbai", "Maharashtra", "Premium Indian makeup brand with 10,000+ retail touchpoints.", 500),
        ("boat-lifestyle.com", "boAt", "electronics", "New Delhi", "Delhi", "India's #1 audio and wearables brand with 8M+ products sold.", 500),
        ("noise.tech", "Noise", "electronics", "Gurugram", "Haryana", "India's leading connected lifestyle brand for smartwatches and audio.", 400),
        ("fireboltt.com", "Fire-Boltt", "electronics", "New Delhi", "Delhi", "India's fastest growing smartwatch and audio brand.", 350),
        ("thesouledstore.com", "The Souled Store", "fashion", "Mumbai", "Maharashtra", "India's largest pop-culture fashion brand.", 400),
        ("bewakoof.com", "Bewakoof", "fashion", "Mumbai", "Maharashtra", "India's leading D2C fashion brand for youth.", 500),
        ("snitch.co.in", "Snitch", "fashion", "Bangalore", "Karnataka", "Fast-growing Indian menswear D2C brand.", 300),
        ("wowskinscience.com", "WOW Skin Science", "skincare", "Bangalore", "Karnataka", "Indian D2C personal care brand.", 300),
        ("plumgoodness.com", "Plum Goodness", "beauty", "Mumbai", "Maharashtra", "100% vegan and cruelty-free Indian beauty brand.", 250),
        ("bombayshavingcompany.com", "Bombay Shaving Company", "grooming", "New Delhi", "Delhi", "Indian men's grooming D2C brand.", 200),
        ("dermaco.in", "Derma Co", "skincare", "New Delhi", "Delhi", "Indian derma-cosmetics brand by Mamaearth parent.", 250),
        ("juicychemistry.com", "Juicy Chemistry", "organic beauty", "Coimbatore", "Tamil Nadu", "Indian organic and natural skincare brand.", 150),
        ("khadinatural.com", "Khadi Natural", "natural products", "Ahmedabad", "Gujarat", "Indian natural personal care brand.", 200),
        ("rusticart.in", "Rustic Art", "personal care", "Pune", "Maharashtra", "Indian organic personal care brand.", 150),
        ("colorbarcosmetics.com", "Colorbar Cosmetics", "cosmetics", "New Delhi", "Delhi", "Premium Indian cosmetics brand.", 300),
        ("lakmeindia.com", "Lakme", "beauty", "Mumbai", "Maharashtra", "India's leading beauty brand by Unilever.", 500),
        ("theskinstory.in", "The Skin Story", "skincare", "Mumbai", "Maharashtra", "Indian skincare brand for modern consumers.", 150),
        ("deconstruct.in", "Deconstruct", "skincare", "Bangalore", "Karnataka", "Indian skincare brand with active ingredients.", 200),
        ("pepperfry.com", "Pepperfry", "furniture", "Mumbai", "Maharashtra", "India's largest online furniture marketplace.", 800),
        ("urbanladder.com", "Urban Ladder", "furniture", "Bangalore", "Karnataka", "Premium Indian online furniture brand.", 500),
        ("fabindia.com", "Fabindia", "lifestyle", "New Delhi", "Delhi", "India's largest private platform for traditional products.", 600),
        ("firstcry.com", "FirstCry", "kids", "Pune", "Maharashtra", "Asia's largest online store for kids and baby products.", 1000),
        ("hopscotch.in", "Hopscotch", "kids", "Mumbai", "Maharashtra", "India's leading online store for kids' fashion.", 250),
        ("nicobar.com", "Nicobar", "lifestyle", "Mumbai", "Maharashtra", "Modern Indian lifestyle brand for travel and home.", 150),
        ("okhai.org", "Okhai", "handicrafts", "Ahmedabad", "Gujarat", "Handcrafted lifestyle brand empowering rural artisans.", 100),
        ("jaypore.com", "Jaypore", "lifestyle", "New Delhi", "Delhi", "Indian online lifestyle and home decor brand.", 200),
        ("addresshome.com", "Address Home", "home decor", "New Delhi", "Delhi", "Indian premium home decor and lifestyle brand.", 150),
        ("hamleys.com", "Hamleys", "toys", "Mumbai", "Maharashtra", "World's oldest and largest toy retailer with India ops.", 300),
        ("ambraneindia.com", "Ambrane", "electronics", "New Delhi", "Delhi", "Indian consumer electronics brand for power banks.", 150),
        ("pTron.com", "pTron", "electronics", "Hyderabad", "Telangana", "Indian budget audio and mobile accessories brand.", 200),
        ("headsupfortails.com", "Heads Up For Tails", "pets", "New Delhi", "Delhi", "Indian premium pet care brand.", 200),
        ("caratlane.com", "CaratLane", "jewelry", "Chennai", "Tamil Nadu", "India's largest omnichannel jewelry brand.", 500),
        ("bluestone.com", "BlueStone", "jewelry", "Bangalore", "Karnataka", "Leading online jewelry brand in India.", 400),
        ("kalyanjewellers.net", "Kalyan Jewellers", "jewelry", "Kochi", "Kerala", "One of India's largest jewelry retail chains.", 600),
        ("vanheusen.com", "Van Heusen", "fashion", "Mumbai", "Maharashtra", "Premium Indian fashion brand by Aditya Birla.", 400),
        ("parkavenue.in", "Park Avenue", "fashion", "Mumbai", "Maharashtra", "Indian premium menswear brand.", 300),
        ("wooplr.com", "Wooplr", "fashion", "Bangalore", "Karnataka", "Indian fashion discovery platform.", 200),
        ("styched.in", "Styched", "fashion", "Bangalore", "Karnataka", "Indian online fashion brand for youth.", 150),
        ("teabox.com", "Teabox", "tea", "Kolkata", "West Bengal", "Indian premium tea D2C brand.", 100),
        ("rawpressery.com", "Raw Pressery", "beverages", "Mumbai", "Maharashtra", "India's largest cold-pressed juice brand.", 200),
        ("box8.in", "BOX8", "food", "Mumbai", "Maharashtra", "Indian all-in-one meal delivery brand.", 150),
        ("freshmenu.com", "FreshMenu", "food", "Bangalore", "Karnataka", "Indian cloud kitchen and food delivery brand.", 100),
        ("naturebasket.co.in", "Nature's Basket", "grocery", "Mumbai", "Maharashtra", "Indian premium grocery and gourmet store.", 300),
        ("snitch.co.in", "Snitch", "fashion", "Bangalore", "Karnataka", "Fast-growing Indian menswear D2C brand.", 300),
        ("myntra.com", "Myntra", "fashion marketplace", "Bangalore", "Karnataka", "India's leading fashion e-commerce platform.", 1000),
        ("nykaa.com", "Nykaa", "beauty marketplace", "Mumbai", "Maharashtra", "India's leading beauty and wellness e-commerce.", 1000),
        ("purplle.com", "Purplle", "beauty marketplace", "Mumbai", "Maharashtra", "India's second largest online beauty destination.", 800),
    ]

    start_time = time.time()
    all_leads = []
    seen_domains = set()

    logger.info("=" * 60)
    logger.info("FAST RECOVERY PIPELINE - Processing %d brands", len(seed_brands))
    logger.info("=" * 60)

    for i, (domain, name, category, city, state, desc, products) in enumerate(seed_brands):
        if domain in seen_domains:
            continue
        seen_domains.add(domain)

        logger.info("[%d/%d] Processing %s...", i+1, len(seed_brands), name)

        raw = RawEcommerceLead(
            company_name=name,
            website=f"https://{domain}",
            domain=domain,
            platform="shopify",  # Will be verified by detector
            industry=category,
            category=category,
            country="India",
            city=city,
            state=state,
            description=desc,
            product_count=products,
            source="seed_data",
        )

        lead = EnrichedEcommerceLead(raw=raw)

        # Technology detection
        try:
            lead = await tech_enricher.enrich(lead)
        except Exception as e:
            logger.warning("  Tech enrichment failed: %s", e)

        # Contact enrichment
        try:
            lead = await contact_enricher.enrich(lead)
        except Exception as e:
            logger.warning("  Contact enrichment failed: %s", e)

        # Social links
        try:
            social = await social_collector.enrich_social_links(raw.website)
            lead.raw.social_links.update(social)
            lead.instagram_url = social.get("instagram", "")
            lead.facebook_url = social.get("facebook", "")
            lead.linkedin_url = social.get("linkedin", "")
        except Exception as e:
            logger.warning("  Social enrichment failed: %s", e)

        # Score
        lead = scorer.score(lead)

        # Quality gate
        lead = quality_gate.evaluate(lead)

        # Sales intelligence
        lead = sales_intel.generate(lead)

        all_leads.append(lead)
        logger.info("  Platform: %s | Chatbot: %s | WhatsApp: %s | Email: %s | Phone: %s | Score: %.0f | Gate: %s",
                     lead.raw.platform, lead.chatbot_detected, lead.whatsapp_detected,
                     bool(lead.email), bool(lead.phone), lead.comai_score,
                     "PASS" if lead.quality_gate_passed else "FAIL")

    elapsed = time.time() - start_time
    logger.info("=" * 60)
    logger.info("Pipeline completed in %.1f seconds", elapsed)

    # Separate results
    sales_ready = [l for l in all_leads if l.quality_gate_passed]
    needs_enrichment = [l for l in all_leads if not l.quality_gate_passed]

    logger.info("Total leads: %d", len(all_leads))
    logger.info("Sales Ready (passed gate): %d", len(sales_ready))
    logger.info("Needs Enrichment: %d", len(needs_enrichment))

    # Stats
    platform_detected = sum(1 for l in all_leads if l.raw.platform and l.raw.platform != "unknown")
    tech_detected = sum(1 for l in all_leads if any([l.shopify_detected, l.woocommerce_detected, l.magento_detected, l.chatbot_detected, l.whatsapp_detected, l.crm_detected]))
    contacts = sum(1 for l in all_leads if l.email or l.phone)
    pain_points = sum(1 for l in all_leads if l.pain_points)
    decision_makers = sum(1 for l in all_leads if l.founder_name)

    logger.info("")
    logger.info("DETECTION RATES:")
    logger.info("  Platform: %d/%d (%.0f%%)", platform_detected, len(all_leads), platform_detected/len(all_leads)*100)
    logger.info("  Technology: %d/%d (%.0f%%)", tech_detected, len(all_leads), tech_detected/len(all_leads)*100)
    logger.info("  Contacts: %d/%d (%.0f%%)", contacts, len(all_leads), contacts/len(all_leads)*100)
    logger.info("  Pain Points: %d/%d (%.0f%%)", pain_points, len(all_leads), pain_points/len(all_leads)*100)
    logger.info("  Decision Makers: %d/%d (%.0f%%)", decision_makers, len(all_leads), decision_makers/len(all_leads)*100)

    # Export
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["SALES READY LEADS - EMERGENCY RECOVERY"])
    ws_summary.append([])
    ws_summary.append(["Generated", time.strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Total Leads", len(all_leads)])
    ws_summary.append(["Sales Ready", len(sales_ready)])
    ws_summary.append(["Platform Detection", f"{platform_detected/len(all_leads)*100:.0f}%"])
    ws_summary.append(["Technology Detection", f"{tech_detected/len(all_leads)*100:.0f}%"])
    ws_summary.append(["Contact Availability", f"{contacts/len(all_leads)*100:.0f}%"])

    ws_leads = wb.create_sheet("Sales Ready Leads")
    headers = [
        "Company", "Website", "Platform", "Industry", "City",
        "Email", "Phone", "Founder", "Role",
        "Shopify", "Chatbot", "WhatsApp", "CRM",
        "COMAI Score", "Priority", "Confidence", "Quality Gate",
        "Sales Reason", "Pain Points",
        "Call Opener", "Pitch Angle", "Recommended Feature",
    ]
    ws_leads.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws_leads[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    # Export all leads with a Quality Gate column
    for lead in all_leads:
        ws_leads.append([
            lead.raw.company_name,
            lead.raw.website,
            lead.raw.platform,
            lead.raw.industry,
            lead.raw.city,
            lead.email,
            lead.phone,
            lead.founder_name,
            lead.decision_maker_role,
            lead.shopify_detected,
            lead.chatbot_detected,
            lead.whatsapp_detected,
            lead.crm_detected,
            lead.comai_score,
            lead.lead_priority,
            lead.confidence_score,
            "PASS" if lead.quality_gate_passed else "FAIL",
            lead.sales_reason,
            "; ".join(lead.pain_points),
            (lead.call_opener or "")[:200],
            lead.pitch_angle,
            lead.recommended_feature,
        ])

    # Auto-width
    for col in ws_leads.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_leads.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output_path = Path(__file__).parent / "sales_ready_india.xlsx"
    wb.save(str(output_path))
    logger.info("Exported to: %s", output_path)

    return len(sales_ready)


if __name__ == "__main__":
    result = asyncio.run(run_fast_pipeline())
    print(f"\nDone! Generated {result} sales-ready leads.")
