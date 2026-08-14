"""Emergency Recovery Pipeline Runner - Collect 100 sales-ready Indian ecommerce leads."""

import asyncio
import logging
import sys
import os
import time
from pathlib import Path

# Add project paths
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / "apps" / "api"))
sys.path.insert(0, str(Path(__file__).parent / "apps" / "worker"))
sys.path.insert(0, str(Path(__file__).parent / "packages"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("recovery_pipeline")


async def run_pipeline():
    """Run the full recovery pipeline."""
    from packages.ecommerce_leads.services.lead_pipeline import LeadPipeline
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    pipeline = LeadPipeline()
    
    logger.info("=" * 60)
    logger.info("EMERGENCY RECOVERY PIPELINE - Starting")
    logger.info("=" * 60)
    
    start_time = time.time()
    
    # Run discovery with 150 limit to ensure we get 100+ after quality gate
    results = await pipeline.run_discovery(limit=150, country="India")
    
    elapsed = time.time() - start_time
    logger.info("Pipeline completed in %.1f seconds", elapsed)
    logger.info("Total leads generated: %d", len(results))
    
    # Separate quality gate passed vs failed
    sales_ready = [r for r in results if r.get("quality_gate_passed", False)]
    needs_enrichment = [r for r in results if not r.get("quality_gate_passed", False)]
    
    logger.info("Sales Ready (passed quality gate): %d", len(sales_ready))
    logger.info("Needs Enrichment (failed quality gate): %d", len(needs_enrichment))
    
    # Generate statistics
    stats = generate_stats(results, sales_ready, needs_enrichment)
    print_stats(stats)
    
    # Export to Excel
    export_path = Path(__file__).parent / "sales_ready_india.xlsx"
    export_sales_ready_xlsx(sales_ready, stats, export_path)
    logger.info("Exported to: %s", export_path)
    
    # Also export full results for reference
    full_export_path = Path(__file__).parent / "full_pipeline_results.xlsx"
    export_full_results(results, full_export_path)
    logger.info("Full results exported to: %s", full_export_path)
    
    return stats


def generate_stats(all_leads, sales_ready, needs_enrichment):
    """Generate comprehensive statistics."""
    stats = {
        "total_leads": len(all_leads),
        "sales_ready": len(sales_ready),
        "needs_enrichment": len(needs_enrichment),
        "platform_detection": 0,
        "technology_detection": 0,
        "contact_availability": 0,
        "pain_points_generated": 0,
        "sales_intel_generated": 0,
        "avg_comai_score": 0,
        "avg_confidence": 0,
        "platforms": {},
        "industries": {},
        "priorities": {},
    }
    
    if not all_leads:
        return stats
    
    for lead in all_leads:
        # Platform detection
        if lead.get("platform") and lead["platform"] != "unknown":
            stats["platform_detection"] += 1
            platform = lead["platform"]
            stats["platforms"][platform] = stats["platforms"].get(platform, 0) + 1
        
        # Technology detection
        if any([
            lead.get("shopify_detected"),
            lead.get("woocommerce_detected"),
            lead.get("magento_detected"),
            lead.get("chatbot_detected"),
            lead.get("whatsapp_detected"),
            lead.get("crm_detected"),
        ]):
            stats["technology_detection"] += 1
        
        # Contact availability
        if lead.get("email") or lead.get("phone"):
            stats["contact_availability"] += 1
        
        # Pain points
        if lead.get("pain_points") and len(lead["pain_points"]) > 0:
            stats["pain_points_generated"] += 1
        
        # Sales intelligence
        if lead.get("call_opener"):
            stats["sales_intel_generated"] += 1
        
        # Scores
        stats["avg_comai_score"] += lead.get("comai_score", 0)
        stats["avg_confidence"] += lead.get("confidence_score", 0)
        
        # Industry distribution
        industry = lead.get("industry", "unknown")
        stats["industries"][industry] = stats["industries"].get(industry, 0) + 1
        
        # Priority distribution
        priority = lead.get("lead_priority", "unknown")
        stats["priorities"][priority] = stats["priorities"].get(priority, 0) + 1
    
    # Calculate averages
    total = len(all_leads)
    stats["avg_comai_score"] = round(stats["avg_comai_score"] / total, 1) if total else 0
    stats["avg_confidence"] = round(stats["avg_confidence"] / total, 1) if total else 0
    
    # Calculate percentages
    stats["platform_detection_pct"] = round(stats["platform_detection"] / total * 100, 1)
    stats["technology_detection_pct"] = round(stats["technology_detection"] / total * 100, 1)
    stats["contact_availability_pct"] = round(stats["contact_availability"] / total * 100, 1)
    stats["pain_points_pct"] = round(stats["pain_points_generated"] / total * 100, 1)
    stats["sales_intel_pct"] = round(stats["sales_intel_generated"] / total * 100, 1)
    
    return stats


def print_stats(stats):
    """Print formatted statistics."""
    print("\n" + "=" * 60)
    print("PIPELINE RESULTS")
    print("=" * 60)
    print(f"Total Leads:          {stats['total_leads']}")
    print(f"Sales Ready:          {stats['sales_ready']}")
    print(f"Needs Enrichment:     {stats['needs_enrichment']}")
    print()
    print("DETECTION RATES:")
    print(f"  Platform Detection: {stats['platform_detection_pct']}% ({stats['platform_detection']}/{stats['total_leads']})")
    print(f"  Tech Detection:     {stats['technology_detection_pct']}% ({stats['technology_detection']}/{stats['total_leads']})")
    print(f"  Contact Available:  {stats['contact_availability_pct']}% ({stats['contact_availability']}/{stats['total_leads']})")
    print(f"  Pain Points:        {stats['pain_points_pct']}% ({stats['pain_points_generated']}/{stats['total_leads']})")
    print(f"  Sales Intelligence: {stats['sales_intel_pct']}% ({stats['sales_intel_generated']}/{stats['total_leads']})")
    print()
    print("SCORES:")
    print(f"  Avg COMAI Score:    {stats['avg_comai_score']}")
    print(f"  Avg Confidence:     {stats['avg_confidence']}%")
    print()
    print("PLATFORMS:")
    for platform, count in sorted(stats["platforms"].items(), key=lambda x: -x[1]):
        print(f"  {platform}: {count}")
    print()
    print("PRIORITIES:")
    for priority, count in sorted(stats["priorities"].items(), key=lambda x: -x[1]):
        print(f"  {priority}: {count}")
    print()
    print("TOP INDUSTRIES:")
    for industry, count in sorted(stats["industries"].items(), key=lambda x: -x[1])[:10]:
        print(f"  {industry}: {count}")
    print("=" * 60)


def export_sales_ready_xlsx(leads, stats, path):
    """Export sales-ready leads to Excel with formatting."""
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    
    ws_summary.append(["SALES READY LEADS - SUMMARY"])
    ws_summary.append([])
    ws_summary.append(["Generated", time.strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Sales Ready Leads", stats["sales_ready"]])
    ws_summary.append(["Platform Detection", f"{stats['platform_detection_pct']}%"])
    ws_summary.append(["Technology Detection", f"{stats['technology_detection_pct']}%"])
    ws_summary.append(["Contact Availability", f"{stats['contact_availability_pct']}%"])
    ws_summary.append(["Pain Points Generated", f"{stats['pain_points_pct']}%"])
    ws_summary.append(["Sales Intelligence", f"{stats['sales_intel_pct']}%"])
    ws_summary.append(["Avg COMAI Score", stats["avg_comai_score"]])
    ws_summary.append(["Avg Confidence", f"{stats['avg_confidence']}%"])
    
    # Leads sheet
    ws_leads = wb.create_sheet("Sales Ready Leads")
    
    headers = [
        "Company Name", "Website", "Domain", "Platform", "Industry", "Category",
        "Country", "City", "Description", "Product Count",
        "Email", "Phone", "Contact Source", "Contact Confidence",
        "Founder Name", "Decision Maker Role",
        "Shopify", "WooCommerce", "Magento", "Chatbot", "WhatsApp", "CRM",
        "COMAI Score", "Lead Priority", "Sales Reason", "Pain Points",
        "Call Opener", "Pitch Angle", "Recommended Feature",
        "Opportunity Summary", "Confidence Score",
    ]
    
    ws_leads.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    for cell in ws_leads[1]:
        cell.fill = header_fill
        cell.font = header_font_white
    
    for lead in leads:
        row = [
            lead.get("company_name", ""),
            lead.get("website", ""),
            lead.get("domain", ""),
            lead.get("platform", ""),
            lead.get("industry", ""),
            lead.get("category", ""),
            lead.get("country", ""),
            lead.get("city", ""),
            lead.get("description", ""),
            lead.get("product_count", 0),
            lead.get("email", ""),
            lead.get("phone", ""),
            lead.get("contact_source", ""),
            lead.get("contact_confidence", 0),
            lead.get("founder_name", ""),
            lead.get("decision_maker_role", ""),
            lead.get("shopify_detected", False),
            lead.get("woocommerce_detected", False),
            lead.get("magento_detected", False),
            lead.get("chatbot_detected", False),
            lead.get("whatsapp_detected", False),
            lead.get("crm_detected", False),
            lead.get("comai_score", 0),
            lead.get("lead_priority", ""),
            lead.get("sales_reason", ""),
            "; ".join(lead.get("pain_points", [])),
            lead.get("call_opener", ""),
            lead.get("pitch_angle", ""),
            lead.get("recommended_feature", ""),
            lead.get("opportunity_summary", ""),
            lead.get("confidence_score", 0),
        ]
        ws_leads.append(row)
    
    # Auto-width columns
    for column in ws_leads.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(len(str(cell.value)), 50)
            except:
                pass
        ws_leads.column_dimensions[column_letter].width = max_length + 2
    
    wb.save(str(path))


def export_full_results(leads, path):
    """Export all results to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Leads"
    
    headers = [
        "Company Name", "Website", "Platform", "Industry", "Email", "Phone",
        "COMAI Score", "Priority", "Quality Gate", "Confidence",
        "Shopify", "Chatbot", "WhatsApp", "Sales Reason",
    ]
    ws.append(headers)
    
    for lead in leads:
        row = [
            lead.get("company_name", ""),
            lead.get("website", ""),
            lead.get("platform", ""),
            lead.get("industry", ""),
            lead.get("email", ""),
            lead.get("phone", ""),
            lead.get("comai_score", 0),
            lead.get("lead_priority", ""),
            "PASS" if lead.get("quality_gate_passed") else "FAIL",
            lead.get("confidence_score", 0),
            lead.get("shopify_detected", False),
            lead.get("chatbot_detected", False),
            lead.get("whatsapp_detected", False),
            lead.get("sales_reason", ""),
        ]
        ws.append(row)
    
    wb.save(str(path))


if __name__ == "__main__":
    asyncio.run(run_pipeline())
