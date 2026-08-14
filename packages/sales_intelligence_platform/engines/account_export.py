"""Account Export - Export sales-ready accounts."""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from packages.sales_intelligence_platform.models import Account

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    ("Company Name", 30),
    ("Website", 35),
    ("Platform", 15),
    ("Category", 20),
    ("City", 15),
    ("Status", 15),
    ("Primary Decision Maker", 25),
    ("Role", 20),
    ("Email", 35),
    ("Email Verified", 15),
    ("Phone", 18),
    ("Phone Verified", 15),
    ("LinkedIn", 35),
    ("Account Score", 15),
    ("Pain Score", 12),
    ("Growth Score", 12),
    ("Probability to Buy", 15),
    ("Why COMAI", 50),
    ("Recommended Pitch", 50),
    ("Evidence Count", 15),
    ("Completeness %", 15),
]


def export_accounts(accounts: list[Account]) -> bytes:
    """Export accounts to XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Ready Accounts"

    # Headers
    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data
    for row_idx, account in enumerate(accounts, 2):
        primary_dm = account.decision_makers[0] if account.decision_makers else None
        email_ch = next(
            (ch for ch in account.contact_channels if ch.kind.endswith("_email")),
            None,
        )
        phone_ch = next(
            (ch for ch in account.contact_channels if ch.kind in ("business_phone", "founder_phone")),
            None,
        )
        linkedin_ch = next(
            (ch for ch in account.contact_channels if ch.kind == "linkedin_company"),
            None,
        )

        row_data = [
            account.company_name,
            account.website,
            account.platform,
            account.category,
            account.city,
            account.status,
            primary_dm.name if primary_dm else "",
            primary_dm.normalized_role if primary_dm else "",
            email_ch.value if email_ch else "",
            email_ch.verification_level if email_ch else "",
            phone_ch.value if phone_ch else "",
            phone_ch.verification_level if phone_ch else "",
            linkedin_ch.value if linkedin_ch else "",
            round(account.score.total, 1),
            round(account.pain_score, 1),
            round(account.growth_score, 1),
            round(account.probability_to_buy, 1),
            account.status,
            account.status,
            len(account.evidence_records),
            account.health.completeness_pct,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-width
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
