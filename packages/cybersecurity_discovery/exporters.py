"""Write cybersecurity discovery artifacts. Never sends outreach."""

from __future__ import annotations

from pathlib import Path
from typing import Any
import json

from packages.cybersecurity_discovery.pipeline import PipelineResult
from packages.cybersecurity_discovery.schema import CyberOpportunity

EXPORT_FILES = (
    "cyber_sales_ready.json",
    "cyber_sales_ready.xlsx",
    "cyber_needs_research.json",
    "cyber_rejected.json",
    "cyber_evidence_audit.json",
    "cyber_contactability_audit.json",
    "CYBERSECURITY_FINAL_REPORT.md",
)


def write_exports(result: PipelineResult, output_dir: Path) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    written: dict[str, str] = {}

    sales = [o.to_dict() for o in result.sales_ready]
    research = [o.to_dict() for o in result.needs_research]
    rejected = [o.to_dict() for o in result.rejected]

    written["cyber_sales_ready.json"] = _write_json(output_dir / "cyber_sales_ready.json", sales)
    written["cyber_needs_research.json"] = _write_json(output_dir / "cyber_needs_research.json", research)
    written["cyber_rejected.json"] = _write_json(output_dir / "cyber_rejected.json", rejected)
    written["cyber_evidence_audit.json"] = _write_json(
        output_dir / "cyber_evidence_audit.json",
        _evidence_audit(result),
    )
    written["cyber_contactability_audit.json"] = _write_json(
        output_dir / "cyber_contactability_audit.json",
        _contactability_audit(result),
    )
    written["CYBERSECURITY_FINAL_REPORT.md"] = _write_text(
        output_dir / "CYBERSECURITY_FINAL_REPORT.md",
        render_report(result),
    )
    xlsx_path = output_dir / "cyber_sales_ready.xlsx"
    write_xlsx(sales, xlsx_path)
    written["cyber_sales_ready.xlsx"] = str(xlsx_path)
    return written


def render_report(result: PipelineResult) -> str:
    c = result.counters
    lines = [
        "# CYBERSECURITY FINAL REPORT",
        "",
        f"Generated: {result.generated_at}",
        "Lane: CYBER",
        "Outreach: NOT SENT",
        "",
        "## COUNTERS",
        "",
        f"- TOTAL_DISCOVERED: {c.get('TOTAL_DISCOVERED', 0)}",
        f"- BUYING_EVENTS: {c.get('BUYING_EVENTS', 0)}",
        f"- VERIFIED_REQUIREMENTS: {c.get('VERIFIED_REQUIREMENTS', 0)}",
        f"- HOT: {c.get('HOT', 0)}",
        f"- HIGH_INTENT: {c.get('HIGH_INTENT', 0)}",
        f"- CONTACTABLE: {c.get('CONTACTABLE', 0)}",
        f"- SALES_READY: {c.get('SALES_READY', 0)}",
        f"- PARTNER_OPPORTUNITIES: {c.get('PARTNER_OPPORTUNITIES', 0)}",
        f"- NEEDS_RESEARCH: {c.get('NEEDS_RESEARCH', 0)}",
        f"- REJECTED: {c.get('REJECTED', 0)}",
        "",
        "## FUNNEL",
        "",
        "DISCOVERED",
        "↓",
        "BUYING EVENT",
        "↓",
        "SECURITY PROBLEM VERIFIED",
        "↓",
        "BUYER VERIFIED",
        "↓",
        "CURRENT",
        "↓",
        "COMMERCIAL INTENT",
        "↓",
        "SERVICE MATCH",
        "↓",
        "CONTACT VERIFIED",
        "↓",
        "SALES_READY",
        "",
        "Counts:",
    ]
    for stage, count in result.funnel.items():
        lines.append(f"- {stage}: {count}")

    lines.extend(["", "## CTO 15-MINUTE — STRONGEST OPPORTUNITIES", ""])
    ranked = sorted(
        result.sales_ready + result.needs_research,
        key=_rank_key,
    )[:8]
    if not ranked:
        ranked = sorted(
            [o for o in result.rejected if o.buying_event_verified],
            key=_rank_key,
        )[:8]
    if not ranked:
        lines.append("No opportunities passed far enough to recommend a 15-minute contact.")
    for opp in ranked:
        lines.append(f"### {opp.company or opp.title or opp.opportunity_id}")
        lines.append(f"- verdict: {opp.final_verdict}")
        lines.append(f"- type: {opp.opportunity_type}")
        lines.append(f"- intent: {opp.intent_level} | currentness: {opp.currentness}")
        lines.append(f"- buyer: {opp.buyer_name or 'unknown'} ({opp.buyer_role or 'unknown'}) identity={opp.identity_confidence}")
        lines.append(f"- service: {opp.service_match or 'none'} ({opp.service_match_confidence})")
        lines.append(f"- contactability: {opp.contactability} email_status={opp.email_status}")
        lines.append(f"- CTO 15-minute: {opp.cto_15_minute_test} — {opp.cto_decision_reason}")
        if opp.failed_gates:
            lines.append(f"- exact gate failures: {', '.join(opp.failed_gates)}")
        lines.append(f"- source: {opp.source_url}")
        lines.append("")

    lines.extend(["", "## QUALITY RULES HONORED", ""])
    lines.append("- Buying events searched first. Industry or missing security badge is not a lead.")
    lines.append("- No email guessing. No outreach. No vulnerability scanning.")
    lines.append("- Empty SALES_READY is a valid outcome.")
    lines.append("- Partners are excluded from the direct-client sales-ready file.")

    reasons = {}
    for opp in result.rejected:
        key = opp.rejection_reason or "unspecified"
        reasons[key] = reasons.get(key, 0) + 1
    if reasons:
        lines.extend(["", "## REJECTION BREAKDOWN", ""])
        for reason, count in sorted(reasons.items(), key=lambda kv: -kv[1]):
            lines.append(f"- {reason}: {count}")

    lines.append("")
    lines.append(
        "Source note: Reddit listing RSS (old.reddit.com) was used because reddit.com JSON is 403, PullPush is 429, and search RSS is often 429. "
        "Hacker News Algolia was searched for buyer phrases in the last 90 days. Generic keyword hits without a commercial buying event were rejected. "
        "No outreach was sent."
    )
    lines.append("")
    lines.append("STOP. DO NOT SEND OUTREACH.")
    lines.append("")
    return "\n".join(lines)


def write_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        path.write_text("openpyxl not installed\n", encoding="utf-8")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "SALES_READY"
    headers = [
        "opportunity_id",
        "company",
        "country",
        "buyer_name",
        "buyer_role",
        "intent_level",
        "currentness",
        "buying_event",
        "service_match",
        "email",
        "email_status",
        "linkedin_url",
        "contactability",
        "cto_15_minute_test",
        "outreach_priority",
        "source_url",
        "final_verdict",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, item in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=_cell(item.get(header)))
    wb.save(path)


def _evidence_audit(result: PipelineResult) -> dict[str, Any]:
    rows = []
    for opp in result.all_opportunities:
        rows.append(
            {
                "opportunity_id": opp.opportunity_id,
                "final_verdict": opp.final_verdict,
                "buying_event_verified": opp.buying_event_verified,
                "requirement_verified": opp.requirement_verified,
                "requirement_evidence": opp.requirement_evidence,
                "security_problem_evidence": opp.security_problem_evidence,
                "outsourcing_evidence": opp.outsourcing_evidence,
                "currentness_evidence": opp.currentness_evidence,
                "source_url": opp.source_url,
                "source_status": opp.source_status,
                "failed_gates": opp.failed_gates,
                "rejection_reason": opp.rejection_reason,
            }
        )
    return {"generated_at": result.generated_at, "count": len(rows), "items": rows}


def _contactability_audit(result: PipelineResult) -> dict[str, Any]:
    rows = []
    for opp in result.all_opportunities:
        rows.append(
            {
                "opportunity_id": opp.opportunity_id,
                "company": opp.company,
                "buyer_name": opp.buyer_name,
                "buyer_role": opp.buyer_role,
                "identity_confidence": opp.identity_confidence,
                "email": opp.email,
                "email_status": opp.email_status,
                "email_evidence": opp.email_evidence,
                "linkedin_url": opp.linkedin_url,
                "linkedin_status": opp.linkedin_status,
                "contactability": opp.contactability,
                "contactability_evidence": opp.contactability_evidence,
                "guessed_email": False,
            }
        )
    return {"generated_at": result.generated_at, "count": len(rows), "items": rows}


def _rank_key(opp: CyberOpportunity) -> tuple[int, int, int]:
    verdict_rank = {"SALES_READY": 0, "NEEDS_RESEARCH": 1, "REJECT": 2}.get(opp.final_verdict, 3)
    intent_rank = {"HOT": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "UNKNOWN": 4}.get(opp.intent_level, 5)
    fail_count = len(opp.failed_gates)
    return (verdict_rank, intent_rank, fail_count)


def _write_json(path: Path, data: Any) -> str:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    return str(path)


def _write_text(path: Path, text: str) -> str:
    path.write_text(text, encoding="utf-8")
    return str(path)


def _cell(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value
