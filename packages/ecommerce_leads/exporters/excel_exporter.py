"""Excel exporter for ecommerce leads."""

from __future__ import annotations

import io
import logging
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXCEL_COLUMNS = [
    ("Company Name", "company_name", 30),
    ("Website", "website", 40),
    ("Platform", "platform", 15),
    ("Category", "category", 20),
    ("City", "city", 15),
    ("State", "state", 20),
    ("Owner Name", "owner_name", 25),
    ("Founder", "founder_name", 25),
    ("Email", "email", 35),
    ("Phone", "phone", 18),
    ("Instagram", "instagram_url", 35),
    ("LinkedIn", "linkedin_url", 35),
    ("WhatsApp", "whatsapp", 18),
    ("Product Count", "product_count", 15),
    ("Chatbot Available", "chatbot_detected", 18),
    ("COMAI Score", "comai_score", 15),
    ("Lead Priority", "lead_priority", 15),
    ("Reason To Contact", "sales_reason", 50),
    ("Source", "source", 20),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HOT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelExporter:
    """Export ecommerce leads to a professional XLSX file."""

    def export_leads(self, leads: list[dict[str, Any]]) -> bytes:
        """Export leads to an Excel file and return bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "India Ecommerce Leads"

        self._write_headers(ws)
        self._write_data(ws, leads)
        self._auto_width(ws)
        self._add_summary_sheet(wb, leads)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _write_headers(self, ws: Any) -> None:
        """Write styled header row."""
        for col_idx, (header, _, width) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 30
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}1"

    def _write_data(self, ws: Any, leads: list[dict[str, Any]]) -> None:
        """Write lead data rows."""
        for row_idx, lead in enumerate(leads, 2):
            for col_idx, (_, field, _) in enumerate(EXCEL_COLUMNS, 1):
                value = lead.get(field, "")
                if isinstance(value, bool):
                    value = "Yes" if value else "No"
                elif isinstance(value, float):
                    value = round(value, 1)
                elif isinstance(value, list):
                    value = ", ".join(str(v) for v in value)

                cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            priority = lead.get("lead_priority", "LOW")
            score = lead.get("comai_score", 0)
            score_cell = ws.cell(row=row_idx, column=EXCEL_COLUMNS.index(
                ("COMAI Score", "comai_score", 15)
            ) + 1)
            priority_cell = ws.cell(row=row_idx, column=EXCEL_COLUMNS.index(
                ("Lead Priority", "lead_priority", 15)
            ) + 1)

            if priority == "HOT":
                score_cell.fill = HOT_FILL
                priority_cell.fill = HOT_FILL
            elif priority == "WARM":
                score_cell.fill = WARM_FILL
                priority_cell.fill = WARM_FILL
            else:
                score_cell.fill = LOW_FILL
                priority_cell.fill = LOW_FILL

    def _auto_width(self, ws: Any) -> None:
        """Auto-adjust column widths."""
        for col_idx, (_, _, default_width) in enumerate(EXCEL_COLUMNS, 1):
            max_len = default_width
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    def _add_summary_sheet(self, wb: Any, leads: list[dict[str, Any]]) -> None:
        """Add a summary sheet."""
        summary = wb.create_sheet("Summary", 0)

        summary_data = [
            ("Metric", "Value"),
            ("Total Leads", len(leads)),
            ("HOT Leads", sum(1 for l in leads if l.get("lead_priority") == "HOT")),
            ("WARM Leads", sum(1 for l in leads if l.get("lead_priority") == "WARM")),
            ("LOW Leads", sum(1 for l in leads if l.get("lead_priority") == "LOW")),
            ("Average COMAI Score", round(
                sum(l.get("comai_score", 0) for l in leads) / max(len(leads), 1), 1
            )),
            ("Shopify Stores", sum(1 for l in leads if l.get("shopify_detected"))),
            ("WooCommerce Stores", sum(1 for l in leads if l.get("woocommerce_detected"))),
            ("With Email", sum(1 for l in leads if l.get("email"))),
            ("With Phone", sum(1 for l in leads if l.get("phone"))),
            ("With WhatsApp", sum(1 for l in leads if l.get("whatsapp_detected"))),
            ("Generated At", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 1):
            cell_a = summary.cell(row=row_idx, column=1, value=metric)
            cell_b = summary.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                cell_a.fill = HEADER_FILL
                cell_a.font = HEADER_FONT
                cell_b.fill = HEADER_FILL
                cell_b.font = HEADER_FONT
            cell_a.border = THIN_BORDER
            cell_b.border = THIN_BORDER

        summary.column_dimensions["A"].width = 25
        summary.column_dimensions["B"].width = 20

    def export_to_file(self, leads: list[dict[str, Any]], filepath: str) -> str:
        """Export leads to a file and return the filepath."""
        data = self.export_leads(leads)
        with open(filepath, "wb") as f:
            f.write(data)
        return filepath
